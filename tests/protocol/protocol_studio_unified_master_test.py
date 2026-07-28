from __future__ import annotations

from pathlib import Path

from _test_support import add_repo_to_import_path, configure_process_runtime

add_repo_to_import_path()
TEST_RUNTIME = configure_process_runtime("mcgs-unified-master")
OUTPUT_ROOT = TEST_RUNTIME / "artifacts"

from openpyxl import load_workbook

from mvp_generator.generator import ProtocolGenerator
from mvp_generator.excel_renderer import ClassicCombinedRenderer
from mvp_generator.library import TemplateLibrary
from mvp_generator.validate_rendered_workbook import validate_model_and_workbook
from mvp_generator.validator import ConfigError
from protocol_studio.alarm_codegen import generate_alarm_code_from_workbook
from protocol_studio.app import normalize_config, render_excel
from protocol_studio.program_upload import write_program_upload_csv

def build_config(mode: str) -> dict:
    config = {
        "workflow_version": "unified_protocol_v1",
        "project": {
            "name": f"统一母版-{mode}",
            "code": f"MASTER-{mode}",
            "protocol_title": "动环通讯协议",
        },
        "protocol_layout": {
            "measurement_layout_mode": mode,
        },
        "routes": {
            "A": {
                "start_boxes": {"count": 1, "instance_names": ["S1"]},
                "plug_boxes": {
                    "board_number_start": 101,
                    "sequence": [
                        {"type_code": "3P*2", "count": 1, "layout_pattern": "1+1"},
                    ],
                },
                "branch_modules": {
                    "module_count": 2,
                    "module_number_start": 1,
                    "output_number_start": 1,
                    "branch_device_number_start": 101,
                    "branches_per_module": 2,
                },
            },
            "B": {"copy_from_A": True},
        },
        "extensions": {
            "single_cabinet": {"enabled": True, "cabinet_count": 3},
            "repeater": {
                "enabled": True,
                "A_count": 1,
                "B_count": 1,
                "alias": "中继器",
            },
            "alarm_state_word": {
                "enabled": True,
                "base_address": 9200 if mode == "by_branch" else 6000,
                "word_mode": "16bit",
            },
        },
        "profiles": {},
    }
    return config


def find_row_by_value(ws, column: int, value: str) -> int:
    for row_no in range(1, ws.max_row + 1):
        if ws.cell(row=row_no, column=column).value == value:
            return row_no
    raise AssertionError(f"未找到 {value}")


def merged_range_containing(ws, row_no: int, column_no: int):
    return next(
        (
            cell_range
            for cell_range in ws.merged_cells.ranges
            if cell_range.min_row <= row_no <= cell_range.max_row
            and cell_range.min_col <= column_no <= cell_range.max_col
        ),
        None,
    )


def assert_unified_main_contract(main) -> None:
    assert main["I10"].value == "分路"
    assert main["J10"].value == "设备"
    assert main.column_dimensions["O"].hidden is True
    assert "K11:N11" in {str(item) for item in main.merged_cells.ranges}
    assert main["K11"].value == "断路器状态字说明"
    assert "变量类型" not in [main.cell(row=10, column=column).value for column in range(1, main.max_column + 1)]
    assert "通道名称" not in [main.cell(row=10, column=column).value for column in range(1, main.max_column + 1)]
    note = str(main["A8"].value or "")
    assert "IEEE 754" in note
    assert "低地址寄存器存放高 16 位" in note
    assert "1061" not in note
    assert "小寄存器" not in note
    assert "高字节" not in note


def test_branch_mode() -> None:
    library = TemplateLibrary.load()
    config = normalize_config(build_config("by_branch"), library)
    assert config["profiles"]["start_box_template_id"] == "start_box_extended_load_unbalance_reactive"
    assert config["profiles"]["plug_branch_template_id"] == "plug_branch_extended_load_reactive"
    assert config["profiles"]["single_cabinet_template_id"] == "single_cabinet_liquidcool_ia_pa_ea_ka"

    model = ProtocolGenerator(library).generate(config)
    assert model["protocol_layout"]["measurement_layout_mode"] == "by_branch"
    assert len(model["routes"][0]["start_boxes"][0]["points"]) == 46
    for route_model, expected_codes in zip(
        model["routes"],
        (["101", "102", "103", "104"], ["201", "202", "203", "204"]),
        strict=True,
    ):
        assert len(route_model["physical_plug_boxes"]) == 2
        assert all(item["entity_kind"] == "monitor_module" for item in route_model["physical_plug_boxes"])
        branches = [
            branch
            for physical_box in route_model["physical_plug_boxes"]
            for board in physical_box["boards"]
            for branch in board["branches"]
        ]
        assert [item["variable_device_code"] for item in branches] == expected_codes
        assert [item["output_no"] for item in branches] == [1, 2, 3, 4]
        assert all(len(item["points"]) == 41 for item in branches)

    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    excel_path = OUTPUT_ROOT / "unified-master-by-branch.xlsx"
    render_excel(model, excel_path)
    validate_model_and_workbook(model, excel_path)

    workbook = load_workbook(excel_path, data_only=True)
    try:
        assert workbook.sheetnames == ["始端箱和插接箱", "中继器", "报警状态"]
        main = workbook["始端箱和插接箱"]
        assert main["A1"].value == "项目名称：统一母版-by_branch"
        assert "MASTER-by_branch" in str(main["A2"].value)
        assert main.row_dimensions[1].hidden is True
        assert main.row_dimensions[2].hidden is True
        assert main["A10"].value == "通道号"
        assert all(sheet.freeze_panes is None for sheet in workbook.worksheets)
        assert_unified_main_contract(main)
        start_state_row = find_row_by_value(main, 2, "StateS1")
        assert main.cell(row=start_state_row, column=9).value == "A路始端箱S1"
        assert main.cell(row=start_state_row, column=7).font.bold is True
        assert "按监控模块组织" in str(main["A5"].value or "")
        assert "按输出分路组织" not in str(main["A5"].value or "")
        assert "A路：始端箱" not in str(main["A5"].value or "")

        row_a = find_row_by_value(main, 2, "StateC101")
        row_b = find_row_by_value(main, 2, "StateC201")
        assert main.cell(row=row_a, column=9).value == "A路输出分路1"
        assert main.cell(row=row_a, column=10).value == "A路1#监控模块"
        assert main.cell(row=row_b, column=9).value == "B路输出分路1"
        assert main.cell(row=row_b, column=10).value == "B路1#监控模块"
        assert main.cell(row=row_a, column=2).fill.fgColor.rgb != main.cell(row=row_b, column=2).fill.fgColor.rgb

        row_ia = find_row_by_value(main, 2, "IA01")
        row_pa = find_row_by_value(main, 2, "PA01")
        row_ea = find_row_by_value(main, 2, "EA01")
        row_ka = find_row_by_value(main, 2, "KA01")
        cabinet_colors = {
            main.cell(row=row_ia, column=2).fill.fgColor.rgb,
            main.cell(row=row_pa, column=2).fill.fgColor.rgb,
            main.cell(row=row_ea, column=2).fill.fgColor.rgb,
            main.cell(row=row_ka, column=2).fill.fgColor.rgb,
        }
        assert len(cabinet_colors) == 4
        assert main.cell(row=row_ia, column=6).value == 8200
        assert main.cell(row=row_pa, column=6).value == 8400
        assert main.cell(row=row_ea, column=6).value == 8600
        assert main.cell(row=row_ka, column=6).value == 8800
        assert "总电流" in str(main.cell(row=row_ia, column=9).value or "")
        assert "1#机柜" in str(main.cell(row=row_ia, column=9).value or "")
        assert "总功率" in str(main.cell(row=row_pa, column=9).value or "")
        assert "总电能" in str(main.cell(row=row_ea, column=9).value or "")
        assert "状态字" in str(main.cell(row=row_ka, column=9).value or "")

        alarm = workbook["报警状态"]
        descriptions = [
            str(alarm.cell(row=row, column=7).value or "")
            for row in range(8, alarm.max_row + 1)
        ]
        assert descriptions[0].startswith("A路始端箱")
        first_common = next(index for index, item in enumerate(descriptions) if "设备通讯异常" in item)
        assert first_common > 0
        bit_text = "\n".join(
            str(alarm.cell(row=row, column=9).value or "")
            for row in range(8, alarm.max_row + 1)
        )
        assert "A路1#监控模块通讯异常" in bit_text
        assert "A路输出分路1-A相电压超上限" in bit_text
    finally:
        workbook.close()

    alarm_code = generate_alarm_code_from_workbook(excel_path)
    assert "Comm_EC101" in alarm_code
    assert "Ua101" in alarm_code
    csv_result = write_program_upload_csv(
        excel_path,
        OUTPUT_ROOT / "unified-master-by-branch.csv",
    )
    assert csv_result["point_count"] > 0


def test_default_protocol_title() -> None:
    library = TemplateLibrary.load()
    config_payload = build_config("by_plug_box")
    config_payload["project"].pop("protocol_title")
    config = normalize_config(config_payload, library)
    assert config["protocol_title"] == "上位机通讯协议"
    model = ProtocolGenerator(library).generate(config)
    assert model["project"]["protocol_title"] == "上位机通讯协议"

    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    excel_path = OUTPUT_ROOT / "unified-master-default-title.xlsx"
    render_excel(model, excel_path)
    workbook = load_workbook(excel_path, data_only=True)
    try:
        main = workbook["始端箱和插接箱"]
        assert "协议标题：上位机通讯协议" in str(main["A2"].value or "")
        assert main.row_dimensions[1].hidden is True
        assert main.row_dimensions[2].hidden is True
    finally:
        workbook.close()


def test_synthetic_segmented_addresses() -> None:
    config_payload = build_config("by_branch")
    config_payload["routes"]["A"]["branch_modules"]["module_count"] = 20
    config_payload["extensions"]["single_cabinet"]["cabinet_count"] = 38
    config_payload["extensions"]["repeater"]["A_count"] = 8
    config_payload["extensions"]["repeater"]["B_count"] = 8

    library = TemplateLibrary.load()
    config = normalize_config(config_payload, library)
    model = ProtocolGenerator(library).generate(config)

    branch_addresses: dict[tuple[str, int], int] = {}
    for route_model in model["routes"]:
        for physical_box in route_model["physical_plug_boxes"]:
            for board in physical_box["boards"]:
                for branch in board["branches"]:
                    branch_addresses[(route_model["route"], branch["output_no"])] = branch["points"][0]["address"]

    assert branch_addresses[("A", 1)] == 2000
    assert branch_addresses[("A", 38)] == 4997
    assert branch_addresses[("A", 39)] == 9500
    assert branch_addresses[("A", 40)] == 9581
    assert branch_addresses[("B", 1)] == 5078
    assert branch_addresses[("B", 38)] == 8075
    assert branch_addresses[("B", 39)] == 9662
    assert branch_addresses[("B", 40)] == 9743
    assert model["address_summary"]["downstream_segments"]["primary_next_address"] == 8156
    assert model["address_summary"]["downstream_segments"]["extension_next_address"] == 9824
    alarm_row_count = len(ClassicCombinedRenderer(model)._build_alarm_rows())
    assert alarm_row_count == 148, alarm_row_count

    cabinet_addresses = {item["var_name"]: item["address"] for item in model["single_cabinet_rows"]}
    assert cabinet_addresses["IA01"] == 8200
    assert cabinet_addresses["PA01"] == 8400
    assert cabinet_addresses["EA01"] == 8600
    assert cabinet_addresses["KA01"] == 8800


def test_plug_box_mode() -> None:
    library = TemplateLibrary.load()
    config_payload = build_config("by_plug_box")
    config_payload["routes"]["A"]["plug_boxes"]["sequence"].append(
        {"type_code": "1P*3", "count": 1, "layout_pattern": "1"}
    )
    config = normalize_config(config_payload, library)
    model = ProtocolGenerator(library).generate(config)
    excel_path = OUTPUT_ROOT / "unified-master-by-plug-box.xlsx"
    render_excel(model, excel_path)
    validate_model_and_workbook(model, excel_path)

    workbook = load_workbook(excel_path, data_only=True)
    try:
        assert workbook.sheetnames == ["始端箱和插接箱", "中继器", "报警状态"]
        main = workbook["始端箱和插接箱"]
        assert_unified_main_contract(main)
        assert "按插接箱组织" in str(main["A5"].value or "")
        row_101 = find_row_by_value(main, 2, "StateC101")
        row_102 = find_row_by_value(main, 2, "StateC102")
        assert main.cell(row=row_101, column=9).value == "分路1"
        assert main.cell(row=row_102, column=9).value == "分路2"
        device_range = merged_range_containing(main, row_101, 10)
        assert device_range is not None
        assert device_range.min_row <= row_102 <= device_range.max_row
        assert main.cell(row=device_range.min_row, column=10).value == "A路插接箱101"

        row_shared = find_row_by_value(main, 2, "StateC103")
        assert main.cell(row=row_shared, column=9).value == "分路1–3（共享测量点集）"
        shared_device_range = merged_range_containing(main, row_shared, 10)
        assert shared_device_range is not None
        assert main.cell(row=shared_device_range.min_row, column=10).value == "A路插接箱102"
        assert "IA01" in [main.cell(row=row, column=2).value for row in range(11, main.max_row + 1)]
    finally:
        workbook.close()

    alarm_code = generate_alarm_code_from_workbook(excel_path)
    assert "Comm_EC101" in alarm_code


def test_address_overlap_rejected() -> None:
    config_payload = build_config("by_branch")
    config_payload["extensions"]["repeater"]["base_address"] = 9000
    config_payload["extensions"]["alarm_state_word"]["base_address"] = 9000

    library = TemplateLibrary.load()
    config = normalize_config(config_payload, library)
    try:
        ProtocolGenerator(library).generate(config)
    except ConfigError as exc:
        message = str(exc)
        assert "寄存器地址冲突" in message
        assert "9000" in message
    else:
        raise AssertionError("重叠的中继和报警地址必须硬失败")


def main() -> int:
    test_default_protocol_title()
    test_branch_mode()
    test_plug_box_mode()
    test_synthetic_segmented_addresses()
    test_address_overlap_rejected()
    print("protocol_studio_unified_master_ok")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
