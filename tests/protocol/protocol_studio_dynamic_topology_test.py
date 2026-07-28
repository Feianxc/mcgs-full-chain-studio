from __future__ import annotations

import csv
import re
import time
from collections import Counter
from pathlib import Path

from _test_support import (
    add_repo_to_import_path,
    configure_process_runtime,
    generated_artifact_path,
)

add_repo_to_import_path()
TEST_RUNTIME = configure_process_runtime("mcgs-dynamic-topology")
OUTPUT_ROOT = TEST_RUNTIME / "artifacts"

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from mvp_generator.excel_renderer import ClassicCombinedRenderer
from mvp_generator.generator import ProtocolGenerator
from mvp_generator.library import TemplateLibrary
from mvp_generator.validate_rendered_workbook import validate_model_and_workbook
from protocol_studio import app as app_module
from protocol_studio.alarm_codegen import (
    CANONICAL_TOPOLOGY_METADATA_PREFIX,
    WorkbookContext,
    collect_combined_plug_map,
    generate_alarm_code_from_workbook,
)
from protocol_studio.app import normalize_config, render_excel
from protocol_studio.program_upload import (
    extract_protocol_point_rows,
    write_program_upload_csv_from_config,
)

def build_dynamic_config() -> dict:
    return {
        "workflow_version": "unified_protocol_v1",
        "project": {
            "name": "动态监控模块拓扑回归",
            "code": "DYNAMIC-TOPOLOGY-AB",
            "protocol_title": "动环通讯协议",
        },
        "protocol_layout": {
            "measurement_layout_mode": "by_branch",
            "main_base_address": 1000,
            "downstream_base_address": 2000,
            "downstream_primary_outputs_per_route": 38,
            "downstream_extension_base_address": 9500,
        },
        "routes": {
            "A": {
                "start_boxes": {"count": 1, "instance_names": ["S1"]},
                "branch_modules": {
                    "module_sequence": [
                        {"type_code": "3P*1", "count": 1, "layout_pattern": "1"},
                        {"type_code": "3P*2", "count": 1, "layout_pattern": "2"},
                    ],
                    "module_number_start": 1,
                    "output_number_start": 1,
                    "branch_device_number_start": 101,
                    "variable_numbering_mode": "per_output_contiguous",
                },
            },
            "B": {
                "start_boxes": {"count": 1, "instance_names": ["S2"]},
                "branch_modules": {
                    "module_sequence": [
                        {"type_code": "3P*3", "count": 1, "layout_pattern": "2+1"},
                        {"type_code": "3P*4", "count": 1, "layout_pattern": "2+2"},
                    ],
                    "module_number_start": 1,
                    "output_number_start": 1,
                    "branch_device_number_start": 201,
                    "variable_numbering_mode": "per_output_contiguous",
                },
            },
        },
        "extensions": {
            "single_cabinet": {"enabled": False, "cabinet_count": 0},
            "repeater": {"enabled": False, "A_count": 0, "B_count": 0},
            "alarm_state_word": {
                "enabled": True,
                "base_address": 6000,
                "word_mode": "16bit",
            },
        },
        "profiles": {},
    }


def build_legacy_segmented_config() -> dict:
    return {
        "workflow_version": "unified_protocol_v1",
        "project": {
            "name": "合成分段地址兼容回归",
            "code": "SYNTH-20X2-LEGACY",
            "protocol_title": "动环通讯协议",
        },
        "protocol_layout": {
            "measurement_layout_mode": "by_branch",
            "main_base_address": 1000,
            "downstream_base_address": 2000,
            "downstream_primary_outputs_per_route": 38,
            "downstream_extension_base_address": 9500,
        },
        "routes": {
            "A": {
                "start_boxes": {"count": 1, "instance_names": ["S1"]},
                "branch_modules": {
                    "module_count": 20,
                    "module_number_start": 1,
                    "output_number_start": 1,
                    "branch_device_number_start": 101,
                    "branches_per_module": 2,
                    "variable_numbering_mode": "per_output_contiguous",
                },
            },
            "B": {"copy_from_A": True},
        },
        "extensions": {
            "single_cabinet": {
                "enabled": True,
                "cabinet_count": 38,
                "base_address": 8200,
                "metric_base_addresses": {
                    "IA": 8200,
                    "PA": 8400,
                    "EA": 8600,
                    "KA": 8800,
                },
            },
            "repeater": {
                "enabled": True,
                "A_count": 8,
                "B_count": 8,
                "alias": "中继单元",
                "base_address": 9000,
            },
            "alarm_state_word": {
                "enabled": True,
                "base_address": 9200,
                "word_mode": "16bit",
            },
        },
        "profiles": {},
    }


def flatten_branches(route_model: dict) -> list[dict]:
    return [
        branch
        for module in route_model["physical_plug_boxes"]
        for board in module["boards"]
        for branch in board["branches"]
    ]


def assert_dynamic_canonical(model: dict) -> None:
    routes = {route["route"]: route for route in model["routes"]}
    expected = {
        "A": [
            ("3P*1", "1", [1], ["101"]),
            ("3P*2", "2", [1, 2], ["102", "102_2"]),
        ],
        "B": [
            ("3P*3", "2+1", [1, 2, 3], ["201", "201_2", "202"]),
            ("3P*4", "2+2", [1, 2, 3, 4], ["203", "203_2", "204", "204_2"]),
        ],
    }

    all_var_names: list[str] = []
    for route_name, expected_modules in expected.items():
        route_model = routes[route_name]
        modules = route_model["physical_plug_boxes"]
        assert len(modules) == len(expected_modules)
        expected_output_no = 1
        for module_index, (module, module_expectation) in enumerate(
            zip(modules, expected_modules, strict=True),
            start=1,
        ):
            type_code, layout_pattern, local_numbers, device_codes = module_expectation
            assert module["entity_kind"] == "monitor_module"
            assert module["module_id"] == f"{route_name}-M{module_index:03d}"
            assert module["module_no"] == module_index
            assert module["communication_alarm_slot"] == module_index
            assert module["communication_variable_device_code"] == device_codes[0]
            assert module["module_sequence_source"] == "explicit_module_sequence"
            assert module["variable_numbering_mode"] == "per_board_suffix"
            assert module["type_code"] == type_code
            assert module["layout_pattern"] == layout_pattern

            branches = [
                branch
                for board in module["boards"]
                for branch in board["branches"]
            ]
            assert module["module_branch_count"] == len(branches) == len(local_numbers)
            assert [branch["module_local_branch_no"] for branch in branches] == local_numbers
            assert [branch["variable_device_code"] for branch in branches] == device_codes
            assert [branch["output_no"] for branch in branches] == list(
                range(expected_output_no, expected_output_no + len(branches))
            )
            expected_output_no += len(branches)
            for branch in branches:
                assert branch["module_id"] == module["module_id"]
                assert branch["module_no"] == module["module_no"]
                assert branch["communication_alarm_slot"] == module["communication_alarm_slot"]
                assert (
                    branch["communication_variable_device_code"]
                    == module["communication_variable_device_code"]
                )
                assert branch["variable_numbering_mode"] == "per_board_suffix"
                all_var_names.extend(point["var_name"] for point in branch["points"])

    duplicates = [name for name, count in Counter(all_var_names).items() if count > 1]
    assert not duplicates, duplicates[:10]
    assert [branch["output_no"] for branch in flatten_branches(routes["A"])] == [1, 2, 3]
    assert [branch["output_no"] for branch in flatten_branches(routes["B"])] == list(range(1, 8))


def canonical_main_rows(model: dict) -> list[tuple[str, int]]:
    rows: list[tuple[str, int]] = []
    for route in model["routes"]:
        for start_box in route["start_boxes"]:
            rows.extend((point["var_name"], point["address"]) for point in start_box["points"])
    for route in model["routes"]:
        for branch in flatten_branches(route):
            rows.extend((point["var_name"], point["address"]) for point in branch["points"])
    return rows


def read_csv_rows(path: Path) -> list[list[str]]:
    raw = path.read_bytes()
    for encoding in ("gb18030", "utf-8-sig", "utf-8"):
        try:
            return list(csv.reader(raw.decode(encoding).splitlines()))
        except UnicodeDecodeError:
            continue
    raise AssertionError(f"无法解码 CSV：{path}")


def test_explicit_and_legacy_workbook_mapping() -> None:
    explicit_book = Workbook()
    explicit_sheet = explicit_book.active
    explicit_sheet.title = "始端箱和插接箱"
    explicit_sheet.append(
        [
            "StateC101",
            "A路输出分路5（A路7#监控模块·模块内分路3）",
            (
                f"{CANONICAL_TOPOLOGY_METADATA_PREFIX}|route=A|output_no=5|"
                "module_no=7|module_local_branch_no=3|communication_alarm_slot=9|"
                "communication_variable_device_code=777"
            ),
        ]
    )
    explicit_context = WorkbookContext(repeater_comm_prefix="Comm_EZ")
    collect_combined_plug_map(explicit_context, explicit_sheet)
    assert explicit_context.output_branch_map[("A", 5)] == "101"
    assert explicit_context.extended_module_branch_map[("A", 7, 3)] == "101"
    assert explicit_context.extended_module_comm_map[("A", 9)] == "777"

    legacy_book = Workbook()
    legacy_sheet = legacy_book.active
    legacy_sheet.title = "始端箱和插接箱"
    for output_no, device_code in enumerate((101, 102, 103, 104), start=1):
        legacy_sheet.append([f"StateC{device_code}", f"A路输出分路{output_no}"])
    legacy_context = WorkbookContext(repeater_comm_prefix="Comm_EZ")
    collect_combined_plug_map(legacy_context, legacy_sheet)
    assert legacy_context.extended_module_branch_map[("A", 1, 1)] == "101"
    assert legacy_context.extended_module_branch_map[("A", 1, 2)] == "102"
    assert legacy_context.extended_module_branch_map[("A", 2, 1)] == "103"
    assert legacy_context.extended_module_branch_map[("A", 2, 2)] == "104"
    assert legacy_context.extended_module_comm_map[("A", 1)] == "101"
    assert legacy_context.extended_module_comm_map[("A", 2)] == "103"


def test_dynamic_api_three_file_delivery() -> float:
    client = TestClient(app_module.app)
    started_at = time.perf_counter()
    response = client.post("/api/generate", json={"config": build_dynamic_config()})
    elapsed_seconds = time.perf_counter() - started_at
    assert response.status_code == 200, response.text
    payload = response.json()

    assert isinstance(payload["delivery_status"], dict)
    assert payload["delivery_status"]["status"] == "deliverable"
    assert payload["delivery_bundle"]["status"] == "complete"
    assert payload["alarm_codegen"]["status"] == "generated"
    assert payload["program_upload"]["status"] == "generated"
    assert isinstance(payload["program_upload"]["point_count"], int)
    assert elapsed_seconds < 30, elapsed_seconds

    model = payload["canonical"]
    assert_dynamic_canonical(model)
    excel_path = generated_artifact_path(app_module, payload, "excel_path")
    alarm_code_path = generated_artifact_path(app_module, payload, "alarm_code_path")
    csv_path = generated_artifact_path(app_module, payload, "program_upload_path")
    for path in (excel_path, alarm_code_path, csv_path):
        assert path.exists() and path.stat().st_size > 0, path

    validate_model_and_workbook(model, excel_path)
    workbook = load_workbook(excel_path, data_only=True)
    try:
        assert workbook.sheetnames == ["始端箱和插接箱", "报警状态"]
        main = workbook["始端箱和插接箱"]
        assert main["I10"].value == "分路"
        assert main["J10"].value == "设备"
        assert main.column_dimensions["O"].hidden is True
        assert all(
            main.column_dimensions[column].hidden is not True
            for column in ("J", "K", "L", "M", "N")
        )
        assert "K11:N11" in {str(item) for item in main.merged_cells.ranges}
        assert main["K11"].value == "断路器状态字说明"
        assert main["K11"].font.bold is True
        assert main["K11"].fill.fill_type == "solid"
        assert not any(
            str(main.cell(row=row, column=column).value or "").startswith(
                CANONICAL_TOPOLOGY_METADATA_PREFIX
            )
            for row in range(1, main.max_row + 1)
            for column in range(1, 15)
        )
        labels = [
            str(main.cell(row=row, column=9).value or "")
            for row in range(11, main.max_row + 1)
        ]
        expected_labels = {
            "A路输出分路1",
            "A路输出分路2",
            "A路输出分路3",
            "B路输出分路1",
            "B路输出分路4",
            "B路输出分路7",
        }
        assert expected_labels.issubset(set(labels))
        first_label = "A路输出分路1"
        first_label_row = labels.index(first_label) + 11
        assert main.cell(row=first_label_row, column=9).font.bold is True
        assert main.cell(row=first_label_row, column=9).font.sz == 11
        assert main.cell(row=first_label_row, column=9).alignment.wrap_text is True
        device_labels = {
            str(main.cell(row=row, column=10).value or "")
            for row in range(11, main.max_row + 1)
            if main.cell(row=row, column=10).value not in (None, "")
        }
        assert {
            "A路1#监控模块",
            "A路2#监控模块",
            "B路1#监控模块",
            "B路2#监控模块",
        }.issubset(device_labels)
        assert main.cell(row=first_label_row, column=10).value == "A路1#监控模块"
        assert main.cell(row=first_label_row, column=10).font.bold is True
        assert main.cell(row=first_label_row, column=10).font.sz == 12
        metadata_cells = [
            str(main.cell(row=row, column=15).value or "")
            for row in range(11, main.max_row + 1)
            if str(main.cell(row=row, column=15).value or "").startswith(
                CANONICAL_TOPOLOGY_METADATA_PREFIX
            )
        ]
        assert len(metadata_cells) == 10
        assert any(
            "module_no=2" in value
            and "module_local_branch_no=4" in value
            and "communication_variable_device_code=203" in value
            for value in metadata_cells
        )

        excel_main_rows = [
            (str(main.cell(row=row, column=2).value), int(main.cell(row=row, column=6).value))
            for row in range(11, main.max_row + 1)
            if main.cell(row=row, column=2).value is not None
            and isinstance(main.cell(row=row, column=6).value, int)
        ]
        assert excel_main_rows == canonical_main_rows(model)

        alarm = workbook["报警状态"]
        alarm_state_vars = [
            str(alarm.cell(row=row, column=2).value)
            for row in range(8, alarm.max_row + 1)
            if str(alarm.cell(row=row, column=2).value or "").startswith("State_")
        ]
        alarm_bit_text = "\n".join(
            str(alarm.cell(row=row, column=9).value or "")
            for row in range(8, alarm.max_row + 1)
        )
        module_comm_labels = re.findall(r"[AB]路\d+#监控模块通讯异常", alarm_bit_text)
        assert Counter(module_comm_labels) == Counter(
            {
                "A路1#监控模块通讯异常": 1,
                "A路2#监控模块通讯异常": 1,
                "B路1#监控模块通讯异常": 1,
                "B路2#监控模块通讯异常": 1,
            }
        )
    finally:
        workbook.close()

    alarm_code = alarm_code_path.read_text(encoding="utf-8")
    assert len(re.findall(r"^' === State_", alarm_code, flags=re.MULTILINE)) == len(
        alarm_state_vars
    )
    for device_code in ("101", "102", "201", "203"):
        assert f"!GetAlmValue(Comm_EC{device_code}, 0," in alarm_code
    for legacy_wrong_code in ("102_2", "201_2", "202", "204", "203_2", "204_2"):
        assert f"!GetAlmValue(Comm_EC{legacy_wrong_code}, 0," not in alarm_code

    protocol_rows = extract_protocol_point_rows(excel_path)
    csv_rows = read_csv_rows(csv_path)
    assert len(csv_rows) == len(protocol_rows) + 5
    assert payload["program_upload"]["point_count"] == len(protocol_rows)
    for protocol_row, csv_row in zip(protocol_rows, csv_rows[5:], strict=True):
        assert csv_row[1] == protocol_row.var_name
        assert int(csv_row[7]) == protocol_row.register_address + 1
    return elapsed_seconds


def test_legacy_segmented_regression() -> float:
    library = TemplateLibrary.load()
    config = normalize_config(build_legacy_segmented_config(), library)
    model = ProtocolGenerator(library).generate(config)

    branch_addresses: dict[tuple[str, int], int] = {}
    for route_model in model["routes"]:
        for branch in flatten_branches(route_model):
            branch_addresses[(route_model["route"], branch["output_no"])] = branch["points"][0][
                "address"
            ]
    assert len(branch_addresses) == 80
    assert branch_addresses[("A", 1)] == 2000
    assert branch_addresses[("A", 38)] == 4997
    assert branch_addresses[("A", 39)] == 9500
    assert branch_addresses[("A", 40)] == 9581
    assert branch_addresses[("B", 1)] == 5078
    assert branch_addresses[("B", 38)] == 8075
    assert branch_addresses[("B", 39)] == 9662
    assert branch_addresses[("B", 40)] == 9743
    assert len(ClassicCombinedRenderer(model)._build_alarm_rows()) == 148

    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    excel_path = OUTPUT_ROOT / "synthetic-legacy-dynamic-topology.xlsx"
    csv_path = OUTPUT_ROOT / "synthetic-legacy-dynamic-topology.csv"
    alarm_code_path = OUTPUT_ROOT / "synthetic-legacy-dynamic-topology.txt"
    started_at = time.perf_counter()
    render_excel(model, excel_path)
    rendered_at = time.perf_counter()
    validate_model_and_workbook(model, excel_path)
    validated_at = time.perf_counter()
    alarm_code = generate_alarm_code_from_workbook(excel_path)
    alarm_code_path.write_text(alarm_code, encoding="utf-8")
    alarm_generated_at = time.perf_counter()
    csv_result = write_program_upload_csv_from_config(excel_path, csv_path, config)
    elapsed_seconds = time.perf_counter() - started_at
    assert csv_result["point_count"] == 3736
    assert "!GetAlmValue(Comm_EC139, 0," in alarm_code
    assert "!GetAlmValue(Comm_EC239, 0," in alarm_code
    for path in (excel_path, alarm_code_path, csv_path):
        assert path.exists() and path.stat().st_size > 0
    stage_times = {
        "render": rendered_at - started_at,
        "validate": validated_at - rendered_at,
        "alarm": alarm_generated_at - validated_at,
        "csv": elapsed_seconds - (alarm_generated_at - started_at),
    }
    assert stage_times["render"] < 45, stage_times
    assert all(stage_times[key] < 10 for key in ("validate", "alarm", "csv")), stage_times
    # This is a full QA triplet (render + model/workbook validation + alarm +
    # CSV), not just the user-facing API's file-writing portion.
    assert elapsed_seconds < 60, elapsed_seconds
    print(
        "legacy_segmented_dynamic_topology_stages "
        + " ".join(f"{key}={value:.3f}s" for key, value in stage_times.items()),
        flush=True,
    )
    return elapsed_seconds


def main() -> int:
    test_explicit_and_legacy_workbook_mapping()
    dynamic_elapsed = test_dynamic_api_three_file_delivery()
    print(f"dynamic_topology_api_stage elapsed={dynamic_elapsed:.3f}s", flush=True)
    legacy_elapsed = test_legacy_segmented_regression()
    print(
        "protocol_studio_dynamic_topology_ok "
        f"dynamic_api={dynamic_elapsed:.3f}s legacy_triplet={legacy_elapsed:.3f}s"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
