from __future__ import annotations

from pathlib import Path

from _test_support import (
    add_repo_to_import_path,
    configure_process_runtime,
    generated_artifact_path,
)

add_repo_to_import_path()
configure_process_runtime("mcgs-unified-workflow")

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import protocol_studio.app as app_module


def fast_validation_report(canonical_path: Path, excel_path: Path) -> dict:
    return {
        "checked_at": "unified-workflow-test",
        "status": "passed",
        "ok": True,
        "message": "unified workflow structure stub passed",
        "canonical_path": str(canonical_path),
        "excel_path": str(excel_path),
    }


def source_compare_must_not_run(
    excel_path: Path,
    family: str,
    export_profile_id: str | None = None,
    address_profile_id: str | None = None,
) -> dict:
    raise AssertionError("统一工作流不应执行历史模板相似度对比")


def build_unified_config(*, alarm_enabled: bool = True) -> dict:
    return {
        "workflow_version": "unified_protocol_v1",
        "project": {
            "name": "合成统一流程回归",
            "code": "UNIFIED-001",
            "protocol_title": "动环通讯协议",
        },
        "routes": {
            "A": {
                "start_boxes": {"count": 1, "instance_names": ["S1"]},
                "plug_boxes": {
                    "board_number_start": 101,
                    "sequence": [
                        {"type_code": "3P*1", "count": 1, "layout_pattern": "1"},
                    ],
                },
            },
            "B": {"copy_from_A": True},
        },
        "extensions": {
            "single_cabinet": {"enabled": False, "cabinet_count": 0},
            "repeater": {"enabled": False, "A_count": 0, "B_count": 0},
            "alarm_state_word": {
                "enabled": alarm_enabled,
                "base_address": 6100,
                "word_mode": "32bit",
            },
        },
        # These deliberately wrong legacy selections must not drive unified mode.
        "profiles": {
            "export_profile_id": "extended_split_default",
            "address_profile_id": "split_main1000_plug2000_cabinet8200_repeater9000_alarm9200_16bit",
        },
    }


def assert_business_bundle(client: TestClient, payload: dict) -> None:
    assert set(payload["downloads"]) >= {"excel", "alarm_code", "program_upload"}
    assert payload["downloads"]["excel"]
    assert payload["downloads"]["alarm_code"]
    assert payload["downloads"]["program_upload"]
    assert payload["delivery_bundle"]["required_keys"] == ["excel", "alarm_code", "program_upload"]
    assert payload["delivery_bundle"]["status"] == "complete"

    excel_path = generated_artifact_path(app_module, payload, "excel_path")
    alarm_path = generated_artifact_path(app_module, payload, "alarm_code_path")
    upload_path = generated_artifact_path(app_module, payload, "program_upload_path")
    assert excel_path.name == "合成统一流程回归-动环通讯协议.xlsx"
    assert alarm_path.name == "合成统一流程回归-报警状态字上传代码.txt"
    assert upload_path.name == "合成统一流程回归-MCGS动环上传设备导入.csv"
    assert excel_path.exists() and excel_path.stat().st_size > 0
    assert alarm_path.exists() and alarm_path.stat().st_size > 0
    assert upload_path.exists() and upload_path.stat().st_size > 0

    for key in ("excel", "alarm_code", "program_upload"):
        response = client.get(payload["downloads"][key])
        assert response.status_code == 200, (key, response.text)


def main() -> int:
    app_module.run_validation_report = fast_validation_report
    app_module.run_source_compare_report = source_compare_must_not_run
    app_module.generate_alarm_code_from_workbook = lambda workbook_path: "' unified alarm upload code\n"
    client = TestClient(app_module.app)

    bootstrap_response = client.get("/api/bootstrap")
    assert bootstrap_response.status_code == 200, bootstrap_response.text
    bootstrap = bootstrap_response.json()
    assert bootstrap["workflow"]["id"] == "unified_protocol_v1"
    assert bootstrap["workflow"]["product_model"] == "parameterized_protocol_compiler"
    assert bootstrap["workflow"]["template_selection_exposed"] is False
    assert [item["id"] for item in bootstrap["workflow"]["steps"]] == [
        "project",
        "route_a",
        "route_b",
        "extensions",
        "review",
        "generate",
    ]
    assert bootstrap["delivery_bundle"]["required_keys"] == ["excel", "alarm_code", "program_upload"]
    assert bootstrap["capabilities"]["automatic_internal_profiles"] is True
    assert bootstrap["scenarios"], "legacy scenarios must remain for API compatibility"

    extension_config = build_unified_config()
    extension_config["extensions"]["single_cabinet"] = {"enabled": True, "cabinet_count": 2}
    extension_config["extensions"]["repeater"] = {"enabled": True, "A_count": 1, "B_count": 1}
    extension_library = app_module.TemplateLibrary.load()
    normalized_extensions = app_module.normalize_config(extension_config, extension_library)
    selected_export = extension_library.export_profile_map[
        normalized_extensions["profiles"]["export_profile_id"]
    ]
    assert selected_export["include_repeater_sheet"] is True
    assert selected_export["include_single_cabinet_sheet"] is False
    assert selected_export["embed_single_cabinet_in_base_sheet"] is True
    assert selected_export["sheet_order"] == [
        "始端箱和插接箱",
        "中继器",
        "报警状态",
    ]
    assert normalized_extensions["profiles"]["device_library_id"] == extension_library.device_library["id"]

    result_response = client.post("/api/generate", json={"config": build_unified_config()})
    assert result_response.status_code == 200, result_response.text
    payload = result_response.json()
    assert_business_bundle(client, payload)
    assert payload["canonical"]["profiles"]["export_profile"]["id"].startswith("unified_master_")
    assert payload["canonical"]["profiles"]["export_profile"]["family"] == "classic_combined"
    assert payload["canonical"]["profiles"]["address_profile"]["alarm_base"] == 6100
    assert payload["canonical"]["profiles"]["address_profile"]["alarm_word_mode"] == "32bit"
    assert payload["canonical"]["project"]["topology"]["screen_topology_mode"] == "single_screen_one_column"
    assert payload["delivery_status"]["status"] == "deliverable"
    assert payload["delivery_status"]["basis"] == "validation_and_three_file_bundle"
    assert payload["delivery_status"]["source_compare_role"] == "internal_regression_only"
    assert payload["source_compare"]["verdict"] == "skipped"
    assert payload["source_compare"]["skipped_for_unified_workflow"] is True
    workbook = load_workbook(
        generated_artifact_path(app_module, payload, "excel_path"),
        read_only=True,
    )
    assert workbook.sheetnames == ["始端箱和插接箱", "报警状态"]
    workbook.close()

    disabled_alarm_response = client.post(
        "/api/generate",
        json={"config": build_unified_config(alarm_enabled=False)},
    )
    assert disabled_alarm_response.status_code == 200, disabled_alarm_response.text
    disabled_payload = disabled_alarm_response.json()
    assert_business_bundle(client, disabled_payload)
    assert disabled_payload["alarm_codegen"]["status"] == "generated"
    assert disabled_payload["alarm_codegen"]["content_status"] == "not_applicable"
    disabled_excel_path = generated_artifact_path(app_module, disabled_payload, "excel_path")
    disabled_workbook = load_workbook(disabled_excel_path, read_only=True)
    assert disabled_workbook.sheetnames == ["始端箱和插接箱"]
    disabled_workbook.close()
    disabled_alarm_path = generated_artifact_path(app_module, disabled_payload, "alarm_code_path")
    alarm_notice = disabled_alarm_path.read_text(encoding="utf-8")
    assert "未启用报警状态字" in alarm_notice

    print("protocol_studio_unified_workflow_ok")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
