from __future__ import annotations

from collections import Counter
from pathlib import Path

from _test_support import (
    add_repo_to_import_path,
    configure_process_runtime,
    generated_artifact_path,
)

add_repo_to_import_path()
configure_process_runtime("mcgs-two-column")

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import protocol_studio.app as app_module
from mvp_generator.generator import ProtocolGenerator
from mvp_generator.library import TemplateLibrary
from mvp_generator.validator import ConfigError


def build_two_column_config() -> dict:
    return {
        "workflow_version": "unified_protocol_v1",
        "project": {
            "name": "单双列与端口模型回归",
            "code": "TWO-COLUMN-PORTS-001",
            "protocol_title": "动环通讯协议",
        },
        "topology": {
            "screen_topology_mode": "single_screen_two_columns",
            "hardware_form_factor": "horizontal",
            "environment_rs485_port": "A4B4",
            "upload_port_profile": "A4B4",
            "bus_data_port_mode": "double_column_by_column",
            "bus_data_port_assignments": {
                "column_1": "A2B2",
                "column_2": "A3B3",
            },
        },
        "protocol_layout": {
            "measurement_layout_mode": "by_plug_box",
        },
        "routes": {
            "A": {
                "start_boxes": {
                    "count": 1,
                    "instance_names": ["S1"],
                    "device_code_start": 1,
                },
                "plug_boxes": {
                    "box_number_start": 101,
                    "board_number_start": 101,
                    "sequence": [
                        {
                            "type_code": "3P*2",
                            "count": 1,
                            "layout_pattern": "1+1",
                        }
                    ],
                },
            },
            "B": {
                "start_boxes": {
                    "count": 1,
                    "instance_names": ["S2"],
                    "device_code_start": 2,
                },
                "plug_boxes": {
                    "box_number_start": 201,
                    "board_number_start": 201,
                    "sequence": [
                        {
                            "type_code": "3P*2",
                            "count": 1,
                            "layout_pattern": "1+1",
                        }
                    ],
                },
            },
        },
        "devices": {
            "screen_columns": {
                "column_2": {
                    "start_boxes": {
                        "A": {
                            "count": 1,
                            "instance_names": ["S3"],
                            "device_code_start": 3,
                        },
                        "B": {
                            "count": 1,
                            "instance_names": ["S4"],
                            "device_code_start": 4,
                        },
                    },
                    "plug_boxes": {
                        "A": {
                            "box_number_start": 301,
                            "board_number_start": 301,
                            "sequence": [
                                {
                                    "type_code": "3P*1",
                                    "count": 1,
                                    "layout_pattern": "1",
                                }
                            ],
                        },
                        "B": {
                            "box_number_start": 401,
                            "board_number_start": 401,
                            "sequence": [
                                {
                                    "type_code": "3P*1",
                                    "count": 1,
                                    "layout_pattern": "1",
                                }
                            ],
                        },
                    },
                    "branch_modules": {
                        "A": {
                            "module_sequence": [],
                            "module_number_start": 1,
                            "output_number_start": 1,
                            "branch_device_number_start": 301,
                            "variable_numbering_mode": "per_output_contiguous",
                            "names": [],
                        },
                        "B": {
                            "module_sequence": [],
                            "module_number_start": 1,
                            "output_number_start": 1,
                            "branch_device_number_start": 401,
                            "variable_numbering_mode": "per_output_contiguous",
                            "names": [],
                        },
                    },
                }
            }
        },
        "extensions": {
            "single_cabinet": {"enabled": False, "cabinet_count": 0},
            "repeater": {"enabled": False, "A_count": 0, "B_count": 0},
            "alarm_state_word": {
                "enabled": True,
                "base_address": 6000,
                "word_mode": "16bit",
            },
        },
        "profiles": {},
    }


def build_two_column_monitor_config() -> dict:
    config = build_two_column_config()
    config["project"]["name"] = "双列监控模块模型回归"
    config["protocol_layout"]["measurement_layout_mode"] = "by_branch"
    for route, device_start in (("A", 101), ("B", 201)):
        config["routes"][route]["plug_boxes"]["sequence"] = []
        config["routes"][route]["branch_modules"] = {
            "module_sequence": [
                {
                    "type_code": "3P*2",
                    "count": 1,
                    "layout_pattern": "1+1",
                }
            ],
            "module_number_start": 1,
            "output_number_start": 1,
            "branch_device_number_start": device_start,
            "variable_numbering_mode": "per_output_contiguous",
            "names": [],
        }

    second_column = config["devices"]["screen_columns"]["column_2"]
    for route, device_start in (("A", 301), ("B", 401)):
        second_column["plug_boxes"][route]["sequence"] = []
        second_column["branch_modules"][route] = {
            "module_sequence": [
                {
                    "type_code": "3P*1",
                    "count": 1,
                    "layout_pattern": "1",
                }
            ],
            "module_number_start": 1,
            "output_number_start": 1,
            "branch_device_number_start": device_start,
            "variable_numbering_mode": "per_output_contiguous",
            "names": [],
        }
    return config


def find_row_by_value(ws, column: int, value: str) -> int:
    for row_no in range(1, ws.max_row + 1):
        if ws.cell(row=row_no, column=column).value == value:
            return row_no
    raise AssertionError(f"未找到 {value}")


def collect_variable_names(model: dict) -> list[str]:
    names: list[str] = []
    for route in model["routes"]:
        for start_box in route["start_boxes"]:
            names.extend(point["var_name"] for point in start_box["points"])
        for physical_box in route["physical_plug_boxes"]:
            for board in physical_box["boards"]:
                for branch in board["branches"]:
                    names.extend(point["var_name"] for point in branch["points"])
    return names


def expect_config_error(raw_config: dict, expected_text: str) -> None:
    library = TemplateLibrary.load()
    normalized = app_module.normalize_config(raw_config, library)
    try:
        ProtocolGenerator(library).generate(normalized)
    except ConfigError as exc:
        assert expected_text in str(exc), str(exc)
    else:
        raise AssertionError(f"非法端口组合必须硬失败：{expected_text}")


def test_port_validation() -> None:
    collision = build_two_column_config()
    collision["topology"]["bus_data_port_assignments"]["column_2"] = "A4B4"
    expect_config_error(collision, "不能与动环 RS-485 上传口")

    duplicate = build_two_column_config()
    duplicate["topology"]["bus_data_port_assignments"]["column_2"] = "A2B2"
    expect_config_error(duplicate, "两个不同的物理口")

    unavailable = build_two_column_config()
    unavailable["topology"]["hardware_form_factor"] = "din_rail"
    expect_config_error(unavailable, "不支持动环上传口 A4B4")

    alias_conflict = build_two_column_config()
    alias_conflict["topology"]["upload_port_profile"] = "A3B3"
    expect_config_error(alias_conflict, "兼容字段")


def test_valid_port_matrix() -> None:
    library = TemplateLibrary.load()
    hardware_cases = {
        "horizontal": {
            "environment": "A4B4",
            "data_ports": ("A2B2", "A3B3"),
        },
        "din_rail": {
            "environment": "A3B3",
            "data_ports": ("A1B1", "A2B2"),
        },
    }
    for hardware_form_factor, ports in hardware_cases.items():
        for topology_mode, data_mode, assignment_keys in (
            (
                "single_screen_one_column",
                "single_column_shared",
                ("shared",),
            ),
            (
                "single_screen_one_column",
                "single_column_split_ab",
                ("A", "B"),
            ),
            (
                "single_screen_two_columns",
                "double_column_by_column",
                ("column_1", "column_2"),
            ),
            (
                "single_screen_two_columns",
                "double_column_by_route",
                ("A", "B"),
            ),
        ):
            config = build_two_column_config()
            config["topology"].update(
                {
                    "screen_topology_mode": topology_mode,
                    "hardware_form_factor": hardware_form_factor,
                    "environment_rs485_port": ports["environment"],
                    "upload_port_profile": ports["environment"],
                    "bus_data_port_mode": data_mode,
                    "bus_data_port_assignments": {
                        key: ports["data_ports"][min(index, 1)]
                        for index, key in enumerate(assignment_keys)
                    },
                }
            )
            normalized = app_module.normalize_config(config, library)
            model = ProtocolGenerator(library).generate(normalized)
            assert model["project"]["topology"]["hardware_form_factor"] == (
                hardware_form_factor
            )
            assert model["project"]["topology"]["environment_rs485_port"] == (
                ports["environment"]
            )


def test_start_box_count_validation() -> None:
    first_column = build_two_column_config()
    first_column["routes"]["A"]["start_boxes"]["count"] = 2
    first_column["routes"]["A"]["start_boxes"]["instance_names"] = ["S1", "S2"]
    expect_config_error(first_column, "第一列A路始端箱数量只能为 0 或 1")

    second_column = build_two_column_config()
    second_column["devices"]["screen_columns"]["column_2"]["start_boxes"]["A"][
        "count"
    ] = 2
    second_column["devices"]["screen_columns"]["column_2"]["start_boxes"]["A"][
        "instance_names"
    ] = ["S3", "S4"]
    expect_config_error(second_column, "第二列A路始端箱数量只能为 0 或 1")


def test_bootstrap_pointsets_and_port_matrix() -> None:
    client = TestClient(app_module.app)
    response = client.get("/api/bootstrap")
    assert response.status_code == 200, response.text
    payload = response.json()

    expected_point_counts = {
        "plug_branch_standard_29row_connector_temp": 29,
        "plug_branch_compact_21row": 21,
        "plug_branch_mid_26row_partial_connector": 26,
        "plug_branch_dual_dataset_47row": 47,
        "plug_branch_compact_22row_freq": 22,
        "plug_branch_standard_30row_full_connector": 30,
        "plug_branch_extended_load_reactive": 41,
        "plug_branch_single_phase_triplet_30row_full_connector": 30,
    }
    templates = {
        item["id"]: item
        for item in payload["templates"]["plug_branch_templates"]
    }
    for template_id, expected_count in expected_point_counts.items():
        template = templates[template_id]
        points = template["points"]
        assert isinstance(points, list)
        assert template["point_count"] == len(points) == expected_count
        assert [point["index"] for point in points] == list(
            range(1, expected_count + 1)
        )
        for point in points:
            assert set(point) == {
                "index",
                "prefix",
                "variable_pattern",
                "dataset_group",
                "name",
                "unit",
                "data_type",
            }
            assert point["prefix"]
            if template_id == "plug_branch_dual_dataset_47row" and point["dataset_group"] == 2:
                assert point["variable_pattern"] == f"{point['prefix']}{{设备号}}_2"
            else:
                assert point["variable_pattern"] == f"{point['prefix']}{{设备号}}"
            assert point["dataset_group"] in {None, 1, 2}
            assert point["name"] and point["name"] != point["prefix"]
            assert point["unit"] is None or isinstance(point["unit"], str)
            assert point["data_type"] in {
                "16位 无符号二进制",
                "32位 无符号二进制",
                "32位 浮点数",
            }

    hardware = {
        item["value"]: item
        for item in payload["options"]["hardware_form_factors"]
    }
    assert hardware["horizontal"]["available_ports"] == [
        "A2B2",
        "A3B3",
        "A4B4",
    ]
    assert hardware["horizontal"]["default_environment_port"] == "A4B4"
    assert hardware["din_rail"]["available_ports"] == [
        "A1B1",
        "A2B2",
        "A3B3",
    ]
    assert hardware["din_rail"]["default_environment_port"] == "A3B3"
    bus_modes = payload["options"]["bus_data_port_modes"]
    assert {item["value"] for item in bus_modes["single_screen_one_column"]} == {
        "single_column_shared",
        "single_column_split_ab",
    }
    assert {item["value"] for item in bus_modes["single_screen_two_columns"]} == {
        "double_column_by_column",
        "double_column_by_route",
    }


def test_canonical_and_three_file_delivery() -> None:
    library = TemplateLibrary.load()
    config = app_module.normalize_config(build_two_column_config(), library)
    model = ProtocolGenerator(library).generate(config)

    assert model["project"]["topology"]["screen_topology_mode"] == "single_screen_two_columns"
    routes = {route["route"]: route for route in model["routes"]}
    assert [item["device_code"] for item in routes["A"]["start_boxes"]] == [1, 3]
    assert [item["device_code"] for item in routes["B"]["start_boxes"]] == [2, 4]
    assert [item["screen_column"] for item in routes["A"]["start_boxes"]] == [1, 2]
    assert [item["screen_column"] for item in routes["B"]["start_boxes"]] == [1, 2]
    assert [item["physical_box_no"] for item in routes["A"]["physical_plug_boxes"]] == [101, 301]
    assert [item["physical_box_no"] for item in routes["B"]["physical_plug_boxes"]] == [201, 401]
    assert [item["screen_column"] for item in routes["A"]["physical_plug_boxes"]] == [1, 2]
    assert [item["screen_column"] for item in routes["B"]["physical_plug_boxes"]] == [1, 2]

    variable_names = collect_variable_names(model)
    duplicates = [name for name, count in Counter(variable_names).items() if count > 1]
    assert not duplicates, duplicates[:10]
    for required_name in (
        "StateS1",
        "StateS2",
        "StateS3",
        "StateS4",
        "StateC101",
        "StateC201",
        "StateC301",
        "StateC401",
    ):
        assert required_name in variable_names

    client = TestClient(app_module.app)
    response = client.post("/api/generate", json={"config": build_two_column_config()})
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["delivery_bundle"]["status"] == "complete"
    assert payload["alarm_codegen"]["status"] == "generated"
    assert payload["program_upload"]["status"] == "generated"
    assert set(payload["downloads"]) >= {"excel", "alarm_code", "program_upload"}

    artifact_paths = {
        key: generated_artifact_path(app_module, payload, f"{key}_path")
        for key in ("excel", "alarm_code", "program_upload")
    }
    for path in artifact_paths.values():
        assert path.exists() and path.stat().st_size > 0, path

    workbook = load_workbook(artifact_paths["excel"], data_only=True)
    try:
        main = workbook["始端箱和插接箱"]
        expected_start_labels = {
            "StateS1": "第一列A路始端箱S1",
            "StateS2": "第一列B路始端箱S2",
            "StateS3": "第二列A路始端箱S3",
            "StateS4": "第二列B路始端箱S4",
        }
        for var_name, expected_label in expected_start_labels.items():
            row_no = find_row_by_value(main, 2, var_name)
            assert main.cell(row=row_no, column=9).value == expected_label

        expected_plug_labels = {
            "StateC101": ("分路1", "第一列A路插接箱101"),
            "StateC201": ("分路1", "第一列B路插接箱201"),
            "StateC301": ("分路1", "第二列A路插接箱301"),
            "StateC401": ("分路1", "第二列B路插接箱401"),
        }
        for var_name, (branch_label, device_label) in expected_plug_labels.items():
            row_no = find_row_by_value(main, 2, var_name)
            assert main.cell(row=row_no, column=9).value == branch_label
            assert main.cell(row=row_no, column=10).value == device_label
    finally:
        workbook.close()

    alarm_code = artifact_paths["alarm_code"].read_text(encoding="utf-8")
    assert alarm_code.strip()
    csv_bytes = artifact_paths["program_upload"].read_bytes()
    assert csv_bytes


def test_two_column_monitor_module_ids() -> None:
    library = TemplateLibrary.load()
    config = app_module.normalize_config(build_two_column_monitor_config(), library)
    model = ProtocolGenerator(library).generate(config)
    routes = {route["route"]: route for route in model["routes"]}

    expected_ids = {
        "A": ["A-M001", "C2-A-M002"],
        "B": ["B-M001", "C2-B-M002"],
    }
    expected_device_codes = {
        "A": ["101", "301"],
        "B": ["201", "401"],
    }
    for route in ("A", "B"):
        modules = routes[route]["physical_plug_boxes"]
        assert [module["module_id"] for module in modules] == expected_ids[route]
        assert [module["module_no"] for module in modules] == [1, 2]
        assert [module["display_module_no"] for module in modules] == [1, 1]
        assert [module["communication_alarm_slot"] for module in modules] == [1, 2]
        assert [module["screen_column"] for module in modules] == [1, 2]
        assert [module["communication_variable_device_code"] for module in modules] == (
            expected_device_codes[route]
        )
        branch_output_numbers = [
            branch["output_no"]
            for module in modules
            for board in module["boards"]
            for branch in board["branches"]
        ]
        assert branch_output_numbers == [1, 2, 3]

    variable_names = collect_variable_names(model)
    duplicates = [
        name for name, count in Counter(variable_names).items() if count > 1
    ]
    assert not duplicates, duplicates[:10]

    client = TestClient(app_module.app)
    response = client.post(
        "/api/generate",
        json={"config": build_two_column_monitor_config()},
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["delivery_bundle"]["status"] == "complete"
    assert payload["alarm_codegen"]["status"] == "generated"
    assert payload["program_upload"]["status"] == "generated"

    excel_path = generated_artifact_path(app_module, payload, "excel_path")
    alarm_code_path = generated_artifact_path(app_module, payload, "alarm_code_path")
    csv_path = generated_artifact_path(app_module, payload, "program_upload_path")
    for path in (excel_path, alarm_code_path, csv_path):
        assert path.exists() and path.stat().st_size > 0, path

    workbook = load_workbook(excel_path, data_only=True)
    try:
        main = workbook["始端箱和插接箱"]
        expected_module_labels = {
            "StateC101": (
                "第一列A路模块内分路1",
                "第一列A路1#监控模块",
            ),
            "StateC201": (
                "第一列B路模块内分路1",
                "第一列B路1#监控模块",
            ),
            "StateC301": (
                "第二列A路模块内分路1",
                "第二列A路1#监控模块",
            ),
            "StateC401": (
                "第二列B路模块内分路1",
                "第二列B路1#监控模块",
            ),
        }
        for var_name, (branch_label, module_label) in expected_module_labels.items():
            row_no = find_row_by_value(main, 2, var_name)
            assert main.cell(row=row_no, column=9).value == branch_label
            assert main.cell(row=row_no, column=10).value == module_label

        alarm = workbook["报警状态"]
        alarm_bit_text = "\n".join(
            str(alarm.cell(row=row_no, column=9).value or "")
            for row_no in range(1, alarm.max_row + 1)
        )
        assert "第二列A路1#监控模块" in alarm_bit_text
        assert "第二列B路1#监控模块" in alarm_bit_text
        assert "第二列A路2#监控模块" not in alarm_bit_text
        assert "第二列B路2#监控模块" not in alarm_bit_text
        assert "第二列A路输出分路3" not in alarm_bit_text
        assert "第二列B路输出分路3" not in alarm_bit_text
    finally:
        workbook.close()

    alarm_code = alarm_code_path.read_text(encoding="utf-8")
    for communication_code in ("101", "201", "301", "401"):
        assert f"Comm_EC{communication_code}" in alarm_code
    for wrong_cross_column_code in ("Comm_EC102", "Comm_EC202"):
        assert wrong_cross_column_code not in alarm_code

    csv_bytes = csv_path.read_bytes()
    csv_text = None
    for encoding in ("gb18030", "utf-8-sig", "utf-8"):
        try:
            csv_text = csv_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    assert csv_text is not None
    for var_name in ("StateC101", "StateC201", "StateC301", "StateC401"):
        assert var_name in csv_text


def main() -> int:
    test_bootstrap_pointsets_and_port_matrix()
    test_port_validation()
    test_valid_port_matrix()
    test_start_box_count_validation()
    test_two_column_monitor_module_ids()
    test_canonical_and_three_file_delivery()
    print("protocol_studio_two_column_topology_ok")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
