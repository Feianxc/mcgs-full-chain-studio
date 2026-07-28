from __future__ import annotations

from copy import deepcopy
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from .excel_renderer import (
    AlarmRow,
    ClassicCombinedRenderer,
    FILL_BLUE,
    FILL_CLASSIC_PLUG,
    FILL_INTRO,
    FILL_LIQUIDCOOL_CABINET_IA,
    FILL_ORANGE,
    FILL_THEME0,
    FILL_THEME3,
    FILL_WHITE,
    FILL_YELLOW,
    PHASE_LABELS,
    TOPOLOGY_MODE_LABELS,
    build_alarm_rows_from_group_specs,
    clear_workbook_freeze_panes,
    fill_cell,
    merge_with_alignment,
    normalize_metric_code,
    set_fill_range,
    temperature_phase_label,
)


FILL_AB_ALT = PatternFill("solid", fgColor="FFEEECE1")
FILL_EXTENDED_CABINET_EA = PatternFill("solid", fgColor="FFDCE6F1")

EXTENDED_START_PREFIX_META: dict[str, tuple[str, str | None]] = {
    "LoadS": ("始端箱负载率", "%"),
    "UBS": ("三相不平衡度", "%"),
    "Pa": ("A相有功功率", "kW"),
    "Pb": ("B相有功功率", "kW"),
    "Pc": ("C相有功功率", "kW"),
    "P": ("总有功功率", "kW"),
    "Qa": ("A相无功功率", "kVar"),
    "Qb": ("B相无功功率", "kVar"),
    "Qc": ("C相无功功率", "kVar"),
    "Q": ("总无功功率", "kVar"),
    "Sa": ("A相视在功率", "kVA"),
    "Sb": ("B相视在功率", "kVA"),
    "Sc": ("C相视在功率", "kVA"),
    "S_": ("总视在功率", "kVA"),
    "Ea": ("A相有功电度", "kWh"),
    "Eb": ("B相有功电度", "kWh"),
    "Ec": ("C相有功电度", "kWh"),
    "E": ("总有功电度", "kWh"),
    "Eqa": ("A相无功电度", "kVarh"),
    "Eqb": ("B相无功电度", "kVarh"),
    "Eqc": ("C相无功电度", "kVarh"),
    "Eq": ("总无功电度", "kVarh"),
}

EXTENDED_PLUG_PREFIX_META: dict[str, tuple[str, str | None]] = {
    "StateC": ("输出支路断路器状态字", None),
    "Loada": ("A相负载率", "%"),
    "Loadb": ("b相负载率", "%"),
    "Loadc": ("c相负载率", "%"),
    "Load": ("三相输出总负载率", "%"),
    "Pa": ("A相有功功率", "kW"),
    "Pb": ("B相有功功率", "kW"),
    "Pc": ("C相有功功率", "kW"),
    "P": ("总有功功率", "kW"),
    "Qa": ("A相无功功率", "kVar"),
    "Qb": ("B相无功功率", "kVar"),
    "Qc": ("C相无功功率", "kVar"),
    "Q": ("总无功功率", "kVar"),
    "Sa": ("A相视在功率", "kVA"),
    "Sb": ("B相视在功率", "kVA"),
    "Sc": ("C相视在功率", "kVA"),
    "S_": ("总视在功率", "kVA"),
    "Ta": ("插脚接点温度A", "℃"),
    "Tb": ("插脚接点温度B", "℃"),
    "Tc": ("插脚接点温度C", "℃"),
    "Tn": ("插脚接点温度N", "℃"),
    "Th": ("箱内环境温度", "℃"),
    "Ea": ("A相有功电度", "kWh"),
    "Eb": ("B相有功电度", "kWh"),
    "Ec": ("C相有功电度", "kWh"),
    "E": ("总有功电度", "kWh"),
    "Eqa": ("A相无功电度", "kVarh"),
    "Eqb": ("B相无功电度", "kVarh"),
    "Eqc": ("C相无功电度", "kVarh"),
    "Eq": ("总无功电度", "kVarh"),
}


class StructuredSplitRendererBase(ClassicCombinedRenderer):
    DATA_HEADERS = ["通道号", "变量名", "变量类型", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "单位", "", "", "", "", ""]
    REPEATER_HEADERS = [
        "通道号",
        "变量名",
        "变量类型",
        "通道名称",
        "读写类型",
        "寄存器名称",
        "数据类型",
        "寄存器地址",
        "",
        "单位",
        "",
        "",
        "",
        "",
    ]
    ALARM_HEADERS = ["通道号", "变量名", "变量类型", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "单位", "", "", "", "", ""]
    DATA_NOTE = (
        "说明：模拟量采用 IEEE 754 单精度（32 位）浮点格式，每个测点连续占用两个 16 位 Modbus 寄存器；"
        "低地址寄存器存放高 16 位，高地址寄存器存放低 16 位。状态字的数据宽度按所选点集模板生成。"
    )
    ALARM_NOTE = "下面的数据中各Bit位0代表正常，1代表监控屏/母线本地产生了对应类型的报警(没用到的Bit位不写出来)"
    BREAKER_STATUS_TEXT = "1：闭合，0：断开"
    COMMON_32BIT_NOTE = (
        "说明：模拟量采用 IEEE 754 单精度（32 位）浮点格式，每个测点连续占用两个 16 位 Modbus 寄存器。"
        "读取时按高字在前的字序组合：低地址寄存器存放高 16 位，高地址寄存器存放低 16 位；"
        "若测点起始地址为 N，则 N 为高字，N+1 为低字。"
    )

    def _prepare_data_sheet(
        self,
        ws,
        intro_text: str,
        *,
        headers: list[str] | None = None,
        note_text: str | None = None,
        header_row: int = 11,
        note_row: int = 9,
    ) -> None:
        line1, line2 = self._communication_lines()
        ws["A1"] = line1
        ws["A3"] = line2
        ws["A5"] = intro_text
        ws[f"A{note_row}"] = note_text or self.DATA_NOTE
        self._write_header_row(ws, header_row, headers or self.DATA_HEADERS)

    def _prepare_alarm_sheet(self, ws, intro_text: str, note_text: str | None = None) -> None:
        line1, line2 = self._communication_lines()
        ws["A1"] = line1
        ws["A3"] = line2
        ws["A5"] = intro_text
        self._write_header_row(ws, 7, self.ALARM_HEADERS)
        ws["A8"] = note_text or self.ALARM_NOTE
        ws["J8"] = "各Bit位含义"

    def _append_plain_point_block(
        self,
        ws,
        row_no: int,
        channel_no: int,
        points: list[dict[str, Any]],
        *,
        primary_label: str | None = None,
        secondary_label: str | None = None,
        primary_label_column: int = 10,
        secondary_label_column: int = 11,
        breaker_note_columns: tuple[int, int, int] | None = None,
        breaker_title: str | None = None,
        describe_point: Callable[[dict[str, Any]], tuple[str, str | None]] | None = None,
    ) -> tuple[int, int]:
        if not points:
            return row_no, channel_no

        describe = describe_point or (lambda point: self._describe_prefix(point["prefix"]))
        state_row = row_no if points and points[0]["prefix"].startswith("State") else None

        for index, point in enumerate(points):
            desc, unit = describe(point)
            data_type_label = self._point_data_type_label(point)
            ws.cell(row=row_no, column=1, value=channel_no)
            ws.cell(row=row_no, column=2, value=point["var_name"])
            ws.cell(row=row_no, column=3, value="SINGLE")
            ws.cell(row=row_no, column=4, value="只读")
            ws.cell(row=row_no, column=5, value="[4区]输出寄存器")
            ws.cell(row=row_no, column=6, value=data_type_label)
            ws.cell(row=row_no, column=7, value=point["address"])
            ws.cell(row=row_no, column=8, value=desc)
            ws.cell(row=row_no, column=9, value=unit)
            if index == 0 and primary_label:
                ws.cell(row=row_no, column=primary_label_column, value=primary_label)
            if index == 0 and secondary_label:
                ws.cell(row=row_no, column=secondary_label_column, value=secondary_label)
            row_no += 1
            channel_no += 1

        if state_row is not None and breaker_note_columns and breaker_title:
            bit_col, label_col, status_col = breaker_note_columns
            ws.cell(row=state_row, column=bit_col, value=breaker_title)
            for offset, phase_label in enumerate(PHASE_LABELS, start=1):
                ws.cell(row=state_row + offset, column=bit_col, value=f"BIT{offset - 1}")
                ws.cell(row=state_row + offset, column=label_col, value=f"{phase_label}相断路器")
                ws.cell(row=state_row + offset, column=status_col, value=self.BREAKER_STATUS_TEXT)

        return row_no, channel_no

    def _render_repeater_sheet_structured(self, ws, intro_text: str) -> None:
        self._prepare_data_sheet(ws, intro_text, headers=self.REPEATER_HEADERS, note_text=self.COMMON_32BIT_NOTE)
        row_no = 12
        channel_no = 0
        for route_model in self.model["routes"]:
            for repeater in route_model["repeater_units"]:
                block_start_row = row_no
                display_route = route_model["route"]
                if self.export_profile.get("family") == "extended_split" and route_model["route"] == "B":
                    display_route = "A"
                group_label = self._repeater_entity_label(display_route, repeater)
                row_no, channel_no = self._append_point_block(
                    ws=ws,
                    row_no=row_no,
                    channel_no=channel_no,
                    points=repeater["points"],
                    group_label=group_label,
                    annotation_lines=None,
                )
                if row_no - 1 > block_start_row:
                    ws.merge_cells(start_row=block_start_row, start_column=11, end_row=row_no - 1, end_column=11)
        ws.freeze_panes = "A12"

    def _render_alarm_rows(self, ws, intro_text: str, rows: list[AlarmRow], note_text: str | None = None) -> None:
        self._prepare_alarm_sheet(ws, intro_text, note_text=note_text)
        row_no = 9
        channel_no = 0
        backfill_description = self.export_profile.get("family") == "extended_split"
        last_description: str | None = None
        for row in rows:
            description = row.description
            if backfill_description and description in (None, ""):
                description = last_description
            ws.cell(row=row_no, column=1, value=channel_no)
            ws.cell(row=row_no, column=2, value=row.var_name)
            ws.cell(row=row_no, column=3, value="SINGLE")
            ws.cell(row=row_no, column=4, value="只读")
            ws.cell(row=row_no, column=5, value="[4区]输出寄存器")
            ws.cell(row=row_no, column=6, value=row.data_type_label)
            ws.cell(row=row_no, column=7, value=row.register_address)
            ws.cell(row=row_no, column=8, value=description)
            ws.cell(row=row_no, column=10, value="\n".join(row.bit_lines))
            if description not in (None, ""):
                last_description = description
            row_no += 1
            channel_no += 1
        ws.freeze_panes = "A9"

    def _merge_ranges(self, ws, ranges: Iterable[str]) -> None:
        existing = {str(item) for item in ws.merged_cells.ranges}
        for cell_range in ranges:
            if cell_range not in existing:
                ws.merge_cells(cell_range)
                existing.add(cell_range)

    def _set_row_heights(self, ws, heights: dict[int, float]) -> None:
        for row_index, height in heights.items():
            ws.row_dimensions[row_index].height = height

    def _apply_common_style(self, ws) -> None:
        super()._apply_common_style(ws)
        family = self.export_profile.get("family")
        widths: dict[str, float] | None = None
        if family == "extended_split":
            widths = self._extended_widths(ws.title)
        elif family == "ab_screen_split":
            widths = self._ab_widths(ws.title)
        if widths:
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

    def _extended_widths(self, title: str) -> dict[str, float] | None:
        mapping = {
            "始端箱": {
                "A": 8.9062, "B": 14.4531, "C": 13.8164, "D": 13.0, "E": 20.0, "F": 19.1797, "G": 12.9062,
                "H": 29.8164, "I": 8.9062, "J": 26.8164, "K": 13.0, "L": 27.543, "M": 22.4531, "N": 27.7266, "O": 8.9062,
            },
            "插接箱": {
                "A": 13.0, "B": 11.4531, "C": 13.0, "D": 13.0, "E": 17.4531, "F": 17.543, "G": 20.9062,
                "H": 33.0, "I": 13.0, "J": 25.9062, "K": 22.6328, "L": 21.4531, "M": 19.4531, "N": 18.0, "O": 13.0,
            },
            "单机柜数据": {
                "A": 13.0, "B": 13.0, "C": 12.4531, "D": 11.6328, "E": 18.2695, "F": 20.7266, "G": 13.3633,
                "H": 24.0, "I": 13.0, "J": 25.6328, "K": 20.2695, "L": 27.3633, "M": 13.0, "N": 13.0, "O": 13.0,
            },
            self._repeater_sheet_name(): {
                "A": 13.0, "B": 9.6328, "C": 14.3633, "D": 13.1797, "E": 13.8164, "F": 20.9062, "G": 15.9062,
                "H": 14.1797, "I": 15.8164, "J": 13.0, "K": 19.0898,
            },
            "报警状态": {
                "A": 8.9062, "B": 17.0, "C": 13.8164, "D": 13.0, "E": 17.3633, "F": 19.1797, "G": 13.8164,
                "H": 37.8164, "I": 8.9062, "J": 51.4531, "K": 8.9062, "L": 13.0, "M": 13.0, "N": 13.0, "O": 13.0,
            },
        }
        return mapping.get(title)

    def _ab_widths(self, title: str) -> dict[str, float] | None:
        data_widths = {
            "A": 8.8867, "B": 14.4414, "C": 13.7773, "D": 13.0, "E": 17.332, "F": 19.2188, "G": 12.8867,
            "H": 21.8867, "I": 8.8867, "J": 26.7773, "K": 27.5547, "L": 17.1094, "M": 27.6641, "N": 8.8867, "O": 13.0,
        }
        alarm_widths = {
            "A": 8.8867, "B": 14.4414, "C": 13.7773, "D": 13.0, "E": 17.332, "F": 19.2188, "G": 13.7773,
            "H": 32.8867, "I": 8.8867, "J": 51.4414, "K": 8.8867, "L": 13.0, "M": 13.0, "N": 13.0, "O": 13.0,
        }
        if title in {"A路屏数据", "B路屏数据"}:
            return data_widths
        if title in {"A路屏报警", "B路屏报警"}:
            return alarm_widths
        return None

    def _copy_points_with_new_device_suffix(
        self,
        points: list[dict[str, Any]],
        suffix_builder: Callable[[dict[str, Any]], str],
    ) -> list[dict[str, Any]]:
        copied = []
        for point in points:
            new_point = deepcopy(point)
            new_point["var_name"] = f"{point['prefix']}{suffix_builder(point)}"
            copied.append(new_point)
        return copied

    def _assign_addresses(self, points: list[dict[str, Any]], start_address: int) -> tuple[list[dict[str, Any]], int]:
        copied = deepcopy(points)
        cursor = start_address
        for point in copied:
            point["address"] = cursor
            register_size = point.get("register_size") or (1 if "16位" in self._point_data_type_label(point) else 2)
            point["register_size"] = register_size
            cursor += register_size
        return copied, cursor


class ExtendedSplitRenderer(StructuredSplitRendererBase):
    def render_to_path(self, output_path: Path) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        start_ws = workbook.create_sheet("始端箱")
        self._render_start_sheet(start_ws)

        plug_ws = workbook.create_sheet("插接箱")
        self._render_plug_sheet(plug_ws)

        cabinet_ws = workbook.create_sheet("单机柜数据")
        self._render_cabinet_sheet(cabinet_ws)

        repeater_ws = workbook.create_sheet(self._repeater_sheet_name())
        self._render_repeater_sheet_structured(
            repeater_ws,
            f"该页面为{self._repeater_sheet_name()}点位预览，寄存器从{self.address_profile.get('repeater_base', '未配置')}开始",
        )

        alarm_ws = workbook.create_sheet("报警状态")
        self._render_alarm_rows(
            alarm_ws,
            "该界面为报警状态字上传表格，以插接箱内的监控单元寄存器从9200开始，",
            self._build_alarm_rows(),
            note_text="下面的数据中各Bit位0代表正常，1代表母线本地产生了对应类型的报警(没用到的Bit位不写出来)",
        )

        self._apply_extended_layout(start_ws, kind="start")
        self._apply_extended_layout(plug_ws, kind="plug")
        self._apply_extended_layout(cabinet_ws, kind="cabinet")
        self._apply_extended_layout(repeater_ws, kind="repeater")
        self._apply_extended_layout(alarm_ws, kind="alarm")

        for ws in workbook.worksheets:
            self._apply_common_style(ws)

        clear_workbook_freeze_panes(workbook)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    def _render_start_sheet(self, ws) -> None:
        self._prepare_data_sheet(
            ws,
            (
                "该项目由一台小母线监控屏管理 A/B 两路母线。动环 RS-485 上传仅占用一个物理接口，"
                "具体端口由设备形态与现场接线配置决定；以下设备数量及地址均按本次项目参数生成。"
            ),
            note_text=self.COMMON_32BIT_NOTE,
        )
        row_no = 12
        channel_no = 0
        breaker_note_written = False
        for route_model in self.model["routes"]:
            for start_box in route_model["start_boxes"]:
                block_start_row = row_no
                row_no, channel_no = self._append_plain_point_block(
                    ws,
                    row_no,
                    channel_no,
                    start_box["points"],
                    primary_label=f"{route_model['route']}路{start_box['instance_name']}始端箱",
                    breaker_note_columns=(11, 12, 13),
                    breaker_title="始端箱断路器状态字说明" if not breaker_note_written else None,
                    describe_point=self._describe_extended_start_point,
                )
                breaker_note_written = True
                if row_no - 1 > block_start_row:
                    ws.merge_cells(start_row=block_start_row, start_column=10, end_row=row_no - 1, end_column=10)
        ws.freeze_panes = "A12"

    def _render_plug_sheet(self, ws) -> None:
        self._prepare_data_sheet(
            ws,
            "该表按监控模块组织插接箱测点。每个监控模块可配置一块或多块板卡，实际输出分路数量由插接箱类型及板卡布局决定；"
            "通讯报警按监控模块生成，遥测与模拟量报警按实际输出分路生成。插接箱、监控模块、输出分路与机柜的对应关系以现场接线为准，"
            "数据寄存器从 2000 开始。",
            note_text=self.COMMON_32BIT_NOTE,
        )
        row_no = 12
        channel_no = 0
        for route_model in self.model["routes"]:
            output_index = 1
            module_index = 1
            for physical_box in route_model["physical_plug_boxes"]:
                for board in physical_box["boards"]:
                    board_branches = [branch for branch in board["branches"] if branch.get("points")]
                    board_start_row = row_no
                    for branch_position, branch in enumerate(board_branches, start=1):
                        primary_label = f"{route_model['route']}路{output_index}#输出支路"
                        if branch.get("branch_kind") == "single_phase_triplet_aggregate":
                            primary_label += "（A/B/C单相）"
                        block_start_row = row_no
                        row_no, channel_no = self._append_plain_point_block(
                            ws,
                            row_no,
                            channel_no,
                            branch["points"],
                            primary_label=primary_label,
                            secondary_label=f"{route_model['route']}路{module_index}#监控模块" if branch_position == 1 else None,
                            breaker_note_columns=(12, 13, 14),
                            breaker_title="输出断路器状态字说明" if branch_position == 1 else None,
                            describe_point=self._describe_extended_plug_point,
                        )
                        if row_no - 1 > block_start_row:
                            ws.merge_cells(start_row=block_start_row, start_column=10, end_row=row_no - 1, end_column=10)
                        output_index += 1
                    if board_branches:
                        if row_no - 1 > board_start_row:
                            ws.merge_cells(start_row=board_start_row, start_column=11, end_row=row_no - 1, end_column=11)
                        module_index += 1
        ws.freeze_panes = "A12"

    def _render_cabinet_sheet(self, ws) -> None:
        self._prepare_data_sheet(
            ws,
            "该表包含每列单机柜的主备路数据之和上传点位，数据寄存器从8200开始",
            note_text=self.COMMON_32BIT_NOTE,
        )
        row_no = 12
        channel_no = 0
        note_text = (
            "所谓的1#机柜是指：监控屏这头开始，\n"
            "小母线插接箱接线最靠近监控屏的柜子为1#,接的第二个柜子为2#，后面依此类推。\n"
            "1#机柜上面的标签不一定是“XX01柜”"
        )
        for index, item in enumerate(self.model.get("single_cabinet_rows", []), start=1):
            ws.cell(row=row_no, column=1, value=channel_no)
            ws.cell(row=row_no, column=2, value=item["var_name"])
            ws.cell(row=row_no, column=3, value="SINGLE")
            ws.cell(row=row_no, column=4, value="只读")
            ws.cell(row=row_no, column=5, value="[4区]输出寄存器")
            ws.cell(row=row_no, column=6, value=item.get("data_type_label", "32位 浮点数"))
            ws.cell(row=row_no, column=7, value=item["address"])
            description = str(item.get("description") or f"{index}#机柜和电流").replace("总电流", "和电流")
            ws.cell(row=row_no, column=8, value=description)
            ws.cell(row=row_no, column=9, value=item.get("unit", "A"))
            if index == 1:
                ws.cell(row=row_no, column=10, value=note_text)
            row_no += 1
            channel_no += 1
        ws.freeze_panes = "A12"

    def _describe_extended_start_point(self, point: dict[str, Any]) -> tuple[str, str | None]:
        return EXTENDED_START_PREFIX_META.get(point["prefix"], self._describe_prefix(point["prefix"]))

    def _describe_extended_plug_point(self, point: dict[str, Any]) -> tuple[str, str | None]:
        return EXTENDED_PLUG_PREFIX_META.get(point["prefix"], self._describe_prefix(point["prefix"]))

    def _build_alarm_rows(self) -> list[AlarmRow]:
        groups: list[tuple[str, list[tuple[str, list[str]]]]] = []
        routes = self.model["routes"]
        groups.append(("SPD", [(f"{route['route']}路始端箱浪涌报警", self._extended_bits_spd(route)) for route in routes]))
        groups.append(("THD", [(f"{route['route']}路始端箱谐波报警", self._extended_bits_thd(route)) for route in routes]))
        groups.append(("In", [(f"{route['route']}路始端箱漏电流报警", self._extended_bits_in(route)) for route in routes]))
        groups.append(("FH", [(f"{route['route']}路始端箱频率上限报警", self._extended_bits_frequency(route, '频率超上限')) for route in routes]))
        groups.append(("FL", [(f"{route['route']}路始端箱频率下限报警", self._extended_bits_frequency(route, '频率超下限')) for route in routes]))
        groups.append(("LoadH", [(f"{route['route']}路始端箱负载量上限报警", self._extended_bits_load_high(route)) for route in routes]))
        groups.append(("UnbH", [(f"{route['route']}路始端箱三相不平衡度超上限报警", self._extended_bits_unbalance(route)) for route in routes]))
        groups.append(("Power", [(f"{route['route']}路电源模块异常报警", [f"{route['route']}路电源模块异常"]) for route in routes]))
        groups.append(("Com", [(f"{route['route']}路设备通讯异常报警", self._extended_bits_com(route)) for route in routes]))
        groups.append(("LoadHC", [(f"{route['route']}路设备负载率超上限报警", self._extended_bits_load_by_phase(route)) for route in routes]))
        groups.append(("VH", [(f"{route['route']}路设备电压超上限报警", self._extended_bits_metric(route, 'U', '电压超上限')) for route in routes]))
        groups.append(("VL", [(f"{route['route']}路设备电压超下限报警", self._extended_bits_metric(route, 'U', '电压超下限')) for route in routes]))
        groups.append(("PH", [(f"{route['route']}路设备功率超上限报警", self._extended_bits_metric(route, 'P', '功率超上限')) for route in routes]))
        groups.append(("IH", [(f"{route['route']}路设备电流超上限报警", self._extended_bits_metric(route, 'I', '电流超上限')) for route in routes]))
        groups.append(("PFL", [(f"{route['route']}路设备功率因数超下限报警", self._extended_bits_metric(route, 'PF', '功率因数超下限')) for route in routes]))
        groups.append(("VLL", [(f"{route['route']}路设备分闸报警", self._extended_bits_metric(route, 'State', '分闸')) for route in routes]))
        groups.append(("TH", [(f"{route['route']}路设备温度超上限报警", self._extended_bits_temperature(route)) for route in routes]))
        return build_alarm_rows_from_group_specs(self.address_profile, groups)

    def _apply_extended_layout(self, ws, *, kind: str) -> None:
        if kind == "start":
            self._merge_ranges(ws, ["A1:H2", "A3:H4", "A5:G8", "A9:G10", "K12:M12"])
            self._set_row_heights(ws, {idx: 21.0 for idx in range(1, 11)} | {11: 18.0, 12: 18.0, 13: 18.0, 14: 18.0})
            fill_cell(ws["A5"], FILL_INTRO)
            fill_cell(ws["K12"], FILL_YELLOW)
            for route_model in self.model["routes"]:
                for start_box in route_model["start_boxes"]:
                    label = f"{route_model['route']}路{start_box['instance_name']}始端箱"
                    for row_index in range(12, ws.max_row + 1):
                        if ws.cell(row=row_index, column=10).value == label:
                            end_row = row_index
                            while end_row + 1 <= ws.max_row and ws.cell(row=end_row + 1, column=2).value:
                                if ws.cell(row=end_row + 1, column=10).value is not None:
                                    break
                                end_row += 1
                            block_fill = FILL_BLUE if route_model["route"] == "A" else FILL_THEME0
                            if route_model["route"] == "A":
                                set_fill_range(ws, row_index, end_row, 1, 8, block_fill)
                                fill_cell(ws.cell(row=row_index, column=10), block_fill)
                            else:
                                set_fill_range(ws, row_index, end_row, 1, 10, block_fill)
            return

        if kind == "plug":
            self._merge_ranges(ws, ["A1:H2", "A3:H4", "A5:G8", "A9:G10", "L12:N12"])
            self._set_row_heights(ws, {idx: 21.0 for idx in range(1, 11)} | {11: 18.0, 12: 14.0, 13: 14.0, 14: 14.0})
            fill_cell(ws["A5"], FILL_INTRO)
            fill_cell(ws["L12"], FILL_YELLOW)
            current_fill = FILL_CLASSIC_PLUG
            for row_index in range(12, ws.max_row + 1):
                if ws.cell(row=row_index, column=2).value is None:
                    continue
                fill_cell(ws.cell(row=row_index, column=1), FILL_BLUE)
                label_text = str(ws.cell(row=row_index, column=10).value or "")
                if label_text:
                    current_fill = FILL_CLASSIC_PLUG if "A路" in label_text else FILL_THEME0
                set_fill_range(ws, row_index, row_index, 2, 8, current_fill)
            for col_index in (10, 11):
                for row_index in range(12, ws.max_row + 1):
                    if ws.cell(row=row_index, column=col_index).value:
                        label_text = str(ws.cell(row=row_index, column=col_index).value or "")
                        fill_cell(ws.cell(row=row_index, column=col_index), FILL_CLASSIC_PLUG if "A路" in label_text else FILL_THEME0)
            return

        if kind == "cabinet":
            self._merge_ranges(ws, ["A1:H2", "A3:H4", "A5:G8", "A9:G10"])
            self._set_row_heights(ws, {idx: 15.0 for idx in range(1, 11)} | {11: 21.0, 12: 18.0, 13: 18.0, 14: 18.0})
            fill_cell(ws["A5"], FILL_INTRO)
            if ws.max_row >= 15:
                merge_with_alignment(ws, "J12:M15")
            cabinet_rows = self.model.get("single_cabinet_rows", [])
            note_row = None
            for offset, item in enumerate(cabinet_rows, start=12):
                metric_code = normalize_metric_code(item.get("metric_code") or item.get("var_name"))
                fill_cell(ws.cell(row=offset, column=1), FILL_BLUE)
                fill = {
                    "IA": FILL_LIQUIDCOOL_CABINET_IA,
                    "PA": FILL_YELLOW,
                    "EA": FILL_EXTENDED_CABINET_EA,
                    "KA": FILL_INTRO,
                }.get(metric_code, FILL_CLASSIC_PLUG)
                set_fill_range(ws, offset, offset, 2, 8, fill)
                if metric_code == "KA" and note_row is None:
                    note_row = offset
            if note_row is not None:
                merge_with_alignment(ws, f"J{note_row}:L{note_row}")
                fill_cell(ws.cell(row=note_row, column=10), FILL_YELLOW)
                ws.cell(row=note_row, column=10, value="机柜通电状态字说明")
                ws.cell(row=note_row + 1, column=10, value="BIT0")
                ws.cell(row=note_row + 1, column=11, value="机柜通电状态")
                ws.cell(row=note_row + 1, column=12, value="1：通电，0：断电")
            return

        if kind == "repeater":
            self._merge_ranges(ws, ["A1:H2", "A3:H4", "A5:G8", "A9:G10"])
            self._set_row_heights(ws, {idx: 15.0 for idx in range(1, 11)} | {11: 23.0, 12: 15.0, 13: 15.0, 14: 15.0})
            fill_cell(ws["A5"], FILL_INTRO)
            a_route_repeater_count = len(next((route["repeater_units"] for route in self.model["routes"] if route["route"] == "A"), []))
            block_index = -1
            current_fill = FILL_ORANGE
            for row_index in range(12, ws.max_row + 1):
                if ws.cell(row=row_index, column=2).value is None:
                    continue
                if ws.cell(row=row_index, column=11).value:
                    block_index += 1
                    if self.export_profile.get("family") == "extended_split":
                        current_fill = FILL_ORANGE if block_index < a_route_repeater_count else FILL_LIQUIDCOOL_CABINET_IA
                    else:
                        current_fill = FILL_ORANGE if "A路" in str(ws.cell(row=row_index, column=11).value) else FILL_LIQUIDCOOL_CABINET_IA
                fill_cell(ws.cell(row=row_index, column=1), FILL_BLUE)
                set_fill_range(ws, row_index, row_index, 2, 9, current_fill)
            return

        if kind == "alarm":
            self._merge_ranges(ws, ["A1:H2", "A3:H4", "A5:H6", "A8:I8"])
            self._set_row_heights(ws, {idx: 21.0 for idx in range(1, 7)} | {7: 20.0, 8: 30.0, 9: 21.0, 10: 21.0, 11: 21.0, 12: 21.0})
            fill_cell(ws["A5"], FILL_INTRO)
            fill_cell(ws["J8"], FILL_THEME3)
            highlight_started = False
            for row_index in range(9, ws.max_row + 1):
                if ws.cell(row=row_index, column=2).value is None:
                    continue
                fill_cell(ws.cell(row=row_index, column=1), FILL_BLUE)
                description = str(ws.cell(row=row_index, column=8).value or "")
                if description and "通讯异常" in description:
                    highlight_started = True
                if highlight_started:
                    set_fill_range(ws, row_index, row_index, 2, 7, FILL_CLASSIC_PLUG)
                    if ws.cell(row=row_index, column=8).value not in (None, ""):
                        fill_cell(ws.cell(row=row_index, column=8), FILL_CLASSIC_PLUG)
                    fill_cell(ws.cell(row=row_index, column=9), FILL_CLASSIC_PLUG)
                    fill_cell(ws.cell(row=row_index, column=10), FILL_CLASSIC_PLUG)

    def _extended_bits_spd(self, route_model: dict[str, Any]) -> list[str]:
        return [f"{route_model['route']}路始端箱{item['instance_name']}浪涌故障" for item in route_model["start_boxes"]]

    def _extended_bits_thd(self, route_model: dict[str, Any]) -> list[str]:
        return [
            f"{route_model['route']}路始端箱{item['instance_name']}谐波超上限"
            for item in route_model["start_boxes"]
            if self._has_prefix(item["points"], "THD")
        ]

    def _extended_bits_in(self, route_model: dict[str, Any]) -> list[str]:
        return [
            f"{route_model['route']}路始端箱{item['instance_name']}漏电流超上限"
            for item in route_model["start_boxes"]
            if self._has_prefix(item["points"], "In")
        ]

    def _extended_bits_frequency(self, route_model: dict[str, Any], suffix: str) -> list[str]:
        return [
            f"{route_model['route']}路始端箱{item['instance_name']}{suffix}"
            for item in route_model["start_boxes"]
            if self._has_prefix(item["points"], "F")
        ]

    def _extended_bits_load_high(self, route_model: dict[str, Any]) -> list[str]:
        return [
            f"{route_model['route']}路始端箱{item['instance_name']}负载率超上限"
            for item in route_model["start_boxes"]
            if self._has_prefix(item["points"], "LoadS")
        ]

    def _extended_bits_unbalance(self, route_model: dict[str, Any]) -> list[str]:
        return [
            f"{route_model['route']}路始端箱{item['instance_name']}三相不平衡度超上限报警"
            for item in route_model["start_boxes"]
            if self._has_prefix(item["points"], "UBS")
        ]

    def _extended_bits_com(self, route_model: dict[str, Any]) -> list[str]:
        bits = [f"{route_model['route']}路始端箱{item['instance_name']}通讯异常" for item in route_model["start_boxes"]]
        module_index = 1
        for physical_box in route_model["physical_plug_boxes"]:
            for board in physical_box["boards"]:
                if any(branch.get("points") for branch in board["branches"]):
                    bits.append(f"{route_model['route']}路{module_index}#监控模块通讯异常")
                    module_index += 1
        for repeater in route_model["repeater_units"]:
            if repeater.get("points"):
                bits.append(f"{self._repeater_entity_label(route_model['route'], repeater)}通讯异常")
        return bits

    def _iter_extended_branch_entities(self, route_model: dict[str, Any]) -> Iterable[dict[str, Any]]:
        module_index = 1
        for physical_box in route_model["physical_plug_boxes"]:
            for board in physical_box["boards"]:
                board_branches = [branch for branch in board["branches"] if branch.get("points")]
                for branch_position, branch in enumerate(board_branches, start=1):
                    yield {"module_index": module_index, "branch_position": branch_position, "points": branch["points"]}
                if board_branches:
                    module_index += 1

    def _extended_bits_load_by_phase(self, route_model: dict[str, Any]) -> list[str]:
        bits: list[str] = []
        for item in route_model["start_boxes"]:
            if self._has_prefix(item["points"], "LoadS"):
                for phase_label in PHASE_LABELS:
                    bits.append(f"{route_model['route']}路始端箱{item['instance_name']}-{phase_label}相负载率超上限")
        phase_prefixes = {"A": "Loada", "B": "Loadb", "C": "Loadc"}
        for entity in self._iter_extended_branch_entities(route_model):
            for phase_label in PHASE_LABELS:
                if self._has_prefix(entity["points"], phase_prefixes[phase_label]):
                    bits.append(f"{route_model['route']}路{entity['module_index']}#监控模块分路{entity['branch_position']}-{phase_label}相负载率超上限")
        return bits

    def _extended_bits_metric(self, route_model: dict[str, Any], metric_kind: str, suffix: str) -> list[str]:
        bits: list[str] = []
        for item in route_model["start_boxes"]:
            if self._supports_three_phase_metric(item["points"], metric_kind):
                for phase_label in PHASE_LABELS:
                    bits.append(f"{route_model['route']}路始端箱{item['instance_name']}-{phase_label}相{suffix}")
        for entity in self._iter_extended_branch_entities(route_model):
            if self._supports_three_phase_metric(entity["points"], metric_kind):
                for phase_label in PHASE_LABELS:
                    bits.append(f"{route_model['route']}路{entity['module_index']}#监控模块分路{entity['branch_position']}-{phase_label}相{suffix}")
        return bits

    def _extended_bits_temperature(self, route_model: dict[str, Any]) -> list[str]:
        bits: list[str] = []
        for item in route_model["start_boxes"]:
            for phase_label in self._temperature_phase_labels(item["points"]):
                bits.append(f"{route_model['route']}路始端箱{item['instance_name']}-{phase_label}温度超上限")
        for entity in self._iter_extended_branch_entities(route_model):
            for phase_label in self._temperature_phase_labels(entity["points"]):
                bits.append(f"{route_model['route']}路{entity['module_index']}#监控模块分路{entity['branch_position']}-{phase_label}温度超上限")
        return bits

    def _supports_three_phase_metric(self, points: list[dict[str, Any]], metric_kind: str) -> bool:
        if metric_kind == "State":
            return self._has_prefix(points, "StateS") or self._has_prefix(points, "StateC")
        if metric_kind == "I":
            return all(self._has_prefix(points, prefix) for prefix in ("Ia", "Ib", "Ic"))
        if metric_kind == "U":
            return all(self._has_prefix(points, prefix) for prefix in ("Ua", "Ub", "Uc"))
        if metric_kind == "P":
            return all(self._has_prefix(points, prefix) for prefix in ("Pa", "Pb", "Pc"))
        if metric_kind == "PF":
            return all(self._has_prefix(points, prefix) for prefix in ("PFa", "PFb", "PFc"))
        return False

    def _temperature_phase_labels(self, points: list[dict[str, Any]]) -> list[str]:
        labels: list[str] = []
        for point in points:
            if point["prefix"].startswith("THD"):
                continue
            if not point["prefix"].startswith("T"):
                continue
            phase_label = temperature_phase_label(point["prefix"])
            if phase_label and phase_label not in labels:
                labels.append(phase_label)
        return labels

class AbScreenSplitRenderer(StructuredSplitRendererBase):
    COLUMN_SPECS = (
        {"label": "A列", "start_offset": 0, "branch_offset": 0},
        {"label": "B列", "start_offset": 2, "branch_offset": 200},
    )

    AB_POINT_META = {
        "StateS": ("断路器状态字", None),
        "StateC": ("断路器状态字", None),
        "Ia": ("A相电流(1路)", "A"),
        "Ib": ("B相电流(2路)", "A"),
        "Ic": ("C相电流(3路)", "A"),
        "In": ("零序漏电流", "A"),
        "Ua": ("A相电压(1路)", "V"),
        "Ub": ("B相电压(2路)", "V"),
        "Uc": ("C相电压(3路)", "V"),
        "F": ("频率", "Hz"),
        "Pa": ("A相有功功率(1路)", "KW"),
        "Pb": ("B相有功功率(2路)", "KW"),
        "Pc": ("C相有功功率(3路)", "KW"),
        "P": ("总有功功率", "KW"),
        "PFa": ("A相功率因数(1路)", None),
        "PFb": ("B相功率因数(2路)", None),
        "PFc": ("C相功率因数(3路)", None),
        "PF": ("总功率因数", None),
        "Ta": ("入线接点温度A(1路)", "℃"),
        "Tb": ("入线接点温度B(2路)", "℃"),
        "Tc": ("入线接点温度C(3路)", "℃"),
        "Tn": ("入线接点温度N(4路)", "℃"),
        "TaO": ("出线接点温度A(5路)", "℃"),
        "TbO": ("出线接点温度B(6路)", "℃"),
        "TcO": ("出线接点温度C(7路)", "℃"),
        "TnO": ("出线接点温度N(8路)", "℃"),
        "Ea": ("A相有功电量(1路)", "KWH"),
        "Eb": ("B相有功电量(2路)", "KWH"),
        "Ec": ("C相有功电量(3路)", "KWH"),
        "E": ("总有功电量", "KWH"),
    }

    def render_to_path(self, output_path: Path) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        for route in ("A", "B"):
            data_ws = workbook.create_sheet(f"{route}路屏数据")
            self._render_screen_data_sheet(data_ws, self._route_model(route))
            self._apply_ab_layout(data_ws, kind="data")
            alarm_ws = workbook.create_sheet(f"{route}路屏报警")
            self._render_screen_alarm_sheet(alarm_ws, self._route_model(route))
            self._apply_ab_layout(alarm_ws, kind="alarm")

        for ws in workbook.worksheets:
            self._apply_common_style(ws)

        clear_workbook_freeze_panes(workbook)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    def _route_model(self, route: str) -> dict[str, Any]:
        return next(item for item in self.model["routes"] if item["route"] == route)

    def _render_screen_data_sheet(self, ws, route_model: dict[str, Any]) -> None:
        self._prepare_data_sheet(
            ws,
            (
                "该项目现场每组通道中一个小母线屏监控两列A路(主路)，一个屏监控两列B路(备路)，两块屏独立\n"
                "通过屏后转接板的串口上传数据，下面工作表是分为A路屏、B路屏两种来定义的\n"
                "说是现场规定了面朝屏左手那列被称为A列，右手那列称为B列，调试阶段按此规律若后续变化另说"
            ),
            note_text="注：除了断路器状态字外，其它模拟量数据都是32位数据占用两个寄存器，默认小寄存器存中放的高字节，例如寄存器1069和1070组成了该始端箱的总电量，1069中存放的是用电量的高字节",
        )
        row_no = 12
        channel_no = 0
        cursor = self.address_profile["main_base"]

        for column_spec in self.COLUMN_SPECS:
            for start_box in route_model["start_boxes"]:
                block_start_row = row_no
                cloned_points = self._copy_points_with_new_device_suffix(
                    start_box["points"],
                    lambda _: str(start_box["device_code"] + column_spec["start_offset"]),
                )
                cloned_points, cursor = self._assign_addresses(cloned_points, cursor)
                row_no, channel_no = self._append_plain_point_block(
                    ws,
                    row_no,
                    channel_no,
                    cloned_points,
                    primary_label=f"{column_spec['label']}{route_model['route']}路始端箱",
                    breaker_note_columns=(11, 12, 13),
                    breaker_title="断路器状态字说明" if column_spec["label"] == "A列" else None,
                    describe_point=self._describe_ab_point,
                )
                if row_no - 1 > block_start_row:
                    ws.merge_cells(start_row=block_start_row, start_column=10, end_row=row_no - 1, end_column=10)

        for column_spec in self.COLUMN_SPECS:
            box_index = 1
            for physical_box in route_model["physical_plug_boxes"]:
                branch_lines = self._ab_branch_mapping_lines(physical_box)
                box_label = f"{column_spec['label']}{route_model['route']}路{box_index}#插接箱"
                if branch_lines:
                    box_label = box_label + "\n" + "\n".join(branch_lines)
                first_branch_in_box = True
                box_start_row = row_no
                for branch in self._iter_physical_box_branches(physical_box):
                    block_start_row = row_no
                    cloned_points = self._copy_points_with_new_device_suffix(
                        branch["points"],
                        lambda _: self._offset_branch_code(branch["variable_device_code"], column_spec["branch_offset"]),
                    )
                    cloned_points, cursor = self._assign_addresses(cloned_points, cursor)
                    row_no, channel_no = self._append_plain_point_block(
                        ws,
                        row_no,
                        channel_no,
                        cloned_points,
                        primary_label=box_label if first_branch_in_box else None,
                        breaker_note_columns=None,
                        breaker_title=None,
                        describe_point=self._describe_ab_point,
                    )
                    if row_no - 1 > block_start_row:
                        ws.merge_cells(start_row=block_start_row, start_column=10, end_row=row_no - 1, end_column=10)
                    first_branch_in_box = False
                if row_no - 1 > box_start_row and column_spec["label"] == "A列" and box_index == 1:
                    self._merge_ranges(ws, [f"K{box_start_row}:M{box_start_row + 2}"])
                    ws.cell(
                        row=box_start_row,
                        column=11,
                        value="无论A路还是B路插接箱，都是从靠近A路始端箱这头开始算1#插接箱设备",
                    )
                box_index += 1
            if column_spec["label"] == "A列":
                row_no += 1

        ws.freeze_panes = "A12"

    def _render_screen_alarm_sheet(self, ws, route_model: dict[str, Any]) -> None:
        self._render_alarm_rows(
            ws,
            "",
            self._build_screen_alarm_rows(route_model),
            note_text="下面的数据中各Bit位0代表正常，1代表监控屏产生了对应类型的报警(没用到的Bit位不写出来)",
        )

    def _build_screen_alarm_rows(self, route_model: dict[str, Any]) -> list[AlarmRow]:
        groups: list[tuple[str, list[tuple[str, list[str]]]]] = []
        groups.append(("SPD", [(f"{spec['label']}{route_model['route']}路始端箱浪涌报警", self._ab_bits_spd(route_model, spec)) for spec in self.COLUMN_SPECS]))
        groups.append(("THD", [(f"{spec['label']}{route_model['route']}路始端箱谐波报警", self._ab_bits_thd(route_model, spec)) for spec in self.COLUMN_SPECS]))
        groups.append(("In", [(f"{spec['label']}{route_model['route']}路始端箱漏电流报警", self._ab_bits_in(route_model, spec)) for spec in self.COLUMN_SPECS]))
        groups.append(("FH", [(f"{spec['label']}{route_model['route']}路始端箱频率上限报警", self._ab_bits_frequency(route_model, spec, '频率超上限')) for spec in self.COLUMN_SPECS]))
        groups.append(("FL", [(f"{spec['label']}{route_model['route']}路始端箱频率下限报警", self._ab_bits_frequency(route_model, spec, '频率超下限')) for spec in self.COLUMN_SPECS]))
        groups.append(("Com", [(f"监控屏与子设备通讯异常报警\n（{spec['label']}{route_model['route']}路）", self._ab_bits_com(route_model, spec)) for spec in self.COLUMN_SPECS]))
        groups.append(("VH", [(f"电压上限报警\n（{spec['label']}{route_model['route']}路）", self._ab_bits_metric(route_model, spec, 'U', '电压超上限')) for spec in self.COLUMN_SPECS]))
        groups.append(("VL", [(f"电压下限报警\n（{spec['label']}{route_model['route']}路）", self._ab_bits_metric(route_model, spec, 'U', '电压超下限')) for spec in self.COLUMN_SPECS]))
        groups.append(("VLL", [(f"分闸报警\n（{spec['label']}{route_model['route']}路）", self._ab_bits_metric(route_model, spec, 'State', '分闸')) for spec in self.COLUMN_SPECS]))
        groups.append(("IH", [(f"电流上限报警\n（{spec['label']}{route_model['route']}路）", self._ab_bits_metric(route_model, spec, 'I', '电流超上限')) for spec in self.COLUMN_SPECS]))
        groups.append(("PFL", [(f"功率因数下限报警\n（{spec['label']}{route_model['route']}路）", self._ab_bits_metric(route_model, spec, 'PF', '功率因数超下限')) for spec in self.COLUMN_SPECS]))
        groups.append(("PH", [(f"功率上限报警\n（{spec['label']}{route_model['route']}路）", self._ab_bits_metric(route_model, spec, 'P', '功率超上限')) for spec in self.COLUMN_SPECS]))
        groups.append(("TH", [(f"温度上限报警\n（{spec['label']}{route_model['route']}路）", self._ab_bits_temperature(route_model, spec)) for spec in self.COLUMN_SPECS]))
        return build_alarm_rows_from_group_specs(self.address_profile, groups)

    def _describe_ab_point(self, point: dict[str, Any]) -> tuple[str, str | None]:
        return self.AB_POINT_META.get(point["prefix"], self._describe_prefix(point["prefix"]))

    def _offset_branch_code(self, code: str, offset: int) -> str:
        if "_" in code:
            base, suffix = code.split("_", 1)
            return f"{int(base) + offset}_{suffix}"
        return str(int(code) + offset)

    def _iter_physical_box_branches(self, physical_box: dict[str, Any]) -> Iterable[dict[str, Any]]:
        for board in physical_box["boards"]:
            for branch in board["branches"]:
                if branch.get("points"):
                    yield branch

    def _ab_branch_mapping_lines(self, physical_box: dict[str, Any]) -> list[str]:
        branches = list(self._iter_physical_box_branches(physical_box))
        if len(branches) == 1 and branches[0].get("branch_kind") != "single_phase_triplet_aggregate":
            return ["(分路1-A相)", "(分路2-B相)", "(分路3-C相)"]

        lines: list[str] = []
        for branch in branches:
            if branch.get("branch_kind") == "single_phase_triplet_aggregate":
                outputs = "/".join(str(item) for item in branch.get("logical_output_labels", ["A", "B", "C"]))
                lines.append(f"(单板三单相 {outputs})")
                continue
            phase_label = PHASE_LABELS[(branch["physical_branch_index"] - 1) % len(PHASE_LABELS)]
            lines.append(f"(分路{branch['physical_branch_index']}-{phase_label}相)")
        return lines

    def _ab_bits_spd(self, route_model: dict[str, Any], column_spec: dict[str, Any]) -> list[str]:
        return [f"{column_spec['label']}{route_model['route']}路始端箱浪涌故障" for _ in route_model["start_boxes"]]

    def _ab_bits_thd(self, route_model: dict[str, Any], column_spec: dict[str, Any]) -> list[str]:
        bits: list[str] = []
        for start_box in route_model["start_boxes"]:
            if self._has_prefix(start_box["points"], "THD"):
                bits.append(f"{column_spec['label']}{route_model['route']}路始端箱谐波超上限")
        return bits

    def _ab_bits_in(self, route_model: dict[str, Any], column_spec: dict[str, Any]) -> list[str]:
        bits: list[str] = []
        for start_box in route_model["start_boxes"]:
            if self._has_prefix(start_box["points"], "In"):
                bits.append(f"{column_spec['label']}{route_model['route']}路始端箱漏电流超上限")
        return bits

    def _ab_bits_frequency(self, route_model: dict[str, Any], column_spec: dict[str, Any], suffix: str) -> list[str]:
        bits: list[str] = []
        for start_box in route_model["start_boxes"]:
            if self._has_prefix(start_box["points"], "F"):
                bits.append(f"{column_spec['label']}{route_model['route']}路始端箱{suffix}")
        return bits

    def _ab_bits_com(self, route_model: dict[str, Any], column_spec: dict[str, Any]) -> list[str]:
        bits = [f"{column_spec['label']}{route_model['route']}路始端箱通讯异常" for _ in route_model["start_boxes"]]
        box_index = 1
        for physical_box in route_model["physical_plug_boxes"]:
            if any(True for _ in self._iter_physical_box_branches(physical_box)):
                bits.append(f"{column_spec['label']}{route_model['route']}路{box_index}#插接箱通讯异常")
            box_index += 1
        return bits

    def _ab_bits_metric(self, route_model: dict[str, Any], column_spec: dict[str, Any], metric_kind: str, suffix: str) -> list[str]:
        bits: list[str] = []
        for start_box in route_model["start_boxes"]:
            if self._supports_ab_metric(start_box["points"], metric_kind):
                for phase_label in PHASE_LABELS:
                    bits.append(f"{column_spec['label']}{route_model['route']}路始端箱-{phase_label}相{suffix}")
        box_index = 1
        for physical_box in route_model["physical_plug_boxes"]:
            for branch in self._iter_physical_box_branches(physical_box):
                if self._supports_ab_metric(branch["points"], metric_kind):
                    for phase_label in PHASE_LABELS:
                        bits.append(f"{column_spec['label']}{route_model['route']}路{box_index}#插接箱分路{branch['physical_branch_index']}-{phase_label}相{suffix}")
            box_index += 1
        return bits

    def _ab_bits_temperature(self, route_model: dict[str, Any], column_spec: dict[str, Any]) -> list[str]:
        bits: list[str] = []
        for start_box in route_model["start_boxes"]:
            for phase_label in self._temperature_phase_labels(start_box["points"]):
                bits.append(f"{column_spec['label']}{route_model['route']}路始端箱-{phase_label}温度超上限")
        box_index = 1
        for physical_box in route_model["physical_plug_boxes"]:
            labels: list[str] = []
            for branch in self._iter_physical_box_branches(physical_box):
                for phase_label in self._temperature_phase_labels(branch["points"]):
                    if phase_label not in labels:
                        labels.append(phase_label)
            for phase_label in labels:
                bits.append(f"{column_spec['label']}{route_model['route']}路{box_index}#插接箱插脚-{phase_label}温度超上限")
            box_index += 1
        return bits

    def _supports_ab_metric(self, points: list[dict[str, Any]], metric_kind: str) -> bool:
        if metric_kind == "State":
            return self._has_prefix(points, "StateS") or self._has_prefix(points, "StateC")
        if metric_kind == "I":
            return all(self._has_prefix(points, prefix) for prefix in ("Ia", "Ib", "Ic"))
        if metric_kind == "U":
            return all(self._has_prefix(points, prefix) for prefix in ("Ua", "Ub", "Uc"))
        if metric_kind == "P":
            return all(self._has_prefix(points, prefix) for prefix in ("Pa", "Pb", "Pc"))
        if metric_kind == "PF":
            return all(self._has_prefix(points, prefix) for prefix in ("PFa", "PFb", "PFc"))
        return False

    def _temperature_phase_labels(self, points: list[dict[str, Any]]) -> list[str]:
        labels: list[str] = []
        for point in points:
            if point["prefix"].startswith("THD"):
                continue
            if not point["prefix"].startswith("T"):
                continue
            phase_label = temperature_phase_label(point["prefix"])
            if phase_label and phase_label not in labels:
                labels.append(phase_label)
        return labels

    def _apply_ab_layout(self, ws, *, kind: str) -> None:
        if kind == "data":
            self._merge_ranges(ws, ["A1:H2", "A3:H4", "A5:G8", "A9:G10", "K12:M12"])
            self._set_row_heights(ws, {idx: 21.0 for idx in range(1, 11)} | {11: 22.2, 12: 21.0, 13: 21.0, 14: 21.0})
            fill_cell(ws["A5"], FILL_INTRO)
            fill_cell(ws["K12"], FILL_YELLOW)
            block_fill = FILL_BLUE
            for row_index in range(12, ws.max_row + 1):
                if ws.cell(row=row_index, column=2).value is None:
                    continue
                fill_cell(ws.cell(row=row_index, column=1), FILL_BLUE)
                if ws.cell(row=row_index, column=10).value is not None:
                    label = str(ws.cell(row=row_index, column=10).value)
                    if "插接箱" in label:
                        block_fill = FILL_CLASSIC_PLUG if "A列" in label else FILL_AB_ALT
                    elif "B列" in label:
                        block_fill = FILL_WHITE
                    else:
                        block_fill = FILL_BLUE if "A列" in label else FILL_WHITE
                    fill_cell(ws.cell(row=row_index, column=10), block_fill)
                set_fill_range(ws, row_index, row_index, 2, 8, block_fill)
            return

        if kind == "alarm":
            self._merge_ranges(ws, ["A1:H2", "A3:H4", "A8:I8"])
            self._set_row_heights(ws, {idx: 21.0 for idx in range(1, 7)} | {7: 19.95, 8: 30.0, 9: 21.0, 10: 21.0, 11: 21.0, 12: 21.0})
            fill_cell(ws["J8"], FILL_THEME3)
            alt_fill = FILL_CLASSIC_PLUG
            for row_index in range(9, ws.max_row + 1):
                if ws.cell(row=row_index, column=2).value is None:
                    continue
                if ws.cell(row=row_index, column=8).value:
                    label = str(ws.cell(row=row_index, column=8).value)
                    alt_fill = FILL_CLASSIC_PLUG if "A列" in label else FILL_AB_ALT
                fill_cell(ws.cell(row=row_index, column=1), FILL_BLUE)
                set_fill_range(ws, row_index, row_index, 2, 8, alt_fill)
            pair_starts = [row for row in range(21, ws.max_row, 2) if ws.cell(row=row, column=2).value and ws.cell(row + 1, column=2).value]
            for row_index in pair_starts:
                if ws.cell(row=row_index + 1, column=8).value in (None, ""):
                    merge_with_alignment(ws, f"H{row_index}:H{row_index + 1}")
            for row_index in [21, 23, 27, 31, 35, 39, 43]:
                if row_index + 1 > ws.max_row:
                    continue
                for col_index in range(12, 19):
                    merge_with_alignment(
                        ws,
                        f"{ws.cell(row=row_index, column=col_index).column_letter}{row_index}:{ws.cell(row=row_index + 1, column=col_index).column_letter}{row_index + 1}",
                    )
