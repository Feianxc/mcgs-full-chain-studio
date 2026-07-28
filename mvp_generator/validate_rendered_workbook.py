from __future__ import annotations

import argparse
import json
import zipfile
from xml.etree import ElementTree as ET
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

if __package__ in (None, ""):
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from mvp_generator.excel_renderer import ClassicCombinedRenderer, LIQUIDCOOL_THRESHOLD_ROWS
    from mvp_generator.split_renderers import AbScreenSplitRenderer, ExtendedSplitRenderer
else:
    from .excel_renderer import ClassicCombinedRenderer, LIQUIDCOOL_THRESHOLD_ROWS
    from .split_renderers import AbScreenSplitRenderer, ExtendedSplitRenderer


LIQUIDCOOL_EXPORT_PROFILE_ID = "classic_combined_liquidcool_default"
SPREADSHEET_NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _validate_excel_package_structure(excel_path: Path) -> None:
    invalid_sheet_views: list[str] = []
    frozen_sheet_views: list[str] = []
    with zipfile.ZipFile(excel_path) as zf:
        worksheet_xml_names = sorted(
            name
            for name in zf.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        for worksheet_xml_name in worksheet_xml_names:
            root = ET.fromstring(zf.read(worksheet_xml_name))
            sheet_view = root.find("./a:sheetViews/a:sheetView", SPREADSHEET_NS)
            if sheet_view is None:
                continue
            pane = sheet_view.find("./a:pane", SPREADSHEET_NS)
            if pane is not None:
                frozen_sheet_views.append(worksheet_xml_name)
            for selection in sheet_view.findall("./a:selection", SPREADSHEET_NS):
                pane_name = selection.get("pane")
                if pane_name and pane is None:
                    invalid_sheet_views.append(
                        f"{worksheet_xml_name} 的 selection.pane={pane_name}，但缺少 pane 节点"
                    )
    if invalid_sheet_views:
        raise AssertionError("Excel package sheetView 非法：" + "；".join(invalid_sheet_views))
    if frozen_sheet_views:
        raise AssertionError("生成的 Excel 不应包含冻结/拆分窗格：" + "；".join(frozen_sheet_views))


def _non_empty_values(ws, column_index: int, start_row: int) -> list[str]:
    values: list[str] = []
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row=row, column=column_index).value
        if value is not None:
            values.append(str(value))
    return values


def _flatten_start_points(model: dict[str, Any]) -> list[dict[str, Any]]:
    points: list[dict[str, Any]] = []
    for route_model in model["routes"]:
        for start_box in route_model["start_boxes"]:
            points.extend(start_box["points"])
    return points


def _flatten_plug_points(model: dict[str, Any]) -> list[dict[str, Any]]:
    points: list[dict[str, Any]] = []
    for route_model in model["routes"]:
        for physical_box in route_model["physical_plug_boxes"]:
            for board in physical_box["boards"]:
                for branch in board["branches"]:
                    points.extend(branch.get("points", []))
    return points


def _flatten_repeater_points(model: dict[str, Any]) -> list[dict[str, Any]]:
    points: list[dict[str, Any]] = []
    for route_model in model["routes"]:
        for repeater in route_model["repeater_units"]:
            points.extend(repeater["points"])
    return points


def _build_renderer(model: dict[str, Any]):
    family = model["profiles"]["export_profile"]["family"]
    if family == "classic_combined":
        return ClassicCombinedRenderer(model)
    if family == "extended_split":
        return ExtendedSplitRenderer(model)
    if family == "ab_screen_split":
        return AbScreenSplitRenderer(model)
    raise NotImplementedError(f"未支持的 family: {family}")


def _expected_ab_points(renderer: AbScreenSplitRenderer, route_model: dict[str, Any]) -> list[dict[str, Any]]:
    cursor = renderer.address_profile["main_base"]
    points: list[dict[str, Any]] = []
    for column_spec in renderer.COLUMN_SPECS:
        for start_box in route_model["start_boxes"]:
            cloned = renderer._copy_points_with_new_device_suffix(
                start_box["points"],
                lambda _: str(start_box["device_code"] + column_spec["start_offset"]),
            )
            cloned, cursor = renderer._assign_addresses(cloned, cursor)
            points.extend(cloned)
    for column_spec in renderer.COLUMN_SPECS:
        for physical_box in route_model["physical_plug_boxes"]:
            for branch in renderer._iter_physical_box_branches(physical_box):
                cloned = renderer._copy_points_with_new_device_suffix(
                    branch["points"],
                    lambda _: renderer._offset_branch_code(branch["variable_device_code"], column_spec["branch_offset"]),
                )
                cloned, cursor = renderer._assign_addresses(cloned, cursor)
                points.extend(cloned)
    return points


def _validate_classic(model: dict[str, Any], renderer: ClassicCombinedRenderer, wb) -> None:
    export_profile_id = renderer.export_profile.get("id")
    is_liquidcool = export_profile_id == LIQUIDCOOL_EXPORT_PROFILE_ID
    is_two_columns = renderer._is_two_column_profile() and not is_liquidcool
    is_unified = renderer._is_unified_master()
    has_repeater_sheet = renderer._should_render_repeater_sheet()
    has_single_cabinet_sheet = renderer._should_render_single_cabinet_sheet()
    has_alarm_sheet = renderer._should_render_alarm_sheet()

    expected_combined_points = _flatten_start_points(model) + _flatten_plug_points(model)
    embedded_cabinet_rows = (
        model.get("single_cabinet_rows", [])
        if renderer._embed_single_cabinet_in_base_sheet()
        else []
    )
    expected_repeater_points = _flatten_repeater_points(model)

    expected_alarm_rows = (
        len(LIQUIDCOOL_THRESHOLD_ROWS) + len(renderer._build_liquidcool_state_alarm_rows())
        if is_liquidcool
        else len(renderer._build_alarm_rows())
    )
    base_sheet_name = renderer._base_sheet_name()
    expected_sheets = [base_sheet_name]
    if has_repeater_sheet:
        expected_sheets.append(renderer._repeater_sheet_name())
    if has_alarm_sheet:
        expected_sheets.append("报警状态")
    if has_single_cabinet_sheet:
        expected_sheets.append("单机柜数据")

    if wb.sheetnames != expected_sheets:
        raise AssertionError(f"sheet 顺序/名称不一致: {wb.sheetnames} != {expected_sheets}")

    combined_ws = wb[base_sheet_name]
    repeater_ws = wb[renderer._repeater_sheet_name()] if has_repeater_sheet else None
    alarm_ws = wb["报警状态"] if has_alarm_sheet else None

    combined_start_row = renderer._combined_sheet_spec().data_start_row
    if len(_non_empty_values(combined_ws, 2, combined_start_row)) != len(expected_combined_points) + len(embedded_cabinet_rows):
        raise AssertionError("classic combined 行数不一致")
    if repeater_ws is not None and len(_non_empty_values(repeater_ws, 2, 7)) != len(expected_repeater_points):
        raise AssertionError("classic repeater 行数不一致")
    classic_alarm_start_row = 9 if is_liquidcool else renderer._alarm_sheet_spec().data_start_row
    if alarm_ws is not None and len(_non_empty_values(alarm_ws, 2, classic_alarm_start_row)) != expected_alarm_rows:
        raise AssertionError("classic alarm 行数不一致")

    if expected_combined_points:
        first_var_cell = f"B{combined_start_row}"
        first_addr_cell = (
            f"F{combined_start_row}"
            if is_unified
            else (f"G{combined_start_row}" if is_two_columns else f"H{combined_start_row}")
        )
        if combined_ws[first_var_cell].value != expected_combined_points[0]["var_name"] or combined_ws[first_addr_cell].value != expected_combined_points[0]["address"]:
            raise AssertionError("classic 首点不一致")
    if not is_liquidcool and "生成警告" in str(combined_ws["A5"].value or ""):
        raise AssertionError("classic 说明区不应输出内部生成 warning")
    if repeater_ws is not None and expected_repeater_points:
        repeater_address_cell = "F7" if is_unified else "H7"
        if repeater_ws["B7"].value != expected_repeater_points[0]["var_name"] or repeater_ws[repeater_address_cell].value != expected_repeater_points[0]["address"]:
            raise AssertionError("classic 中继首点不一致")

    if has_single_cabinet_sheet:
        cabinet_rows = model.get("single_cabinet_rows", [])
        cabinet_ws = wb["单机柜数据"]
        if len(_non_empty_values(cabinet_ws, 2, 6)) != len(cabinet_rows):
            raise AssertionError("classic 单机柜行数不一致")
        if cabinet_rows:
            if cabinet_ws["B6"].value != cabinet_rows[0]["var_name"] or cabinet_ws["H6"].value != cabinet_rows[0]["address"]:
                raise AssertionError("classic 单机柜首点不一致")
    if embedded_cabinet_rows:
        first_cabinet_row = combined_start_row + len(expected_combined_points)
        first_cabinet = embedded_cabinet_rows[0]
        if (
            combined_ws[f"B{first_cabinet_row}"].value != first_cabinet["var_name"]
            or combined_ws[f"{'F' if is_unified else 'H'}{first_cabinet_row}"].value != first_cabinet["address"]
        ):
            raise AssertionError("classic 合并页单机柜首点不一致")

    if alarm_ws is None:
        return
    if is_liquidcool:
        if alarm_ws["B9"].value != "ALM_VH" or alarm_ws["H9"].value != renderer.address_profile["alarm_base"]:
            raise AssertionError("liquidcool 报警阈值首行不一致")
    else:
        expected_alarm_first = renderer._build_alarm_rows()[0] if renderer._build_alarm_rows() else None
        alarm_spec = renderer._alarm_sheet_spec()
        header_cell = f"A{alarm_spec.header_row}"
        legend_cell = f"{'I' if is_unified else 'J'}{alarm_spec.intro_row}"
        if alarm_ws[header_cell].value != "通道号" or alarm_ws[legend_cell].value != "各Bit位含义":
            raise AssertionError("classic 报警页头部布局不正确")
        alarm_first_row = alarm_spec.data_start_row
        alarm_address_column = "F" if is_unified else "G"
        if expected_alarm_first and (alarm_ws[f"B{alarm_first_row}"].value != expected_alarm_first.var_name or alarm_ws[f"{alarm_address_column}{alarm_first_row}"].value != expected_alarm_first.register_address):
            raise AssertionError("classic 报警首行不一致")

    if is_unified:
        for ws in wb.worksheets:
            header_values = {
                str(ws.cell(row=row_index, column=column_index).value or "")
                for row_index in range(1, min(ws.max_row, 20) + 1)
                for column_index in range(1, ws.max_column + 1)
            }
            if {"变量类型", "通道名称"} & header_values:
                raise AssertionError(f"{ws.title} 仍包含已删除的变量类型/通道名称列")


def _validate_extended(model: dict[str, Any], renderer: ExtendedSplitRenderer, wb) -> None:
    expected_sheets = ["始端箱", "插接箱", "单机柜数据", renderer._repeater_sheet_name(), "报警状态"]
    if wb.sheetnames != expected_sheets:
        raise AssertionError(f"extended sheet 顺序/名称不一致: {wb.sheetnames} != {expected_sheets}")

    start_points = _flatten_start_points(model)
    plug_points = _flatten_plug_points(model)
    repeater_points = _flatten_repeater_points(model)
    cabinet_rows = model.get("single_cabinet_rows", [])
    alarm_rows = renderer._build_alarm_rows()

    start_ws = wb["始端箱"]
    plug_ws = wb["插接箱"]
    cabinet_ws = wb["单机柜数据"]
    repeater_ws = wb[renderer._repeater_sheet_name()]
    alarm_ws = wb["报警状态"]

    if len(_non_empty_values(start_ws, 2, 12)) != len(start_points):
        raise AssertionError("extended 始端箱行数不一致")
    if len(_non_empty_values(plug_ws, 2, 12)) != len(plug_points):
        raise AssertionError("extended 插接箱行数不一致")
    if len(_non_empty_values(cabinet_ws, 2, 12)) != len(cabinet_rows):
        raise AssertionError("extended 单机柜行数不一致")
    if len(_non_empty_values(repeater_ws, 2, 12)) != len(repeater_points):
        raise AssertionError("extended 中继行数不一致")
    if len(_non_empty_values(alarm_ws, 2, 9)) != len(alarm_rows):
        raise AssertionError("extended 报警行数不一致")
    if start_points:
        if start_ws["B12"].value != start_points[0]["var_name"] or start_ws["G12"].value != start_points[0]["address"]:
            raise AssertionError("extended 始端箱首点不一致")
    if plug_points:
        if plug_ws["B12"].value != plug_points[0]["var_name"] or plug_ws["G12"].value != plug_points[0]["address"]:
            raise AssertionError("extended 插接箱首点不一致")


def _validate_ab(model: dict[str, Any], renderer: AbScreenSplitRenderer, wb) -> None:
    expected_sheets = ["A路屏数据", "A路屏报警", "B路屏数据", "B路屏报警"]
    if wb.sheetnames != expected_sheets:
        raise AssertionError(f"ab sheet 顺序/名称不一致: {wb.sheetnames} != {expected_sheets}")

    for route in ("A", "B"):
        route_model = renderer._route_model(route)
        expected_points = _expected_ab_points(renderer, route_model)
        expected_alarm_rows = renderer._build_screen_alarm_rows(route_model)
        data_ws = wb[f"{route}路屏数据"]
        alarm_ws = wb[f"{route}路屏报警"]
        if len(_non_empty_values(data_ws, 2, 12)) != len(expected_points):
            raise AssertionError(f"{route} 路屏数据行数不一致")
        if len(_non_empty_values(alarm_ws, 2, 9)) != len(expected_alarm_rows):
            raise AssertionError(f"{route} 路屏报警行数不一致")
        if expected_points:
            if data_ws["B12"].value != expected_points[0]["var_name"] or data_ws["G12"].value != expected_points[0]["address"]:
                raise AssertionError(f"{route} 路屏数据首点不一致")
        if expected_alarm_rows:
            if alarm_ws["B9"].value != expected_alarm_rows[0].var_name or alarm_ws["G9"].value != expected_alarm_rows[0].register_address:
                raise AssertionError(f"{route} 路屏报警首行不一致")


def validate_model_and_workbook(model: dict[str, Any], excel_path: str | Path) -> dict[str, Any]:
    renderer = _build_renderer(model)
    excel_path = Path(excel_path)
    _validate_excel_package_structure(excel_path)
    wb = load_workbook(excel_path, data_only=True)
    family = model["profiles"]["export_profile"]["family"]
    try:
        if family == "classic_combined":
            _validate_classic(model, renderer, wb)
        elif family == "extended_split":
            _validate_extended(model, renderer, wb)
        elif family == "ab_screen_split":
            _validate_ab(model, renderer, wb)
        else:
            raise NotImplementedError(f"未支持的 family: {family}")
    finally:
        wb.close()

    return {
        "ok": True,
        "family": family,
        "excel_path": str(excel_path.resolve()),
        "sheet_names": list(renderer.export_profile.get("sheet_order", [])),
    }


def validate_rendered_workbook_pair(json_path: str | Path, excel_path: str | Path) -> dict[str, Any]:
    json_path = Path(json_path)
    model = json.loads(json_path.read_text(encoding="utf-8"))
    result = validate_model_and_workbook(model, excel_path)
    result["json_path"] = str(json_path.resolve())
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="校验 canonical JSON 与导出 Excel 是否一致")
    parser.add_argument("--json", required=True, help="canonical JSON 路径")
    parser.add_argument("--excel", required=True, help="Excel 路径")
    args = parser.parse_args()

    validate_rendered_workbook_pair(args.json, args.excel)
    print("validation_ok")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
