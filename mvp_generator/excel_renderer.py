from __future__ import annotations

from copy import copy, deepcopy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.views import Selection


THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
NO_FILL = PatternFill(fill_type=None)
FILL_BLUE = PatternFill("solid", fgColor="FF8DB4E2")
FILL_ORANGE = PatternFill("solid", fgColor="FFF79646")
FILL_INTRO = PatternFill("solid", fgColor="FFFFC000")
FILL_YELLOW = PatternFill("solid", fgColor="FFFFFF00")
FILL_CLASSIC_PLUG = PatternFill("solid", fgColor="FFD8E4BC")
FILL_LIQUIDCOOL_PLUG = PatternFill("solid", fgColor="FFEBF1DE")
FILL_LIQUIDCOOL_CABINET_IA = PatternFill("solid", fgColor="FF92D050")
FILL_LIQUIDCOOL_CABINET_PA = PatternFill(fill_type="solid", fgColor=Color(theme=9))
FILL_LIQUIDCOOL_CABINET_EA = PatternFill(fill_type="solid", fgColor=Color(theme=8))
FILL_UNIFIED_ROUTE_A = PatternFill("solid", fgColor="FFDDEBF7")
FILL_UNIFIED_ROUTE_B = PatternFill("solid", fgColor="FFFCE4D6")
FILL_UNIFIED_CABINET_IA = PatternFill("solid", fgColor="FFE2F0D9")
FILL_UNIFIED_CABINET_PA = PatternFill("solid", fgColor="FFD9EAF7")
FILL_UNIFIED_CABINET_EA = PatternFill("solid", fgColor="FFFFE5CC")
FILL_UNIFIED_CABINET_KA = PatternFill("solid", fgColor="FFFFF2CC")
CANONICAL_TOPOLOGY_METADATA_PREFIX = "__MCGS_CANONICAL_TOPOLOGY__"
FILL_WHITE = PatternFill("solid", fgColor="FFFFFFFF")
FILL_BLACK = PatternFill("solid", fgColor="FF000000")
FILL_THEME0 = PatternFill(fill_type="solid", fgColor=Color(theme=0))
FILL_THEME3 = PatternFill(fill_type="solid", fgColor=Color(theme=3))
FONT_COLOR_INDEXED_BLACK = Color(indexed=8)
FONT_COLOR_RGB_BLACK = Color(rgb="FF000000")
FONT_COLOR_RGB_RED = Color(rgb="FFFF0000")
BODY_FONT = Font(name="宋体", size=11)
BOLD_FONT = Font(name="宋体", size=11, bold=True)
PHASE_LABELS = ("A", "B", "C")
STATE_EXPLANATION_LINES = ["断路器状态字说明", "BIT0", "BIT1", "BIT2"]
TOPOLOGY_MODE_LABELS = {
    "single_screen_one_column": "一屏一列母线",
    "single_screen_two_columns": "一屏两列母线",
    "dual_screens_ab_separated": "A/B 路分屏",
    "single_screen_half_channel": "单屏半列通道",
}


PREFIX_META = {
    "StateS": ("断路器状态字", None),
    "StateC": ("断路器状态字", None),
    "LoadS": ("总负载率", "%"),
    "UBS": ("三相不平衡度", "%"),
    "Loada": ("A相负载率", "%"),
    "Loadb": ("B相负载率", "%"),
    "Loadc": ("C相负载率", "%"),
    "Load": ("总负载率", "%"),
    "Ia": ("A相电流", "A"),
    "Ib": ("B相电流", "A"),
    "Ic": ("C相电流", "A"),
    "In": ("零序漏电流", "A"),
    "Ua": ("A相电压", "V"),
    "Ub": ("B相电压", "V"),
    "Uc": ("C相电压", "V"),
    "F": ("频率", "Hz"),
    "Pa": ("A相功率", "KW"),
    "Pb": ("B相功率", "KW"),
    "Pc": ("C相功率", "KW"),
    "P": ("总有功功率", "KW"),
    "Qa": ("A相无功功率", "kVar"),
    "Qb": ("B相无功功率", "kVar"),
    "Qc": ("C相无功功率", "kVar"),
    "Q": ("总无功功率", "kVar"),
    "Sa": ("A相视在功率", "kVA"),
    "Sb": ("B相视在功率", "kVA"),
    "Sc": ("C相视在功率", "kVA"),
    "S_": ("总视在功率", "kVA"),
    "PFa": ("A相功率因数", None),
    "PFb": ("B相功率因数", None),
    "PFc": ("C相功率因数", None),
    "PF": ("总功率因数", None),
    "Ta": ("入线接点温度A", "℃"),
    "Tb": ("入线接点温度B", "℃"),
    "Tc": ("入线接点温度C", "℃"),
    "Tn": ("入线接点温度N", "℃"),
    "TaO": ("出线接点温度A", "℃"),
    "TbO": ("出线接点温度B", "℃"),
    "TcO": ("出线接点温度C", "℃"),
    "TnO": ("出线接点温度N", "℃"),
    "TaD": ("连接器接点温度A", "℃"),
    "TbD": ("连接器接点温度B", "℃"),
    "TcD": ("连接器接点温度C", "℃"),
    "TnD": ("连接器接点温度N", "℃"),
    "TaZ": ("连接点温度A", "℃"),
    "TbZ": ("连接点温度B", "℃"),
    "TcZ": ("连接点温度C", "℃"),
    "TnZ": ("连接点温度N", "℃"),
    "Th": ("箱内环境温度", "℃"),
    "THDUa": ("A相电压谐波畸变率", "%"),
    "THDUb": ("B相电压谐波畸变率", "%"),
    "THDUc": ("C相电压谐波畸变率", "%"),
    "THDIa": ("A相电流谐波畸变率", "%"),
    "THDIb": ("B相电流谐波畸变率", "%"),
    "THDIc": ("C相电流谐波畸变率", "%"),
    "THD": ("总谐波畸变率", "%"),
    "Ea": ("A相有功电量", "KWH"),
    "Eb": ("B相有功电量", "KWH"),
    "Ec": ("C相有功电量", "KWH"),
    "E": ("总有功电量", "KWH"),
    "Eqa": ("A相无功电量", "kVarh"),
    "Eqb": ("B相无功电量", "kVarh"),
    "Eqc": ("C相无功电量", "kVarh"),
    "Eq": ("总无功电量", "kVarh"),
}

LIQUIDCOOL_THRESHOLD_ROWS = [
    ("ALM_VH", "电压上限", "V"),
    ("ALM_VL", "电压下限", "V"),
    ("ALM_FH", "频率上限", "Hz"),
    ("ALM_FL", "频率下限", "Hz"),
    ("ALM_THDH", "谐波上限", "%"),
    ("ALM_InH", "零序漏电流上限", "A"),
    ("ALM_IH", "始端箱电流上限", "A"),
    ("ALM_IH1", "32A插接箱电流上限", "A"),
    ("ALM_IH2", "63A插接箱电流上限", "A"),
    ("ALM_PH", "始端箱功率上限", "KW"),
    ("ALM_PH1", "32A插接箱功率上限", "KW"),
    ("ALM_PH2", "63A插接箱功率上限", "KW"),
    ("ALM_PFL", "功率因数下限", None),
    ("ALM_TSDX", "始端箱温度上限", "℃"),
    ("ALM_TCJX", "插接箱温度上限", "℃"),
    ("ALM_TLSDX", "始端箱温度下限", "℃"),
    ("ALM_TLCJX", "插接箱温度下限", "℃"),
]

CLASSIC_COMBINED_COLUMN_WIDTHS = {
    "A": 8.9062,
    "B": 14.4531,
    "C": 19.3633,
    "D": 12.0898,
    "E": 8.9062,
    "F": 17.3633,
    "G": 19.1797,
    "H": 12.9062,
    "I": 21.9062,
    "J": 8.9062,
    "K": 26.8164,
    "L": 28.0,
    "M": 17.0898,
    "N": 32.5430,
    "O": 8.9062,
}

UNIFIED_COMBINED_COLUMN_WIDTHS = {
    **CLASSIC_COMBINED_COLUMN_WIDTHS,
    "K": 20.0,
    "L": 24.0,
    "M": 11.0,
    "N": 18.0,
    "O": 18.0,
    "P": 11.0,
    "Q": 2.0,
}

CLASSIC_REPEATER_COLUMN_WIDTHS = {
    "A": 8.9062,
    "B": 14.4531,
    "C": 19.3633,
    "D": 12.0898,
    "E": 8.9062,
    "F": 17.3633,
    "G": 19.1797,
    "H": 13.8164,
    "I": 21.9062,
    "J": 8.9062,
    "K": 18.6328,
    "L": 8.9062,
    "M": 13.0,
    "N": 8.9062,
    "O": 13.0,
}

CLASSIC_ALARM_COLUMN_WIDTHS = {
    "A": 9.0,
    "B": 14.4531,
    "C": 13.8164,
    "D": 13.0,
    "E": 17.3633,
    "F": 19.1797,
    "G": 13.8164,
    "H": 32.9062,
    "I": 9.0,
    "J": 51.4531,
    "K": 9.0,
    "L": 13.0,
    "M": 9.0,
    "N": 13.0,
    "O": 13.0,
}

LIQUIDCOOL_COMBINED_COLUMN_WIDTHS = {
    "A": 8.91,
    "B": 14.45,
    "C": 19.36,
    "D": 12.09,
    "E": 8.91,
    "F": 17.36,
    "G": 19.18,
    "H": 10.91,
    "I": 21.91,
    "J": 24.0,
    "K": 24.45,
    "L": 8.91,
    "M": 22.91,
    "N": 19.18,
    "O": 13.0,
}

LIQUIDCOOL_REPEATER_COLUMN_WIDTHS = {
    "A": 8.91,
    "B": 14.45,
    "C": 19.36,
    "D": 12.09,
    "E": 8.91,
    "F": 17.36,
    "G": 19.18,
    "H": 10.91,
    "I": 21.91,
    "J": 8.91,
    "K": 18.18,
    "L": 8.91,
    "M": 13.0,
    "N": 13.0,
}

LIQUIDCOOL_ALARM_COLUMN_WIDTHS = {
    "A": 7.82,
    "B": 14.91,
    "C": 12.82,
    "D": 13.09,
    "E": 9.91,
    "F": 20.73,
    "G": 19.54,
    "H": 13.18,
    "I": 33.36,
    "J": 8.82,
    "K": 44.09,
    "L": 55.63,
    "M": 13.0,
    "N": 8.73,
}

LIQUIDCOOL_CABINET_COLUMN_WIDTHS = {
    "A": 8.73,
    "B": 13.0,
    "C": 13.0,
    "D": 11.45,
    "E": 8.73,
    "F": 19.91,
    "G": 27.63,
    "H": 28.73,
    "I": 37.27,
    "J": 8.73,
    "K": 36.73,
    "L": 38.27,
    "M": 36.45,
    "N": 32.73,
    "O": 8.73,
}

DEFAULT_COLUMN_WIDTHS = {
    "A": 10,
    "B": 20,
    "C": 12,
    "D": 12,
    "E": 10,
    "F": 16,
    "G": 18,
    "H": 24,
    "I": 12,
    "J": 10,
    "K": 24,
    "L": 18,
    "M": 18,
    "N": 18,
}


@dataclass
class AlarmRow:
    channel_no: int | None
    var_name: str
    register_address: int
    description: str | None
    bit_lines: list[str]
    data_type_label: str


@dataclass(frozen=True)
class CombinedSheetSpec:
    headers: tuple[str, ...]
    header_row: int
    data_start_row: int
    note_cell: str | None = None
    note_kind: str | None = None
    extra_cells: tuple[tuple[str, str], ...] = ()


@dataclass(frozen=True)
class AlarmSheetSpec:
    headers: tuple[str, ...]
    header_row: int
    intro_row: int
    data_start_row: int
    legend_column: int = 10


@dataclass(frozen=True)
class CabinetSheetSpec:
    headers: tuple[str, ...]
    header_row: int
    data_start_row: int
    include_screen_address_note: bool
    label_column_ids: tuple[str, ...] = ()


@dataclass(frozen=True)
class SheetLayoutSpec:
    merge_ranges: tuple[str, ...] = ()
    row_heights: tuple[tuple[int, float], ...] = ()
    left_wrap_cells: tuple[str, ...] = ()
    fill_cells: tuple[tuple[str, PatternFill], ...] = ()


@dataclass(frozen=True)
class SheetStyleSpec:
    column_widths: dict[str, float]


CLASSIC_COMBINED_SHEET_SPECS = {
    "unified_master": CombinedSheetSpec(
        headers=(
            "通道号",
            "变量名",
            "变量类型",
            "通道名称",
            "读写类型",
            "寄存器名称",
            "数据类型",
            "寄存器地址",
            "测点说明",
            "单位",
            "分路",
            "设备",
            "",
            "",
            "",
            "",
        ),
        header_row=10,
        data_start_row=11,
        note_cell="A8",
        note_kind="unified_float_reading",
    ),
    "classic_standard": CombinedSheetSpec(
        headers=("通道号", "变量名", "变量类型", "通道名称", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "单位", "", "", ""),
        header_row=10,
        data_start_row=11,
        note_cell="A8",
        note_kind="classic_intro",
    ),
    "classic_two_columns": CombinedSheetSpec(
        headers=("通道号", "变量名", "变量类型", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "单位", "", "", "", ""),
        header_row=11,
        data_start_row=12,
        note_cell="A9",
        note_kind="classic_intro",
    ),
    "classic_liquidcool": CombinedSheetSpec(
        headers=("通道号", "变量名", "变量类型", "通道名称", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "J01列", "J02列", "", "", "", ""),
        header_row=12,
        data_start_row=13,
        note_cell="A9",
        note_kind="liquidcool_combined",
        extra_cells=(("J11", "接线布局"),),
    ),
}

CLASSIC_ALARM_SHEET_SPECS = {
    "unified_master": AlarmSheetSpec(
        headers=("通道号", "变量名", "变量类型", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "报警类别", "单位", "各Bit位含义"),
        header_row=6,
        intro_row=7,
        data_start_row=8,
    ),
    "classic_standard": AlarmSheetSpec(
        headers=("通道号", "变量名", "变量类型", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "单位", ""),
        header_row=5,
        intro_row=6,
        data_start_row=7,
    ),
    "classic_two_columns": AlarmSheetSpec(
        headers=("通道号", "变量名", "变量类型", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "单位", ""),
        header_row=7,
        intro_row=8,
        data_start_row=9,
    ),
}

CABINET_SHEET_SPECS = {
    "classic_standard": CabinetSheetSpec(
        headers=("通道号", "变量名", "变量类型", "通道名称", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "单位"),
        header_row=5,
        data_start_row=6,
        include_screen_address_note=True,
    ),
    "classic_liquidcool": CabinetSheetSpec(
        headers=("通道号", "变量名", "变量类型", "通道名称", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "单位", "J01列", "J02列", "", ""),
        header_row=5,
        data_start_row=6,
        include_screen_address_note=False,
        label_column_ids=("J01", "J02"),
    ),
}


CLASSIC_COMBINED_LAYOUT_SPECS = {
    "unified_master": SheetLayoutSpec(
        merge_ranges=("A1:I1", "A2:I2", "A3:I3", "A4:I4", "A5:I7", "A8:I9"),
        row_heights=((1, 22), (2, 20), (3, 20), (4, 20), (5, 24), (6, 24), (7, 24), (8, 28), (9, 28), (10, 24)),
        left_wrap_cells=("A1", "A2", "A3", "A4", "A5", "A8"),
        fill_cells=(("A5", FILL_INTRO),),
    ),
    "classic_standard": SheetLayoutSpec(
        merge_ranges=("A1:I2", "A3:I4", "A5:G7", "A8:G9"),
        row_heights=((1, 20), (2, 20), (3, 20), (4, 20), (5, 26), (6, 26), (7, 26), (8, 20), (9, 20), (10, 23)),
        left_wrap_cells=("A5", "A8"),
        fill_cells=(("A5", FILL_INTRO), ("L11", FILL_YELLOW)),
    ),
    "classic_two_columns": SheetLayoutSpec(
        merge_ranges=("A1:H2", "A3:H4", "A5:H8", "A9:H10"),
        row_heights=((1, 20), (2, 20), (3, 20), (4, 20), (5, 24), (6, 24), (7, 24), (8, 24), (9, 20), (10, 20), (11, 23)),
        left_wrap_cells=("A5", "A9"),
        fill_cells=(("A5", FILL_INTRO), ("K12", FILL_YELLOW)),
    ),
    "classic_liquidcool": SheetLayoutSpec(
        merge_ranges=("A1:I2", "A3:I4", "A5:G8", "A9:G10", "J11:K11"),
        row_heights=((1, 14), (2, 15), (3, 14), (4, 15), (5, 14), (6, 15), (7, 15), (8, 15), (9, 28), (10, 22), (11, 25), (12, 15), (13, 14)),
        left_wrap_cells=("A5", "A9"),
        fill_cells=(("A5", FILL_INTRO), ("M13", FILL_YELLOW)),
    ),
}

CLASSIC_AUX_LAYOUT_SPECS = {
    ("repeater", "unified_master"): SheetLayoutSpec(
        merge_ranges=("A1:I1", "A2:I2", "A3:I3", "A4:I4"),
        row_heights=((1, 22), (2, 20), (3, 20), (4, 20), (5, 15), (6, 24)),
        left_wrap_cells=("A1", "A2", "A3", "A4"),
    ),
    ("repeater", "classic_standard"): SheetLayoutSpec(
        merge_ranges=("A1:I2", "A3:I4"),
        row_heights=((1, 20), (2, 20), (3, 20), (4, 20), (5, 14.4), (6, 23)),
    ),
    ("repeater", "classic_liquidcool"): SheetLayoutSpec(
        merge_ranges=("A1:G1", "A2:G2", "A3:G3"),
        row_heights=((1, 15), (2, 15), (3, 15), (4, 15), (5, 15), (6, 15)),
    ),
    ("alarm", "classic_standard"): SheetLayoutSpec(
        merge_ranges=("A1:H2", "A3:H4", "A6:I6"),
        row_heights=((1, 21), (2, 21), (3, 21), (4, 21), (5, 20), (6, 30)),
    ),
    ("alarm", "unified_master"): SheetLayoutSpec(
        merge_ranges=("A1:I1", "A2:I2", "A3:I3", "A4:I4", "A7:I7"),
        row_heights=((1, 22), (2, 20), (3, 20), (4, 20), (5, 15), (6, 24), (7, 30)),
        left_wrap_cells=("A1", "A2", "A3", "A4"),
    ),
    ("alarm", "classic_two_columns"): SheetLayoutSpec(
        merge_ranges=("A1:H2", "A3:H4", "A8:I8"),
        row_heights=((1, 21), (2, 21), (3, 21), (4, 21), (5, 18), (6, 18), (7, 20), (8, 30)),
    ),
    ("alarm", "classic_liquidcool"): SheetLayoutSpec(
        merge_ranges=("A1:I2", "A3:I4", "A5:G6", "A8:J8"),
        row_heights=((7, 13.5), (8, 13.5)),
        fill_cells=(("A5", FILL_INTRO),),
    ),
    ("cabinet", "classic_standard"): SheetLayoutSpec(
        merge_ranges=("A2:G2", "A3:G3"),
    ),
    ("cabinet", "classic_liquidcool"): SheetLayoutSpec(
        merge_ranges=("A1:G1", "A2:G2", "A3:G3"),
        row_heights=((1, 15), (2, 15), (3, 15), (4, 15), (5, 15)),
    ),
}


CLASSIC_SHEET_STYLE_SPECS = {
    ("combined", "unified_master"): SheetStyleSpec(column_widths=UNIFIED_COMBINED_COLUMN_WIDTHS),
    ("combined", "classic_standard"): SheetStyleSpec(column_widths=CLASSIC_COMBINED_COLUMN_WIDTHS),
    ("combined", "classic_two_columns"): SheetStyleSpec(column_widths=CLASSIC_COMBINED_COLUMN_WIDTHS),
    ("combined", "classic_liquidcool"): SheetStyleSpec(column_widths=LIQUIDCOOL_COMBINED_COLUMN_WIDTHS),
    ("repeater", "classic_standard"): SheetStyleSpec(column_widths=CLASSIC_REPEATER_COLUMN_WIDTHS),
    ("repeater", "unified_master"): SheetStyleSpec(column_widths=CLASSIC_REPEATER_COLUMN_WIDTHS),
    ("repeater", "classic_two_columns"): SheetStyleSpec(column_widths=CLASSIC_REPEATER_COLUMN_WIDTHS),
    ("repeater", "classic_liquidcool"): SheetStyleSpec(column_widths=LIQUIDCOOL_REPEATER_COLUMN_WIDTHS),
    ("alarm", "classic_standard"): SheetStyleSpec(column_widths=CLASSIC_ALARM_COLUMN_WIDTHS),
    ("alarm", "unified_master"): SheetStyleSpec(column_widths=CLASSIC_ALARM_COLUMN_WIDTHS),
    ("alarm", "classic_two_columns"): SheetStyleSpec(column_widths=CLASSIC_ALARM_COLUMN_WIDTHS),
    ("alarm", "classic_liquidcool"): SheetStyleSpec(column_widths=LIQUIDCOOL_ALARM_COLUMN_WIDTHS),
    ("cabinet", "classic_standard"): SheetStyleSpec(column_widths=DEFAULT_COLUMN_WIDTHS),
    ("cabinet", "classic_two_columns"): SheetStyleSpec(column_widths=DEFAULT_COLUMN_WIDTHS),
    ("cabinet", "classic_liquidcool"): SheetStyleSpec(column_widths=LIQUIDCOOL_CABINET_COLUMN_WIDTHS),
}


def build_alarm_rows_from_group_specs(
    address_profile: dict[str, Any],
    group_specs: list[tuple[str, list[tuple[str, list[str]]]]],
) -> list[AlarmRow]:
    chunk_size = 32 if address_profile.get("alarm_word_mode") == "32bit" else 16
    register_size = 2 if address_profile.get("alarm_word_mode") == "32bit" else 1
    data_type_label = "32位 无符号二进制" if register_size == 2 else "16位 无符号二进制"
    address = address_profile["alarm_base"]
    channel_no = 0
    category_counter: dict[str, int] = {}
    rows: list[AlarmRow] = []

    for category, scoped_groups in group_specs:
        category_counter.setdefault(category, 1)
        for description, bits in scoped_groups:
            if not bits:
                continue
            chunks = [bits[i : i + chunk_size] for i in range(0, len(bits), chunk_size)]
            first = True
            for chunk in chunks:
                rows.append(
                    AlarmRow(
                        channel_no=channel_no if first else None,
                        var_name=f"State_{category}{category_counter[category]}",
                        register_address=address,
                        description=description if first else None,
                        bit_lines=[f"Bit{idx}： {text}" for idx, text in enumerate(chunk)],
                        data_type_label=data_type_label,
                    )
                )
                first = False
                category_counter[category] += 1
                address += register_size
            channel_no += 1
    return rows


def temperature_phase_label(prefix: str) -> str | None:
    if prefix.endswith("a") or prefix.endswith("A"):
        return "A相"
    if prefix.endswith("b") or prefix.endswith("B"):
        return "B相"
    if prefix.endswith("c") or prefix.endswith("C"):
        return "C相"
    if prefix.endswith("n") or prefix.endswith("N"):
        return "N相"
    return None


def fill_cell(cell, fill: PatternFill | None) -> None:
    if fill is not None:
        cell.fill = fill


def set_fill_range(ws, start_row: int, end_row: int, start_col: int, end_col: int, fill: PatternFill | None) -> None:
    if fill is None:
        return
    for row_index in range(start_row, end_row + 1):
        for col_index in range(start_col, end_col + 1):
            ws.cell(row=row_index, column=col_index).fill = fill


def normalize_metric_code(value: Any) -> str:
    text = str(value or "").upper()
    for metric in ("IA", "PA", "EA", "KA"):
        if text.startswith(metric):
            return metric
    return text[:2]


def merge_with_alignment(ws, cell_range: str, *, horizontal: str = "center", vertical: str = "center") -> None:
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(":")[0]]
    top_left.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=True)


def text_display_line_count(value: Any) -> int:
    if value in (None, ""):
        return 1
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return max(1, text.count("\n") + 1)


def explicit_border(
    *,
    left: Side | None = None,
    right: Side | None = None,
    top: Side | None = None,
    bottom: Side | None = None,
) -> Border:
    return Border(
        left=deepcopy(left or Side()),
        right=deepcopy(right or Side()),
        top=deepcopy(top or Side()),
        bottom=deepcopy(bottom or Side()),
    )


def clear_workbook_freeze_panes(workbook) -> None:
    """Remove freeze/split pane XML and leave one valid A1 selection per sheet."""

    for ws in workbook.worksheets:
        ws.freeze_panes = None
        ws.sheet_view.pane = None
        ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]


class ClassicCombinedRenderer:
    def __init__(self, model: dict[str, Any]):
        self.model = model
        self.project = model["project"]
        self.communication = model.get("communication", {})
        self.address_profile = model["profiles"]["address_profile"]
        self.export_profile = model["profiles"]["export_profile"]
        self.protocol_layout = model.get("protocol_layout", {})
        self.warnings = model.get("warnings", [])

    def render_to_path(self, output_path: Path) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        combined_ws = workbook.create_sheet(self._base_sheet_name())
        self._render_combined_sheet(combined_ws)

        if self._should_render_repeater_sheet():
            repeater_ws = workbook.create_sheet(self._repeater_sheet_name())
            self._render_repeater_sheet(repeater_ws)

        if self._should_render_alarm_sheet():
            alarm_ws = workbook.create_sheet("报警状态")
            if self._is_liquidcool_profile():
                self._render_liquidcool_alarm_sheet(alarm_ws)
            else:
                self._render_alarm_sheet(alarm_ws)

        if self._should_render_single_cabinet_sheet():
            cabinet_ws = workbook.create_sheet("单机柜数据")
            self._render_single_cabinet_sheet(cabinet_ws)

        for ws in workbook.worksheets:
            self._apply_common_style(ws)

        self._finalize_unified_workbook(workbook)

        clear_workbook_freeze_panes(workbook)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    @staticmethod
    def _delete_column_preserving_layout(ws, column_index: int) -> None:
        merged_ranges = [
            (
                cell_range.min_col,
                cell_range.min_row,
                cell_range.max_col,
                cell_range.max_row,
            )
            for cell_range in ws.merged_cells.ranges
        ]
        for cell_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(cell_range))

        column_dimensions: dict[int, dict[str, Any]] = {}
        for column_letter, dimension in list(ws.column_dimensions.items()):
            try:
                source_index = column_index_from_string(column_letter)
            except ValueError:
                continue
            column_dimensions[source_index] = {
                "width": dimension.width,
                "hidden": dimension.hidden,
                "outline_level": dimension.outline_level,
                "collapsed": dimension.collapsed,
                "best_fit": dimension.bestFit,
            }

        ws.delete_cols(column_index, 1)

        for column_letter in list(ws.column_dimensions.keys()):
            del ws.column_dimensions[column_letter]
        for source_index, state in sorted(column_dimensions.items()):
            if source_index == column_index:
                continue
            target_index = source_index - 1 if source_index > column_index else source_index
            dimension = ws.column_dimensions[get_column_letter(target_index)]
            dimension.width = state["width"]
            dimension.hidden = state["hidden"]
            dimension.outline_level = state["outline_level"]
            dimension.collapsed = state["collapsed"]
            dimension.bestFit = state["best_fit"]

        for min_col, min_row, max_col, max_row in merged_ranges:
            if max_col < column_index:
                shifted_min_col = min_col
                shifted_max_col = max_col
            elif min_col > column_index:
                shifted_min_col = min_col - 1
                shifted_max_col = max_col - 1
            else:
                shifted_min_col = min_col
                shifted_max_col = max_col - 1
            if shifted_max_col < shifted_min_col:
                continue
            ws.merge_cells(
                start_row=min_row,
                start_column=shifted_min_col,
                end_row=max_row,
                end_column=shifted_max_col,
            )

    def _finalize_unified_workbook(self, workbook) -> None:
        """Apply the user-facing unified protocol contract after styling.

        The renderer keeps its established internal coordinates while composing
        complex blocks and state-word notes.  The final pass removes the two
        obsolete UI-oriented columns by their header names, shifts merges and
        widths together, hides project identity rows, and preserves hidden
        topology metadata for code generation.
        """

        if not self._is_unified_master():
            return

        obsolete_headers = {"变量类型", "通道名称"}
        for ws in workbook.worksheets:
            if str(ws["A1"].value or "").startswith("项目名称："):
                ws.row_dimensions[1].hidden = True
            if str(ws["A2"].value or "").startswith("项目编号："):
                ws.row_dimensions[2].hidden = True

            obsolete_columns: set[int] = set()
            for row_index in range(1, min(ws.max_row, 20) + 1):
                for column_index in range(1, ws.max_column + 1):
                    if ws.cell(row=row_index, column=column_index).value in obsolete_headers:
                        obsolete_columns.add(column_index)
            for column_index in sorted(obsolete_columns, reverse=True):
                self._delete_column_preserving_layout(ws, column_index)

            if ws.title == self._base_sheet_name():
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value == "断路器状态字":
                            bold_font = copy(cell.font)
                            bold_font.bold = True
                            cell.font = bold_font

            self._normalize_merged_range_borders(ws)

    def _iter_combined_plug_blocks(self) -> Iterable[tuple[str, dict[str, Any], dict[str, Any]]]:
        route_models_by_name = {route_model["route"]: route_model for route_model in self.model["routes"]}
        if not self._is_liquidcool_profile():
            for route_model in self.model["routes"]:
                for physical_box in route_model["physical_plug_boxes"]:
                    for board in physical_box["boards"]:
                        for branch in board["branches"]:
                            yield route_model["route"], physical_box, branch
            return

        route_boxes = {
            route: route_models_by_name.get(route, {}).get("physical_plug_boxes", [])
            for route in ("A", "B")
        }
        max_box_count = max((len(items) for items in route_boxes.values()), default=0)
        for box_index in range(max_box_count):
            route_branch_lists: dict[str, tuple[dict[str, Any], list[dict[str, Any]]]] = {}
            max_branch_count = 0
            for route in ("A", "B"):
                if box_index >= len(route_boxes[route]):
                    continue
                physical_box = route_boxes[route][box_index]
                ordered_branches = [
                    branch
                    for board in physical_box["boards"]
                    for branch in board["branches"]
                ]
                route_branch_lists[route] = (physical_box, ordered_branches)
                max_branch_count = max(max_branch_count, len(ordered_branches))
            for branch_index in range(max_branch_count):
                for route in ("A", "B"):
                    payload = route_branch_lists.get(route)
                    if not payload:
                        continue
                    physical_box, branches = payload
                    if branch_index < len(branches):
                        yield route, physical_box, branches[branch_index]

    def _render_combined_sheet(self, ws) -> None:
        line1, line2 = self._communication_lines(include_screen_address_note=True)
        intro_lines = self._build_intro_lines()
        spec = self._combined_sheet_spec()
        if self._is_unified_master():
            ws["A1"] = f"项目名称：{self.project.get('project_name') or '未填写'}"
            ws["A2"] = (
                f"项目编号：{self.project.get('project_code') or '未填写'}    "
                f"协议标题：{self.project.get('protocol_title') or '上位机通讯协议'}"
            )
            ws["A3"] = line1
            ws["A4"] = line2
        else:
            ws["A1"] = line1
            ws["A3"] = line2
        ws["A5"] = "\n".join(intro_lines)
        if spec.note_cell and spec.note_kind:
            ws[spec.note_cell] = self._note_text(spec.note_kind)
        for cell_ref, value in spec.extra_cells:
            ws[cell_ref] = value
        self._write_header_row(ws, spec.header_row, list(spec.headers))
        row_no = spec.data_start_row
        data_start_row = spec.data_start_row

        channel_no = 0
        classic_start_box_index = 0
        liquidcool_start_box_index = 0

        for route_model in self.model["routes"]:
            for start_box in route_model["start_boxes"]:
                group_label = (
                    f"{self._route_scope_label(route_model['route'], start_box)}"
                    f"始端箱{start_box['instance_name']}"
                )
                block_start_row = row_no
                row_no, channel_no = self._append_point_block(
                    ws=ws,
                    row_no=row_no,
                    channel_no=channel_no,
                    points=start_box["points"],
                    group_label=group_label,
                    annotation_lines=None,
                    layout_kind="combined",
                )
                self._style_combined_block(
                    ws,
                    start_row=block_start_row,
                    end_row=row_no - 1,
                    route=route_model["route"],
                    block_kind="start",
                    block_index=liquidcool_start_box_index if self._is_liquidcool_profile() else classic_start_box_index,
                    group_label=group_label,
                    write_state_note=(block_start_row == data_start_row),
                )
                if self._is_liquidcool_profile():
                    liquidcool_start_box_index += 1
                else:
                    classic_start_box_index += 1

        if self._is_unified_master():
            for route_model in self.model["routes"]:
                route = route_model["route"]
                for physical_box in route_model["physical_plug_boxes"]:
                    device_start_row = row_no
                    for board in physical_box["boards"]:
                        for branch in board["branches"]:
                            points = branch.get("points", [])
                            if not points:
                                continue
                            branch_label = self._unified_branch_label(
                                route,
                                physical_box,
                                branch,
                            )
                            block_start_row = row_no
                            row_no, channel_no = self._append_point_block(
                                ws=ws,
                                row_no=row_no,
                                channel_no=channel_no,
                                points=points,
                                group_label=branch_label,
                                annotation_lines=None,
                                layout_kind="combined",
                                canonical_topology_metadata=(
                                    self._branch_canonical_topology_metadata(
                                        route,
                                        physical_box,
                                        branch,
                                    )
                                    if self._measurement_layout_mode() == "by_branch"
                                    else None
                                ),
                            )
                            self._style_combined_block(
                                ws,
                                start_row=block_start_row,
                                end_row=row_no - 1,
                                route=route,
                                block_kind="plug",
                                block_index=0,
                                group_label=branch_label,
                                write_state_note=False,
                            )
                    if row_no > device_start_row:
                        self._write_unified_device_label(
                            ws,
                            start_row=device_start_row,
                            end_row=row_no - 1,
                            label=self._unified_device_label(route, physical_box),
                        )
        elif self._measurement_layout_mode() == "by_plug_box":
            for route_model in self.model["routes"]:
                route = route_model["route"]
                for physical_box in route_model["physical_plug_boxes"]:
                    points = [
                        point
                        for board in physical_box["boards"]
                        for branch in board["branches"]
                        for point in branch.get("points", [])
                    ]
                    if not points:
                        continue
                    group_label = self._plug_box_label(route, physical_box)
                    block_start_row = row_no
                    row_no, channel_no = self._append_point_block(
                        ws=ws,
                        row_no=row_no,
                        channel_no=channel_no,
                        points=points,
                        group_label=group_label,
                        annotation_lines=None,
                        layout_kind="combined",
                    )
                    self._style_combined_block(
                        ws,
                        start_row=block_start_row,
                        end_row=row_no - 1,
                        route=route,
                        block_kind="plug",
                        block_index=0,
                        group_label=group_label,
                        write_state_note=False,
                    )
        else:
            for route, physical_box, branch in self._iter_combined_plug_blocks():
                points = branch.get("points", [])
                if not points:
                    continue
                group_label = self._branch_group_label(route, physical_box, branch, include_route=True)
                block_start_row = row_no
                row_no, channel_no = self._append_point_block(
                    ws=ws,
                    row_no=row_no,
                    channel_no=channel_no,
                    points=points,
                    group_label=group_label,
                    annotation_lines=None,
                    layout_kind="combined",
                    canonical_topology_metadata=self._branch_canonical_topology_metadata(
                        route,
                        physical_box,
                        branch,
                    ),
                )
                self._style_combined_block(
                    ws,
                    start_row=block_start_row,
                    end_row=row_no - 1,
                    route=route,
                    block_kind="plug",
                    block_index=0,
                    group_label=group_label,
                    write_state_note=False,
                )

        if self._embed_single_cabinet_in_base_sheet():
            row_no, channel_no = self._append_single_cabinet_rows_to_combined(
                ws,
                row_no=row_no,
                channel_no=channel_no,
            )

        ws.freeze_panes = f"A{data_start_row}"
        if self._is_unified_master():
            # Q is reserved for renderer-to-codegen metadata before the final
            # removal of the obsolete C/D columns. It becomes hidden O in the
            # delivered workbook.
            ws.column_dimensions["Q"].hidden = True
        self._apply_classic_combined_layout(ws)

    def _render_repeater_sheet(self, ws) -> None:
        line1, line2 = self._communication_lines(include_screen_address_note=not self._is_liquidcool_profile())
        if self._is_unified_master():
            ws["A1"] = f"项目名称：{self.project.get('project_name') or '未填写'}"
            ws["A2"] = (
                f"项目编号：{self.project.get('project_code') or '未填写'}    "
                f"协议标题：{self.project.get('protocol_title') or '上位机通讯协议'}"
            )
            ws["A3"] = line1
            ws["A4"] = line2
            headers = ["通道号", "变量名", "变量类型", "通道名称", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "测点说明", "单位", "设备分组"]
        elif self._is_liquidcool_profile():
            ws["A2"] = line1
            ws["A3"] = line2
            headers = ["通道号", "变量名", "变量类型", "通道名称", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "", "", ""]
        else:
            ws["A1"] = line1
            ws["A3"] = line2
            headers = ["通道号", "变量名", "变量类型", "通道名称", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "单位", ""]
        if self._is_liquidcool_profile():
            headers = ["通道号", "变量名", "变量类型", "通道名称", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "", ""]
        self._write_header_row(ws, 6, headers)

        row_no = 7
        channel_no = 0
        repeater_block_index = 0
        if self._is_liquidcool_profile():
            route_models_by_name = {route_model["route"]: route_model for route_model in self.model["routes"]}
            max_repeater_count = max((len(route_model["repeater_units"]) for route_model in self.model["routes"]), default=0)
            repeater_entries: list[tuple[dict[str, Any], dict[str, Any]]] = []
            for slot_index in range(max_repeater_count):
                for route_name in ("A", "B"):
                    route_model = route_models_by_name.get(route_name)
                    if not route_model:
                        continue
                    if slot_index < len(route_model["repeater_units"]):
                        repeater_entries.append((route_model, route_model["repeater_units"][slot_index]))
        else:
            repeater_entries = [
                (route_model, repeater)
                for route_model in self.model["routes"]
                for repeater in route_model["repeater_units"]
            ]
        for route_model, repeater in repeater_entries:
            group_label = self._repeater_entity_label(route_model["route"], repeater)
            block_start_row = row_no
            row_no, channel_no = self._append_point_block(
                ws=ws,
                row_no=row_no,
                channel_no=channel_no,
                points=repeater["points"],
                group_label=group_label,
                annotation_lines=None,
                layout_kind="repeater",
            )
            self._style_repeater_block(
                ws,
                start_row=block_start_row,
                end_row=row_no - 1,
                route=route_model["route"],
                block_index=repeater_block_index,
            )
            repeater_block_index += 1
        ws.freeze_panes = "A7"
        self._apply_classic_aux_layout(ws, kind="repeater")

    def _render_alarm_sheet(self, ws) -> None:
        line1, line2 = self._communication_lines(include_screen_address_note=True)
        spec = self._alarm_sheet_spec()
        if self._is_unified_master():
            ws["A1"] = f"项目名称：{self.project.get('project_name') or '未填写'}"
            ws["A2"] = (
                f"项目编号：{self.project.get('project_code') or '未填写'}    "
                f"协议标题：{self.project.get('protocol_title') or '上位机通讯协议'}"
            )
            ws["A3"] = line1
            ws["A4"] = line2
        else:
            ws["A1"] = line1
            ws["A3"] = line2
        self._write_header_row(ws, spec.header_row, list(spec.headers))
        ws.cell(row=spec.intro_row, column=1, value="下面的数据中各Bit位0代表正常，1代表监控屏产生了对应类型的报警(没用到的Bit位不写出来)")
        ws.cell(row=spec.intro_row, column=spec.legend_column, value="各Bit位含义")
        fill_cell(ws.cell(row=spec.intro_row, column=spec.legend_column), FILL_THEME3)

        merge_start_row = None
        row_no = spec.data_start_row
        for row in self._build_alarm_rows():
            ws.cell(row=row_no, column=1, value=row.channel_no)
            ws.cell(row=row_no, column=2, value=row.var_name)
            ws.cell(row=row_no, column=3, value="SINGLE")
            ws.cell(row=row_no, column=4, value="只读")
            ws.cell(row=row_no, column=5, value="[4区]输出寄存器")
            ws.cell(row=row_no, column=6, value=row.data_type_label)
            ws.cell(row=row_no, column=7, value=row.register_address)
            ws.cell(row=row_no, column=8, value=row.description)
            ws.cell(row=row_no, column=10, value="\n".join(row.bit_lines))
            if row.description:
                if merge_start_row is not None and row_no - 1 > merge_start_row:
                    ws.merge_cells(start_row=merge_start_row, start_column=8, end_row=row_no - 1, end_column=8)
                merge_start_row = row_no
            row_no += 1
        if merge_start_row is not None and row_no - 1 > merge_start_row:
            ws.merge_cells(start_row=merge_start_row, start_column=8, end_row=row_no - 1, end_column=8)
        self._style_classic_alarm_rows(ws, spec.data_start_row, row_no - 1)
        ws.freeze_panes = f"A{spec.data_start_row}"
        self._apply_classic_aux_layout(ws, kind="alarm")

    def _render_liquidcool_alarm_sheet(self, ws) -> None:
        line1, line2 = self._communication_lines(include_screen_address_note=True)
        ws["A1"] = line1
        ws["A3"] = line2
        ws["A5"] = (
            "注：以下所有的数据都是32位数据存放于两个寄存器中，除了最上面的报警阈值外，其它报警状态字都是小寄存器存中放的低字节，"
            "例如寄存器6050和6051这个数据中，6050存放的是低字节Bit0~15"
        )
        headers = ["通道号", "变量名", "变量类型", "通道名称", "读写类型", "寄存器名称", "数据类型", "寄存器地址", "", "单位", "", "", ""]
        self._write_header_row(ws, 7, headers)
        ws["A8"] = "监控屏上默认设置的各类报警阈值"

        row_no = 9
        for channel_no, (var_name, description, unit) in enumerate(LIQUIDCOOL_THRESHOLD_ROWS):
            ws.cell(row=row_no, column=1, value=channel_no)
            ws.cell(row=row_no, column=2, value=var_name)
            ws.cell(row=row_no, column=3, value="SINGLE")
            ws.cell(row=row_no, column=4, value="只读4DF")
            ws.cell(row=row_no, column=5, value="只读")
            ws.cell(row=row_no, column=6, value="[4区]输出寄存器")
            ws.cell(row=row_no, column=7, value="32位 浮点数")
            ws.cell(row=row_no, column=8, value=self.address_profile["alarm_base"] + channel_no * 2)
            ws.cell(row=row_no, column=9, value=description)
            ws.cell(row=row_no, column=10, value=unit)
            row_no += 1

        ws.cell(row=row_no, column=1, value="下面的数据中各Bit位0代表正常，1代表监控屏产生了对应类型的报警(没用到的Bit位不写出来)")
        ws.cell(row=row_no, column=11, value="各Bit位含义")
        merge_with_alignment(ws, f"A{row_no}:J{row_no}")
        row_no += 1

        state_rows = self._build_liquidcool_state_alarm_rows()
        base_channel_no = len(LIQUIDCOOL_THRESHOLD_ROWS)
        for index, row in enumerate(state_rows, start=1):
            pair_mode = index > 10
            group_start_row = row_no
            ws.cell(row=row_no, column=1, value=base_channel_no + index - 1)
            ws.cell(row=row_no, column=2, value=row.var_name)
            ws.cell(row=row_no, column=3, value="SINGLE")
            ws.cell(row=row_no, column=4, value="只读4DUB" if "32位" in row.data_type_label else "只读4WUB")
            ws.cell(row=row_no, column=5, value="只读")
            ws.cell(row=row_no, column=6, value="[4区]输出寄存器")
            ws.cell(row=row_no, column=7, value=row.data_type_label)
            ws.cell(row=row_no, column=8, value=row.register_address)
            ws.cell(row=row_no, column=9, value=row.description)
            if index <= 4:
                ws.cell(row=row_no, column=11, value="\n".join(row.bit_lines))
            else:
                ws.cell(row=row_no, column=11, value="\n".join(row.bit_lines))
                if index == 5:
                    ws.cell(row=row_no, column=12, value="报警对应机柜")
                    ws.merge_cells(start_row=row_no, start_column=12, end_row=row_no, end_column=13)
                elif index == 6:
                    ws.cell(row=row_no, column=12, value="J01列")
                    ws.cell(row=row_no, column=13, value="J02列")
                elif index >= 7:
                    ws.cell(row=row_no, column=12, value="\n".join(row.bit_lines))
                    ws.cell(row=row_no, column=13, value="\n".join(row.bit_lines))
            if pair_mode:
                ws.cell(row=row_no + 1, column=8, value=row.register_address + 1)
                for column_index in (1, 2, 3, 4, 5, 6, 7, 9, 10, 11, 12, 13):
                    ws.merge_cells(
                        start_row=group_start_row,
                        start_column=column_index,
                        end_row=group_start_row + 1,
                        end_column=column_index,
                    )
                row_no += 2
            else:
                row_no += 1
        self._style_liquidcool_alarm_rows(ws, 9, row_no - 1)
        ws.freeze_panes = "A9"
        self._apply_classic_aux_layout(ws, kind="alarm")

    def _render_single_cabinet_sheet(self, ws) -> None:
        spec = self._cabinet_sheet_spec()
        line1, line2 = self._communication_lines(include_screen_address_note=spec.include_screen_address_note)
        ws["A2"] = line1
        ws["A3"] = line2
        self._write_header_row(ws, spec.header_row, list(spec.headers))

        row_no = spec.data_start_row
        for channel_no, item in enumerate(self.model.get("single_cabinet_rows", [])):
            data_type_label = item.get("data_type_label", "32位 浮点数")
            channel_name = item.get("channel_name") or ("只读4WUB" if "16位" in data_type_label else "只读4DF")
            ws.cell(row=row_no, column=1, value=channel_no)
            ws.cell(row=row_no, column=2, value=item["var_name"])
            ws.cell(row=row_no, column=3, value="SINGLE")
            ws.cell(row=row_no, column=4, value=channel_name)
            ws.cell(row=row_no, column=5, value="只读")
            ws.cell(row=row_no, column=6, value="[4区]输出寄存器")
            ws.cell(row=row_no, column=7, value=data_type_label)
            ws.cell(row=row_no, column=8, value=item["address"])
            ws.cell(row=row_no, column=9, value=item.get("description"))
            ws.cell(row=row_no, column=10, value=item.get("unit"))
            for col_offset, column_id in enumerate(spec.label_column_ids, start=11):
                ws.cell(row=row_no, column=col_offset, value=self._liquidcool_cabinet_label(column_id, item))
            row_no += 1
        if row_no > spec.data_start_row:
            self._style_cabinet_rows(ws, spec.data_start_row, row_no - 1)
        ws.freeze_panes = f"A{spec.data_start_row}"
        self._apply_classic_aux_layout(ws, kind="cabinet")

    def _append_single_cabinet_rows_to_combined(
        self,
        ws,
        *,
        row_no: int,
        channel_no: int,
    ) -> tuple[int, int]:
        fill_map = {
            "IA": FILL_UNIFIED_CABINET_IA,
            "PA": FILL_UNIFIED_CABINET_PA,
            "EA": FILL_UNIFIED_CABINET_EA,
            "KA": FILL_UNIFIED_CABINET_KA,
            "P": FILL_UNIFIED_CABINET_PA,
            "E": FILL_UNIFIED_CABINET_EA,
        }
        metric_labels = {
            "IA": "单机柜·总电流",
            "PA": "单机柜·总功率",
            "EA": "单机柜·总电能",
            "KA": "单机柜·状态字",
        }
        for item in self.model.get("single_cabinet_rows", []):
            metric_code = str(item.get("metric_code") or "IA").upper()
            data_scope = str(item.get("data_scope") or "total")
            display_cabinet_index = int(
                item.get("display_cabinet_index")
                or item.get("cabinet_index")
                or 0
            )
            screen_column = int(item.get("screen_column", 1) or 1)
            column_prefix = (
                f"{'第一列' if screen_column == 1 else '第二列'} · "
                if self.project.get("topology", {}).get("screen_topology_mode")
                == "single_screen_two_columns"
                else ""
            )
            data_type_label = item.get("data_type_label", "32位 浮点数")
            channel_name = item.get("channel_name") or (
                "只读4WUB" if "16位" in data_type_label else "只读4DF"
            )
            ws.cell(row=row_no, column=1, value=channel_no)
            ws.cell(row=row_no, column=2, value=item["var_name"])
            ws.cell(row=row_no, column=3, value="SINGLE")
            ws.cell(row=row_no, column=4, value=channel_name)
            ws.cell(row=row_no, column=5, value="只读")
            ws.cell(row=row_no, column=6, value="[4区]输出寄存器")
            ws.cell(row=row_no, column=7, value=data_type_label)
            ws.cell(row=row_no, column=8, value=item["address"])
            ws.cell(row=row_no, column=9, value=item.get("description"))
            ws.cell(row=row_no, column=10, value=item.get("unit"))
            if data_scope == "screen_total":
                summary_description = str(item.get("description") or metric_code)
                if summary_description.startswith("整屏"):
                    summary_description = summary_description[2:]
                cabinet_group_label = f"整屏汇总 · {summary_description}"
            else:
                route = item.get("route")
                route_label = f" · {route}路" if route else ""
                cabinet_group_label = (
                    f"{column_prefix}{display_cabinet_index}#机柜{route_label} · "
                    f"{metric_labels.get(metric_code, '单机柜数据')}"
                )
            ws.cell(row=row_no, column=11, value=cabinet_group_label)
            set_fill_range(ws, row_no, row_no, 1, 11, fill_map.get(metric_code, FILL_WHITE))
            row_no += 1
            channel_no += 1
        return row_no, channel_no

    def _append_point_block(
        self,
        ws,
        row_no: int,
        channel_no: int,
        points: list[dict[str, Any]],
        group_label: str,
        annotation_lines: list[str] | None,
        layout_kind: str = "combined",
        canonical_topology_metadata: str | None = None,
    ) -> tuple[int, int]:
        start_row = row_no
        for index, point in enumerate(points):
            desc, unit = self._describe_prefix(point["prefix"])
            data_type_label = self._point_data_type_label(point)
            channel_name = self._channel_name_for_point(point, data_type_label)

            ws.cell(row=row_no, column=1, value=channel_no)
            ws.cell(row=row_no, column=2, value=point["var_name"])
            ws.cell(row=row_no, column=3, value="SINGLE")
            if self._is_two_column_profile() and not self._is_liquidcool_profile() and layout_kind == "combined":
                ws.cell(row=row_no, column=4, value="只读")
                ws.cell(row=row_no, column=5, value="[4区]输出寄存器")
                ws.cell(row=row_no, column=6, value=data_type_label)
                ws.cell(row=row_no, column=7, value=point["address"])
                ws.cell(row=row_no, column=8, value=desc)
                ws.cell(row=row_no, column=9, value=unit)
                if index == 0:
                    ws.cell(row=row_no, column=10, value=group_label)
                if annotation_lines and index < len(annotation_lines):
                    ws.cell(row=row_no, column=11, value=annotation_lines[index])
            else:
                ws.cell(row=row_no, column=4, value=channel_name)
                ws.cell(row=row_no, column=5, value="只读")
                ws.cell(row=row_no, column=6, value="[4区]输出寄存器")
                ws.cell(row=row_no, column=7, value=data_type_label)
                ws.cell(row=row_no, column=8, value=point["address"])
                ws.cell(row=row_no, column=9, value=desc)
            if self._is_liquidcool_profile():
                if layout_kind == "repeater":
                    ws.cell(row=row_no, column=10, value=unit)
                    if index == 0:
                        ws.cell(row=row_no, column=11, value=group_label)
                else:
                    ws.cell(row=row_no, column=12, value=unit)
                    if index == 0:
                        ws.cell(row=row_no, column=10, value=group_label)
                        ws.cell(row=row_no, column=11, value=group_label)
                    if annotation_lines and index < len(annotation_lines):
                        ws.cell(row=row_no, column=13, value=annotation_lines[index])
            else:
                if not (self._is_two_column_profile() and layout_kind == "combined"):
                    ws.cell(row=row_no, column=10, value=unit)
                    if index == 0:
                        ws.cell(row=row_no, column=11, value=group_label)
                        if canonical_topology_metadata:
                            metadata_column = 17 if self._is_unified_master() else 16
                            ws.cell(
                                row=row_no,
                                column=metadata_column,
                                value=canonical_topology_metadata,
                            )
                    if annotation_lines and index < len(annotation_lines):
                        ws.cell(row=row_no, column=12, value=annotation_lines[index])
            row_no += 1
            channel_no += 1
        end_row = row_no - 1
        if self._is_liquidcool_profile() and end_row >= start_row:
            if layout_kind == "repeater":
                if end_row > start_row:
                    ws.merge_cells(start_row=start_row, start_column=11, end_row=end_row, end_column=11)
            elif end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)
                ws.merge_cells(start_row=start_row, start_column=11, end_row=end_row, end_column=11)
        return row_no, channel_no

    def _communication_lines(self, *, include_screen_address_note: bool = True) -> tuple[str, str]:
        protocol = self.communication.get("protocol", "Modbus RTU")
        baud_rate = self.communication.get("baud_rate", 9600)
        parity = self.communication.get("parity", "N")
        data_bits = self.communication.get("data_bits", 8)
        stop_bits = self.communication.get("stop_bits", 1)
        default_address = self.communication.get("default_screen_address", 1)
        line1 = f"1、通信协议：{protocol} （有关协议详细说明，请参见 Modbus 通信协议说明）"
        line2 = f"2、默认通信参数：{baud_rate}  {parity}  {data_bits}  {stop_bits}"
        if include_screen_address_note:
            line2 += f" （监控屏的地址默认都为{default_address}，可自行进设置界面修改）"
        return line1, line2

    def _build_intro_lines(self) -> list[str]:
        render_variant_id = self._render_variant_id()
        if render_variant_id == "unified_master":
            mode_label = (
                "按监控模块组织"
                if self._measurement_layout_mode() == "by_branch"
                else "按插接箱组织"
            )
            topology_mode = self.project.get("topology", {}).get("screen_topology_mode")
            screen_scope = "单屏双列（每列均含 A/B 路）" if topology_mode == "single_screen_two_columns" else "单屏单列（含 A/B 路）"
            return [
                f"本协议按项目最大设备配置编制，屏内结构为{screen_scope}；基础遥测与单机柜数据统一放在“{self._base_sheet_name()}”主表。",
                f"设备组织方式：{mode_label}；A/B 路以不同底色区分，单机柜采集项按数据类型区分颜色。",
            ]
        if render_variant_id == "classic_liquidcool":
            upload_port = self.project.get("topology", {}).get("upload_port_profile") or "A4B4"
            return [
                "该项目一个小母线屏监控一列机柜，一列分为AB两路母线（即主备两路）",
                f"通过屏后转接板的{upload_port}端口上传数据，下面是按照插接箱最多的情况所写，多余的部分可以忽略",
            ]
        if render_variant_id == "classic_two_columns":
            upload_port = self.project.get("topology", {}).get("upload_port_profile") or "A3B3"
            return [
                "该项目现场一个小母线屏监控两列机柜（一个通道），一列分为AB两路母线（即主备两路）",
                f"一般通过屏后转接板的{upload_port}串口上传数据，下面按插接箱最多的情况编制，多余条目可忽略。",
            ]
        route_models = self.model["routes"]
        uses_only_single_phase_triplet = bool(route_models) and all(
            physical_box.get("type_code") == "1P*3"
            for route_model in route_models
            for physical_box in route_model.get("physical_plug_boxes", [])
        )
        if uses_only_single_phase_triplet:
            return [
                "该项目现场一个小母线屏监控一列机柜，一列分为AB两路母线（即主备两路），AB路布局相同，该协议按图纸单列最多设备来写，插接箱均为3*1P规格"
            ]
        repeater_max = max((len(route_model["repeater_units"]) for route_model in route_models), default=0)
        repeater_alias = self._repeater_sheet_name()
        cabinet_count = len(self.model.get("single_cabinet_rows", []))
        sequence = self._summarize_route_box_sequence(route_models[0]["physical_plug_boxes"] if route_models else [])
        capacity_parts: list[str] = []
        if repeater_max:
            capacity_parts.append(f"单路最多{repeater_max}个{repeater_alias}")
        if cabinet_count:
            capacity_parts.append(f"最多{cabinet_count}个机柜")
        lines = [
            "该项目现场一个小母线屏监控一列机柜，一列分为AB两路母线（即主备两路）",
            "AB路布局相同，该协议按图纸单列最多设备来写",
            "单路插接箱依次为：" + (sequence or "无"),
            f"通讯协议按板卡/分路建模，一个板卡代表一个分路；{'、'.join(capacity_parts)}" if capacity_parts else "通讯协议按板卡/分路建模，一个板卡代表一个分路",
        ]
        return lines

    def _classic_intro_note(self) -> str:
        return (
            "注：以下所有的模拟量数据例如电压电量等都是32位数据存放于两个寄存器中，小寄存器存中放的是高字节，"
            "例如寄存器1061和1062组成了该始端箱的总电量，1061中存放的是用电量的高字节"
        )

    def _unified_float_reading_note(self) -> str:
        return (
            "说明：本表模拟量采用 IEEE 754 单精度（32 位）浮点格式，每个测点连续占用两个 16 位 Modbus 寄存器。"
            "读取时按高字在前的字序组合：低地址寄存器存放高 16 位，高地址寄存器存放低 16 位；"
            "若测点起始地址为 N，则 N 为高字，N+1 为低字。"
        )

    def _liquidcool_combined_note(self) -> str:
        return (
            "注：以下所有的数据都是32位数据存放于两个寄存器中，除了断路器状态字是高对高低对低，"
            "其它都是小寄存器中存放的高字节，例如寄存器1070和1071组成了该始端箱的总电量，1070中存放的是用电量的高字节"
        )

    def _liquidcool_alarm_note(self) -> str:
        return (
            "注：以下所有的数据都是32位数据存放于两个寄存器中，除了最上面的报警阈值外，其它报警状态字都是小寄存器存中放的低字节，"
            "例如寄存器6050和6051这个数据中，6050存放的是低字节Bit0~15"
        )

    def _build_alarm_rows(self) -> list[AlarmRow]:
        chunk_size = 32 if self.address_profile.get("alarm_word_mode") == "32bit" else 16
        register_size = 2 if self.address_profile.get("alarm_word_mode") == "32bit" else 1
        data_type_label = "32位 无符号二进制" if register_size == 2 else "16位 无符号二进制"
        address = self.address_profile["alarm_base"]
        channel_no = 0
        category_counter: dict[str, int] = {}
        rows: list[AlarmRow] = []

        for category, title_template, bits_by_route in self._build_alarm_bit_groups():
            category_counter.setdefault(category, 1)
            for route, bits in bits_by_route:
                if not bits:
                    continue
                chunks = [bits[i : i + chunk_size] for i in range(0, len(bits), chunk_size)]
                group_title = self._format_alarm_group_title(title_template, route)
                first = True
                for chunk in chunks:
                    var_name = f"State_{category}{category_counter[category]}"
                    rows.append(
                        AlarmRow(
                            channel_no=channel_no if first else None,
                            var_name=var_name,
                            register_address=address,
                            description=group_title if first else None,
                            bit_lines=[f"Bit{idx}： {text}" for idx, text in enumerate(chunk)],
                            data_type_label=data_type_label,
                        )
                    )
                    first = False
                    category_counter[category] += 1
                    address += register_size
                channel_no += 1
        return rows

    def _build_alarm_bit_groups(self) -> list[tuple[str, str, list[tuple[str, list[str]]]]]:
        routes = self.model["routes"]
        groups: list[tuple[str, str, list[tuple[str, list[str]]]]] = []

        def per_route_bits(builder: Callable[[dict[str, Any]], list[str]]) -> list[tuple[str, list[str]]]:
            return [(route_model["route"], builder(route_model)) for route_model in routes]

        if self._is_unified_master() and self.protocol_layout.get("alarm_start_box_first", True):
            groups.append(("SPD", "{route}路始端箱浪涌报警", per_route_bits(self._bits_spd)))
            groups.append(("THD", "{route}路始端箱谐波报警", per_route_bits(self._bits_thd)))
            groups.append(("In", "{route}路始端箱漏电流报警", per_route_bits(self._bits_in)))
            groups.append(("FH", "{route}路始端箱频率上限报警", per_route_bits(self._bits_frequency_high)))
            groups.append(("FL", "{route}路始端箱频率下限报警", per_route_bits(self._bits_frequency_low)))
            groups.append(("LoadH", "{route}路始端箱负载率上限报警", per_route_bits(self._bits_start_load_high)))
            groups.append(("UnbH", "{route}路始端箱不平衡度上限报警", per_route_bits(self._bits_start_unbalance_high)))
            if self._measurement_layout_mode() == "by_branch":
                groups.append(("Power", "{route}路电源模块异常报警", per_route_bits(self._bits_power_module)))
            groups.append(
                (
                    "Com",
                    "{route}路设备通讯异常报警",
                    per_route_bits(
                        lambda route_model: self._bits_start_com(route_model)
                        + self._bits_other_com(route_model)
                    ),
                )
            )
            groups.append(
                (
                    "LoadHC",
                    "{route}路负载率上限报警",
                    per_route_bits(
                        lambda route_model: self._bits_three_phase_metric(
                            route_model,
                            "Load",
                            "负载率超上限",
                        )
                    ),
                )
            )
            for category, title, prefix, suffix in (
                ("VH", "{route}路电压上限报警", "U", "电压超上限"),
                ("VL", "{route}路电压下限报警", "U", "电压超下限"),
                ("PH", "{route}路功率上限报警", "P", "功率超上限"),
                ("IH", "{route}路电流上限报警", "I", "电流超上限"),
                ("PFL", "{route}路功率因数下限报警", "PF", "功率因数超下限"),
                ("VLL", "{route}路分闸报警", "State", "分闸"),
            ):
                groups.append(
                    (
                        category,
                        title,
                        per_route_bits(
                            lambda route_model, point_prefix=prefix, text_suffix=suffix: self._bits_three_phase_metric(
                                route_model,
                                point_prefix,
                                text_suffix,
                            )
                        ),
                    )
            )
            groups.append(
                (
                    "TH",
                    "{route}路温度上限报警",
                    per_route_bits(
                        self._bits_three_phase_temperature
                        if self._measurement_layout_mode() == "by_branch"
                        else self._bits_temperature_entities
                    ),
                )
            )
            return groups

        groups.append(("SPD", "{route}路始端箱浪涌报警", per_route_bits(self._bits_spd)))
        groups.append(("THD", "{route}路始端箱谐波报警", per_route_bits(self._bits_thd)))
        groups.append(("In", "{route}路始端箱漏电流报警", per_route_bits(self._bits_in)))
        groups.append(("FH", "{route}路始端箱频率上限报警", per_route_bits(self._bits_frequency_high)))
        groups.append(("FL", "{route}路始端箱频率下限报警", per_route_bits(self._bits_frequency_low)))
        groups.append(("Com", "监控屏与子设备通讯异常报警", per_route_bits(self._bits_com)))
        groups.append(("VH", "电压上限报警", per_route_bits(lambda route_model: self._bits_three_phase_metric(route_model, "U", "电压超上限"))))
        groups.append(("VL", "电压下限报警", per_route_bits(lambda route_model: self._bits_three_phase_metric(route_model, "U", "电压超下限"))))
        groups.append(("VLL", "分闸报警", per_route_bits(lambda route_model: self._bits_three_phase_metric(route_model, "State", "分闸"))))
        groups.append(("IH", "电流上限报警", per_route_bits(lambda route_model: self._bits_three_phase_metric(route_model, "I", "电流超上限"))))
        groups.append(("PFL", "功率因数下限报警", per_route_bits(lambda route_model: self._bits_three_phase_metric(route_model, "PF", "功率因数超下限"))))
        groups.append(("PH", "功率上限报警", per_route_bits(lambda route_model: self._bits_three_phase_metric(route_model, "P", "功率超上限"))))
        groups.append(("TH", "温度上限报警", per_route_bits(self._bits_temperature_entities)))
        return groups

    def _format_alarm_group_title(self, title_template: str, route: str) -> str:
        if "{route}" in title_template:
            return title_template.format(route=route)
        return f"{title_template}\n（{route}路）"

    def _bits_spd(self, route_model: dict[str, Any]) -> list[str]:
        return [f"{route_model['route']}路始端箱{item['instance_name']}浪涌故障" for item in route_model["start_boxes"]]

    def _bits_thd(self, route_model: dict[str, Any]) -> list[str]:
        result = []
        for item in route_model["start_boxes"]:
            if self._has_prefix(item["points"], "THD"):
                result.append(f"{route_model['route']}路始端箱{item['instance_name']}谐波超上限")
        return result

    def _bits_in(self, route_model: dict[str, Any]) -> list[str]:
        result = []
        for item in route_model["start_boxes"]:
            if self._has_prefix(item["points"], "In"):
                result.append(f"{route_model['route']}路始端箱{item['instance_name']}漏电流超上限")
        return result

    def _bits_frequency_high(self, route_model: dict[str, Any]) -> list[str]:
        return self._bits_frequency(route_model, "频率超上限")

    def _bits_frequency_low(self, route_model: dict[str, Any]) -> list[str]:
        return self._bits_frequency(route_model, "频率超下限")

    def _bits_frequency(self, route_model: dict[str, Any], suffix: str) -> list[str]:
        result = []
        for item in route_model["start_boxes"]:
            if self._has_prefix(item["points"], "F"):
                result.append(f"{route_model['route']}路始端箱{item['instance_name']}{suffix}")
        return result

    def _bits_com(self, route_model: dict[str, Any]) -> list[str]:
        result = []
        for item in route_model["start_boxes"]:
            result.append(f"{route_model['route']}路始端箱{item['instance_name']}通讯异常")
        for physical_box in route_model["physical_plug_boxes"]:
            for board in physical_box["boards"]:
                for branch in board["branches"]:
                    if branch.get("points"):
                        if branch.get("branch_kind") == "single_phase_triplet_aggregate":
                            result.append(f"{self._plug_box_label(route_model['route'], physical_box)}通讯异常")
                        else:
                            result.append(
                                f"{self._branch_group_label(route_model['route'], physical_box, branch, include_topology_context=False)}通讯异常"
                            )
        for repeater in route_model["repeater_units"]:
            if repeater.get("points"):
                result.append(f"{self._repeater_entity_label(route_model['route'], repeater)}通讯异常")
        return result

    def _bits_start_com(self, route_model: dict[str, Any]) -> list[str]:
        return [
            f"{route_model['route']}路始端箱{item['instance_name']}通讯异常"
            for item in route_model["start_boxes"]
        ]

    def _bits_start_load_high(self, route_model: dict[str, Any]) -> list[str]:
        return [
            f"{route_model['route']}路始端箱{item['instance_name']}负载率超上限"
            for item in route_model["start_boxes"]
            if self._has_prefix(item["points"], "LoadS")
        ]

    def _bits_start_unbalance_high(self, route_model: dict[str, Any]) -> list[str]:
        return [
            f"{route_model['route']}路始端箱{item['instance_name']}三相不平衡度超上限"
            for item in route_model["start_boxes"]
            if self._has_prefix(item["points"], "UBS")
        ]

    def _bits_power_module(self, route_model: dict[str, Any]) -> list[str]:
        return [f"{route_model['route']}路电源模块异常"] if route_model.get("start_boxes") else []

    def _bits_other_com(self, route_model: dict[str, Any]) -> list[str]:
        result: list[str] = []
        for physical_box in route_model["physical_plug_boxes"]:
            if self._measurement_layout_mode() == "by_branch":
                if any(branch.get("points") for board in physical_box["boards"] for branch in board["branches"]):
                    result.append(
                        f"{self._monitor_module_label(route_model['route'], physical_box)}通讯异常"
                    )
                continue
            for board in physical_box["boards"]:
                for branch in board["branches"]:
                    if not branch.get("points"):
                        continue
                    result.append(
                        f"{self._branch_group_label(route_model['route'], physical_box, branch, include_topology_context=False)}通讯异常"
                    )
        for repeater in route_model["repeater_units"]:
            if repeater.get("points"):
                result.append(f"{self._repeater_entity_label(route_model['route'], repeater)}通讯异常")
        return result

    def _bits_start_three_phase_metric(
        self,
        route_model: dict[str, Any],
        point_prefix_base: str,
        suffix: str,
    ) -> list[str]:
        result: list[str] = []
        for item in route_model["start_boxes"]:
            if self._supports_three_phase(item["points"], point_prefix_base):
                result.extend(
                    f"{route_model['route']}路始端箱{item['instance_name']}-{phase_label}相{suffix}"
                    for phase_label in PHASE_LABELS
                )
        return result

    def _bits_other_three_phase_metric(
        self,
        route_model: dict[str, Any],
        point_prefix_base: str,
        suffix: str,
    ) -> list[str]:
        result: list[str] = []
        for physical_box in route_model["physical_plug_boxes"]:
            for board in physical_box["boards"]:
                for branch in board["branches"]:
                    points = branch.get("points", [])
                    if not points or not self._supports_three_phase(points, point_prefix_base):
                        continue
                    if branch.get("branch_kind") == "single_phase_triplet_aggregate":
                        result.extend(
                            f"{label}{suffix}"
                            for label in self._single_phase_triplet_alarm_labels(
                                route_model["route"],
                                physical_box,
                                branch,
                            )
                        )
                        continue
                    entity_label = self._branch_group_label(
                        route_model["route"],
                        physical_box,
                        branch,
                        include_topology_context=False,
                    )
                    result.extend(f"{entity_label}-{phase_label}相{suffix}" for phase_label in PHASE_LABELS)
        return result

    def _bits_start_temperature(self, route_model: dict[str, Any]) -> list[str]:
        return [
            f"{route_model['route']}路始端箱{item['instance_name']}温度超上限"
            for item in route_model["start_boxes"]
            if self._has_temperature_points(item["points"])
        ]

    def _bits_three_phase_temperature(self, route_model: dict[str, Any]) -> list[str]:
        result: list[str] = []
        for item in route_model["start_boxes"]:
            for phase_label, prefix in (("A", "Ta"), ("B", "Tb"), ("C", "Tc")):
                if self._has_prefix(item["points"], prefix):
                    result.append(
                        f"{route_model['route']}路始端箱{item['instance_name']}-{phase_label}相温度超上限"
                    )
        for physical_box in route_model["physical_plug_boxes"]:
            for board in physical_box["boards"]:
                for branch in board["branches"]:
                    points = branch.get("points", [])
                    entity_label = self._branch_group_label(
                        route_model["route"],
                        physical_box,
                        branch,
                        include_topology_context=False,
                    )
                    for phase_label, prefix in (("A", "Ta"), ("B", "Tb"), ("C", "Tc")):
                        if self._has_prefix(points, prefix):
                            result.append(f"{entity_label}-{phase_label}相温度超上限")
        return result

    def _bits_other_temperature(self, route_model: dict[str, Any]) -> list[str]:
        result: list[str] = []
        for physical_box in route_model["physical_plug_boxes"]:
            for board in physical_box["boards"]:
                for branch in board["branches"]:
                    points = branch.get("points", [])
                    if points and self._has_temperature_points(points):
                        result.append(
                            f"{self._branch_group_label(route_model['route'], physical_box, branch, include_topology_context=False)}温度超上限"
                        )
        for repeater in route_model["repeater_units"]:
            points = repeater.get("points", [])
            if points and self._has_temperature_points(points):
                result.append(f"{self._repeater_entity_label(route_model['route'], repeater)}温度超上限")
        return result

    def _bits_three_phase_metric(self, route_model: dict[str, Any], point_prefix_base: str, suffix: str) -> list[str]:
        result = []
        for item in route_model["start_boxes"]:
            if self._supports_three_phase(item["points"], point_prefix_base):
                for phase_label in PHASE_LABELS:
                    result.append(f"{route_model['route']}路始端箱{item['instance_name']}-{phase_label}相{suffix}")
        for physical_box in route_model["physical_plug_boxes"]:
            for board in physical_box["boards"]:
                for branch in board["branches"]:
                    points = branch.get("points", [])
                    if not points or not self._supports_three_phase(points, point_prefix_base):
                        continue
                    if branch.get("branch_kind") == "single_phase_triplet_aggregate":
                        result.extend(f"{label}{suffix}" for label in self._single_phase_triplet_alarm_labels(route_model["route"], physical_box, branch))
                        continue
                    entity_label = self._branch_group_label(
                        route_model["route"],
                        physical_box,
                        branch,
                        include_topology_context=False,
                    )
                    for phase_label in PHASE_LABELS:
                        result.append(f"{entity_label}-{phase_label}相{suffix}")
        return result

    def _bits_temperature_entities(self, route_model: dict[str, Any]) -> list[str]:
        result = []
        for entity_label, points in self._iter_temperature_entities(route_model):
            if self._has_temperature_points(points):
                result.append(f"{entity_label}温度超上限")
        return result

    def _iter_three_phase_entities(self, route_model: dict[str, Any]) -> Iterable[tuple[str, list[dict[str, Any]]]]:
        for item in route_model["start_boxes"]:
            yield f"{route_model['route']}路始端箱{item['instance_name']}", item["points"]
        for physical_box in route_model["physical_plug_boxes"]:
            for board in physical_box["boards"]:
                for branch in board["branches"]:
                    points = branch.get("points", [])
                    if points:
                        if branch.get("branch_kind") == "single_phase_triplet_aggregate":
                            yield self._plug_box_label(route_model["route"], physical_box), points
                        else:
                            yield self._branch_group_label(
                                route_model["route"],
                                physical_box,
                                branch,
                                include_topology_context=False,
                            ), points

    def _iter_temperature_entities(self, route_model: dict[str, Any]) -> Iterable[tuple[str, list[dict[str, Any]]]]:
        yield from self._iter_three_phase_entities(route_model)
        for repeater in route_model["repeater_units"]:
            points = repeater.get("points", [])
            if points:
                yield self._repeater_entity_label(route_model["route"], repeater), points

    def _has_temperature_points(self, points: list[dict[str, Any]]) -> bool:
        return any(point["prefix"].startswith("T") and not point["prefix"].startswith("THD") for point in points)

    def _supports_three_phase(self, points: list[dict[str, Any]], phase_prefix: str) -> bool:
        if phase_prefix == "State":
            return self._has_prefix(points, "StateS") or self._has_prefix(points, "StateC")
        return all(self._has_prefix(points, f"{phase_prefix}{phase.lower()}") for phase in PHASE_LABELS)

    def _has_prefix(self, points: list[dict[str, Any]], prefix: str) -> bool:
        return any(point["prefix"] == prefix for point in points)

    def _point_data_type_label(self, point: dict[str, Any]) -> str:
        explicit = point.get("data_type_label")
        if explicit:
            return explicit
        if point["prefix"].startswith("State"):
            if point.get("register_size") == 2:
                return "32位 无符号二进制"
            return "16位 无符号二进制"
        return "32位 浮点数"

    def _channel_name_for_point(self, point: dict[str, Any], data_type_label: str) -> str:
        if "无符号二进制" not in data_type_label:
            return "只读4DF"
        register_size = point.get("register_size")
        if register_size == 2 or "32位" in data_type_label:
            return "只读4DUB"
        return "只读4WUB"

    def _describe_prefix(self, prefix: str) -> tuple[str, str | None]:
        return PREFIX_META.get(prefix, (prefix, None))

    def _repeater_sheet_name(self) -> str:
        if self._is_unified_master():
            return "中继器"
        aliases = {repeater.get("alias") for route in self.model["routes"] for repeater in route["repeater_units"] if repeater.get("alias")}
        if len(aliases) == 1:
            alias = next(iter(aliases))
            if alias in {"中继器", "中继单元", "连接器测温"}:
                return alias
        return "中继器"

    def _base_sheet_name(self) -> str:
        return str(
            self.export_profile.get("base_sheet_name")
            or self.protocol_layout.get("base_sheet_name")
            or "始端箱和插接箱"
        ).strip() or "设备遥测"

    def _measurement_layout_mode(self) -> str:
        value = str(
            self.protocol_layout.get("measurement_layout_mode") or "by_plug_box"
        ).strip()
        return value if value in {"by_plug_box", "by_branch"} else "by_plug_box"

    def _embed_single_cabinet_in_base_sheet(self) -> bool:
        return bool(
            self.export_profile.get("embed_single_cabinet_in_base_sheet")
            and self.protocol_layout.get("embed_single_cabinet_in_base_sheet", True)
            and self.model.get("single_cabinet_rows")
        )

    def _should_render_repeater_sheet(self) -> bool:
        has_repeater_rows = any(route.get("repeater_units") for route in self.model.get("routes", []))
        return bool(self.export_profile.get("include_repeater_sheet", True) and has_repeater_rows)

    def _should_render_single_cabinet_sheet(self) -> bool:
        return bool(self.export_profile.get("include_single_cabinet_sheet") and self.model.get("single_cabinet_rows"))

    def _should_render_alarm_sheet(self) -> bool:
        return bool(self.export_profile.get("include_alarm_sheet", True))

    def _is_liquidcool_profile(self) -> bool:
        return self._render_variant_id() == "classic_liquidcool"

    def _is_two_column_profile(self) -> bool:
        return self._render_variant_id() == "classic_two_columns"

    def _is_unified_master(self) -> bool:
        return self._render_variant_id() == "unified_master"

    def _render_variant_id(self) -> str:
        explicit_variant = str(self.export_profile.get("render_variant_id") or "").strip()
        if explicit_variant:
            return explicit_variant
        topology_mode = self.project.get("topology", {}).get("screen_topology_mode")
        if self.export_profile.get("id") == "classic_combined_liquidcool_default" or self.export_profile.get("subtype") == "liquidcool_hybrid":
            return "classic_liquidcool"
        if self.export_profile.get("subtype") == "single_screen_two_columns" or topology_mode == "single_screen_two_columns":
            return "classic_two_columns"
        return "classic_standard"

    def _combined_sheet_spec(self) -> CombinedSheetSpec:
        return CLASSIC_COMBINED_SHEET_SPECS.get(self._render_variant_id(), CLASSIC_COMBINED_SHEET_SPECS["classic_standard"])

    def _alarm_sheet_spec(self) -> AlarmSheetSpec:
        render_variant_id = self._render_variant_id()
        if render_variant_id == "classic_liquidcool":
            render_variant_id = "classic_standard"
        return CLASSIC_ALARM_SHEET_SPECS.get(render_variant_id, CLASSIC_ALARM_SHEET_SPECS["classic_standard"])

    def _cabinet_sheet_spec(self) -> CabinetSheetSpec:
        render_variant_id = self._render_variant_id()
        if render_variant_id == "classic_liquidcool":
            return CABINET_SHEET_SPECS["classic_liquidcool"]
        return CABINET_SHEET_SPECS["classic_standard"]

    def _note_text(self, note_kind: str) -> str:
        if note_kind == "classic_intro":
            return self._classic_intro_note()
        if note_kind == "unified_float_reading":
            return self._unified_float_reading_note()
        if note_kind == "liquidcool_combined":
            return self._liquidcool_combined_note()
        if note_kind == "liquidcool_alarm":
            return self._liquidcool_alarm_note()
        raise KeyError(f"未知说明区文案类型: {note_kind}")

    def _build_liquidcool_state_alarm_rows(self) -> list[AlarmRow]:
        address_profile = deepcopy(self.address_profile)
        address_profile["alarm_base"] = self.address_profile["alarm_base"] + len(LIQUIDCOOL_THRESHOLD_ROWS) * 2
        group_specs = []
        for category, title_template, bits_by_route in self._build_alarm_bit_groups():
            scoped_groups = [(self._format_alarm_group_title(title_template, route), bits) for route, bits in bits_by_route]
            group_specs.append((category, scoped_groups))
        return build_alarm_rows_from_group_specs(address_profile, group_specs)

    def _display_box_type_code(self, type_code: str) -> str:
        return "3*1P" if str(type_code) == "1P*3" else str(type_code)

    def _summarize_route_box_sequence(self, physical_boxes: list[dict[str, Any]]) -> str:
        if not physical_boxes:
            return ""
        segments: list[str] = []
        current_type: str | None = None
        current_count = 0
        for physical_box in physical_boxes:
            type_code = self._display_box_type_code(str(physical_box.get("type_code") or ""))
            if current_type == type_code:
                current_count += 1
                continue
            if current_type is not None:
                segments.append(f"{current_type}×{current_count}" if current_count > 1 else current_type)
            current_type = type_code
            current_count = 1
        if current_type is not None:
            segments.append(f"{current_type}×{current_count}" if current_count > 1 else current_type)
        return "、".join(segments)

    def _plug_box_label(self, route: str, physical_box: dict[str, Any], *, include_route: bool = True) -> str:
        prefix = self._route_scope_label(route, physical_box) if include_route else ""
        box_code = f"C{physical_box['physical_box_no']}" if self._is_liquidcool_profile() else str(physical_box["physical_box_no"])
        return f"{prefix}插接箱{box_code}"

    def _route_scope_label(self, route: str, entity: dict[str, Any] | None = None) -> str:
        screen_column = int((entity or {}).get("screen_column", 1) or 1)
        topology_mode = self.project.get("topology", {}).get("screen_topology_mode")
        if topology_mode == "single_screen_two_columns":
            column_label = {1: "第一列", 2: "第二列"}.get(
                screen_column,
                f"第{screen_column}列",
            )
            return f"{column_label}{route}路"
        return f"{route}路"

    def _unified_branch_label(
        self,
        route: str,
        physical_box: dict[str, Any],
        branch: dict[str, Any],
    ) -> str:
        if self._measurement_layout_mode() == "by_branch":
            if (
                self.project.get("topology", {}).get("screen_topology_mode")
                == "single_screen_two_columns"
            ):
                module_local_branch_no = int(
                    branch.get("module_local_branch_no")
                    or branch.get("physical_branch_index")
                    or 1
                )
                return (
                    f"{self._route_scope_label(route, physical_box)}"
                    f"模块内分路{module_local_branch_no}"
                )
            return self._branch_group_label(
                route,
                physical_box,
                branch,
                include_route=True,
                include_topology_context=False,
            )
        if branch.get("branch_kind") == "single_phase_triplet_aggregate":
            numeric_labels = [
                int(value)
                for value in branch.get("logical_output_labels", [])
                if str(value).isdigit()
            ]
            start_index = min(numeric_labels, default=1)
            output_count = int(branch.get("logical_output_count") or len(numeric_labels) or 1)
            end_index = max(numeric_labels, default=start_index + output_count - 1)
            return f"分路{start_index}–{end_index}（共享测量点集）"
        return f"分路{branch.get('physical_branch_index') or 1}"

    def _unified_device_label(self, route: str, physical_box: dict[str, Any]) -> str:
        if self._measurement_layout_mode() == "by_branch":
            return self._monitor_module_label(route, physical_box)
        return self._plug_box_label(route, physical_box)

    def _write_unified_device_label(
        self,
        ws,
        *,
        start_row: int,
        end_row: int,
        label: str,
    ) -> None:
        ws.cell(row=start_row, column=12, value=label)
        if end_row > start_row:
            ws.merge_cells(
                start_row=start_row,
                start_column=12,
                end_row=end_row,
                end_column=12,
            )

    def _single_phase_triplet_alarm_labels(
        self,
        route: str,
        physical_box: dict[str, Any],
        branch: dict[str, Any],
    ) -> list[str]:
        base_label = self._plug_box_label(route, physical_box)
        output_labels = list(branch.get("logical_output_labels") or ["A", "B", "C"])
        return [
            f"{base_label}分路{index}-{phase_label}相"
            for index, phase_label in enumerate(output_labels, start=1)
        ]

    def _branch_group_label(
        self,
        route: str,
        physical_box: dict[str, Any],
        branch: dict[str, Any],
        *,
        include_route: bool = True,
        include_topology_context: bool = True,
    ) -> str:
        if self._measurement_layout_mode() == "by_branch" and branch.get("output_no") is not None:
            output_no = int(branch["output_no"])
            route_scope = self._route_scope_label(route, physical_box)
            if (
                self.project.get("topology", {}).get("screen_topology_mode")
                == "single_screen_two_columns"
            ):
                display_module_no = int(
                    branch.get("display_module_no")
                    or physical_box.get("display_module_no")
                    or branch.get("module_no")
                    or physical_box.get("module_no")
                    or 1
                )
                module_local_branch_no = int(
                    branch.get("module_local_branch_no")
                    or branch.get("physical_branch_index")
                    or 1
                )
                return (
                    f"{route_scope}{display_module_no}#监控模块·"
                    f"模块内分路{module_local_branch_no}"
                )
            stable_output_label = f"{route_scope}输出分路{output_no}"
            custom_output_name = str(branch.get("output_name") or "").strip()
            if custom_output_name and custom_output_name != stable_output_label:
                stable_output_label = f"{stable_output_label}·{custom_output_name}"

            module_no = branch.get("module_no") or physical_box.get("module_no")
            module_local_branch_no = branch.get("module_local_branch_no")
            parent_parts: list[str] = []
            if module_no is not None:
                parent_parts.append(f"{route_scope}{module_no}#监控模块")
            if module_local_branch_no is not None:
                parent_parts.append(f"模块内分路{module_local_branch_no}")
            if parent_parts and include_topology_context:
                return f"{stable_output_label}（{'·'.join(parent_parts)}）"
            return stable_output_label
        if branch.get("branch_kind") == "single_phase_triplet_aggregate":
            return self._plug_box_label(route, physical_box, include_route=include_route)
        base_label = self._plug_box_label(route, physical_box, include_route=include_route)
        if self._is_liquidcool_profile():
            return f"{base_label}-分路{branch['physical_branch_index']}"
        return f"{base_label}分路{branch['physical_branch_index']}"

    def _branch_canonical_topology_metadata(
        self,
        route: str,
        physical_box: dict[str, Any],
        branch: dict[str, Any],
    ) -> str | None:
        """Serialize explicit topology for workbook-first downstream generators."""

        if self._measurement_layout_mode() != "by_branch":
            return None
        module_no = branch.get("module_no") or physical_box.get("module_no")
        if module_no is None:
            return None

        values: list[tuple[str, Any]] = [
            ("route", route),
            ("screen_column", physical_box.get("screen_column", 1)),
            ("output_no", branch.get("output_no")),
            ("module_no", module_no),
            (
                "display_module_no",
                branch.get("display_module_no")
                or physical_box.get("display_module_no")
                or module_no,
            ),
            ("module_local_branch_no", branch.get("module_local_branch_no")),
            (
                "communication_alarm_slot",
                branch.get("communication_alarm_slot")
                or physical_box.get("communication_alarm_slot"),
            ),
            (
                "communication_variable_device_code",
                branch.get("communication_variable_device_code")
                or physical_box.get("communication_variable_device_code"),
            ),
            (
                "variable_numbering_mode",
                branch.get("variable_numbering_mode")
                or physical_box.get("variable_numbering_mode"),
            ),
        ]
        serialized = [CANONICAL_TOPOLOGY_METADATA_PREFIX]
        for key, value in values:
            if value is None or str(value).strip() == "":
                continue
            text = str(value).strip()
            if "|" in text or "=" in text:
                raise ValueError(f"非法 canonical topology 元数据：{key}={text!r}")
            serialized.append(f"{key}={text}")
        return "|".join(serialized)

    def _repeater_entity_label(self, route: str, repeater: dict[str, Any]) -> str:
        alias = repeater.get("alias", "中继器")
        device_code = repeater["device_code"]
        route_scope = self._route_scope_label(route, repeater)
        if alias == "中继器":
            return f"{route_scope}中继{device_code}"
        return f"{route_scope}Z{device_code}{alias}"

    def _monitor_module_label(
        self,
        route: str,
        physical_box: dict[str, Any],
    ) -> str:
        module_no = physical_box.get("module_no") or physical_box.get("physical_box_no")
        display_module_no = physical_box.get("display_module_no") or module_no
        return f"{self._route_scope_label(route, physical_box)}{display_module_no}#监控模块"

    def _combined_block_fill(self, *, block_kind: str, route: str, block_index: int) -> PatternFill:
        if self._is_unified_master():
            return FILL_UNIFIED_ROUTE_A if route == "A" else FILL_UNIFIED_ROUTE_B
        if self._is_liquidcool_profile():
            if block_kind == "plug":
                return FILL_LIQUIDCOOL_PLUG if route == "A" else FILL_WHITE
            if block_kind == "start":
                return FILL_BLUE if route == "A" else FILL_WHITE
            return FILL_BLUE if route == "A" else FILL_WHITE

        if block_kind == "plug":
            return FILL_CLASSIC_PLUG
        if block_kind == "start":
            return FILL_BLUE if route == "A" else FILL_WHITE
        return FILL_BLUE if route == "A" else FILL_WHITE

    def _style_combined_block(
        self,
        ws,
        *,
        start_row: int,
        end_row: int,
        route: str,
        block_kind: str,
        block_index: int,
        group_label: str,
        write_state_note: bool,
    ) -> None:
        fill = self._combined_block_fill(block_kind=block_kind, route=route, block_index=block_index)
        set_fill_range(
            ws,
            start_row,
            end_row,
            1,
            1,
            fill if self._is_unified_master() else FILL_BLUE,
        )
        if self._is_unified_master():
            set_fill_range(ws, start_row, end_row, 2, 12, fill)
            if end_row > start_row:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=11,
                    end_row=end_row,
                    end_column=11,
                )
            if write_state_note:
                set_fill_range(ws, start_row, start_row, 13, 16, FILL_YELLOW)
                ws.merge_cells(
                    start_row=start_row,
                    start_column=13,
                    end_row=start_row,
                    end_column=16,
                )
                ws.cell(row=start_row, column=13, value="断路器状态字说明")
                state_lines = [
                    ("BIT0", "A相断路器", "1：闭合，0：断开"),
                    ("BIT1", "B相断路器", "1：闭合，0：断开"),
                    ("BIT2", "C相断路器", "1：闭合，0：断开"),
                ]
                for offset, (bit_label, label, value) in enumerate(state_lines, start=1):
                    ws.cell(row=start_row + offset, column=13, value=bit_label)
                    ws.cell(row=start_row + offset, column=14, value=label)
                    ws.cell(row=start_row + offset, column=15, value=value)
            return
        if self._is_liquidcool_profile():
            set_fill_range(ws, start_row, end_row, 2, 9, fill)
            fill_cell(ws.cell(row=start_row, column=10), fill)
            fill_cell(ws.cell(row=start_row, column=11), fill)
            ws.cell(row=start_row, column=10, value=group_label)
            ws.cell(row=start_row, column=11, value=group_label)
            if write_state_note:
                fill_cell(ws.cell(row=start_row, column=13), FILL_YELLOW)
                merge_with_alignment(ws, f"M{start_row}:O{start_row}")
                ws.cell(row=start_row, column=13, value="始端箱断路器状态字说明")
                ws.cell(row=start_row + 1, column=13, value="BIT0:A相断路器")
                ws.cell(row=start_row + 1, column=14, value="1：闭合，0：断开")
                ws.cell(row=start_row + 2, column=13, value="BIT1:B相断路器")
                ws.cell(row=start_row + 2, column=14, value="1：闭合，0：断开")
                ws.cell(row=start_row + 3, column=13, value="BIT2:C相断路器")
                ws.cell(row=start_row + 3, column=14, value="1：闭合，0：断开")
                plug_note_row = start_row + 4
                fill_cell(ws.cell(row=plug_note_row, column=13), FILL_YELLOW)
                merge_with_alignment(ws, f"M{plug_note_row}:O{plug_note_row}")
                ws.cell(row=plug_note_row, column=13, value="插接箱断路器状态字说明")
                ws.cell(row=plug_note_row + 1, column=13, value="BIT0:分路1-A相断路器")
                ws.cell(row=plug_note_row + 1, column=14, value="1：闭合，0：断开")
                ws.cell(row=plug_note_row + 2, column=13, value="BIT1:分路1-B相断路器")
                ws.cell(row=plug_note_row + 2, column=14, value="1：闭合，0：断开")
                ws.cell(row=plug_note_row + 3, column=13, value="BIT2:分路1-C相断路器")
                ws.cell(row=plug_note_row + 3, column=14, value="1：闭合，0：断开")
                ws.cell(row=plug_note_row + 4, column=13, value="BIT3:分路2-A相断路器")
                ws.cell(row=plug_note_row + 4, column=14, value="1：闭合，0：断开")
                ws.cell(row=plug_note_row + 5, column=13, value="BIT4:分路2-B相断路器")
                ws.cell(row=plug_note_row + 5, column=14, value="1：闭合，0：断开")
                ws.cell(row=plug_note_row + 6, column=13, value="BIT5:分路2-C相断路器")
                ws.cell(row=plug_note_row + 6, column=14, value="1：闭合，0：断开")
        elif self._is_two_column_profile():
            set_fill_range(ws, start_row, end_row, 2, 8, fill)
            fill_cell(ws.cell(row=start_row, column=10), fill)
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)
            if write_state_note:
                set_fill_range(ws, start_row, start_row, 11, 13, FILL_YELLOW)
                ws.merge_cells(start_row=start_row, start_column=11, end_row=start_row, end_column=13)
                ws.cell(row=start_row, column=11, value="断路器状态字说明")
                classic_state_lines = [
                    ("BIT0", "A相断路器", "1：闭合，0：断开"),
                    ("BIT1", "B相断路器", "1：闭合，0：断开"),
                    ("BIT2", "C相断路器", "1：闭合，0：断开"),
                ]
                for offset, (bit_label, label, value) in enumerate(classic_state_lines, start=1):
                    ws.cell(row=start_row + offset, column=11, value=bit_label)
                    ws.cell(row=start_row + offset, column=12, value=label)
                    ws.cell(row=start_row + offset, column=13, value=value)
        else:
            set_fill_range(ws, start_row, end_row, 2, 9, fill)
            fill_cell(ws.cell(row=start_row, column=11), fill)
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=11, end_row=end_row, end_column=11)
            if write_state_note:
                set_fill_range(ws, start_row, start_row, 12, 15, FILL_YELLOW)
                ws.merge_cells(start_row=start_row, start_column=12, end_row=start_row, end_column=15)
                ws.cell(row=start_row, column=12, value="断路器状态字说明")
                classic_state_lines = [
                    ("BIT0", "A相断路器", "1：闭合，0：断开"),
                    ("BIT1", "B相断路器", "1：闭合，0：断开"),
                    ("BIT2", "C相断路器", "1：闭合，0：断开"),
                ]
                for offset, (bit_label, label, value) in enumerate(classic_state_lines, start=1):
                    ws.cell(row=start_row + offset, column=12, value=bit_label)
                    ws.cell(row=start_row + offset, column=13, value=label)
                    ws.cell(row=start_row + offset, column=14, value=value)

    def _style_repeater_block(self, ws, *, start_row: int, end_row: int, route: str, block_index: int) -> None:
        if self._is_unified_master():
            fill = FILL_UNIFIED_ROUTE_A if route == "A" else FILL_UNIFIED_ROUTE_B
            set_fill_range(ws, start_row, end_row, 1, 1, fill)
        else:
            set_fill_range(ws, start_row, end_row, 1, 1, FILL_BLUE)
        if self._is_unified_master():
            fill = FILL_UNIFIED_ROUTE_A if route == "A" else FILL_UNIFIED_ROUTE_B
        elif self._is_liquidcool_profile():
            fill = FILL_ORANGE if route == "A" else FILL_WHITE
        else:
            fill = FILL_ORANGE
        set_fill_range(ws, start_row, end_row, 2, 9, fill)
        if not self._is_liquidcool_profile() and end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=11, end_row=end_row, end_column=11)
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=11, end_row=end_row, end_column=11)

    def _style_classic_alarm_rows(self, ws, start_row: int, end_row: int) -> None:
        highlight_started = False
        for row_index in range(start_row, end_row + 1):
            fill_cell(ws.cell(row=row_index, column=1), FILL_BLUE)
            description = str(ws.cell(row=row_index, column=8).value or "")
            if description and "通讯异常" in description:
                highlight_started = True
            if highlight_started:
                set_fill_range(ws, row_index, row_index, 2, 6, FILL_CLASSIC_PLUG)
                if ws.cell(row=row_index, column=8).value not in (None, ""):
                    fill_cell(ws.cell(row=row_index, column=8), FILL_CLASSIC_PLUG)
                fill_cell(ws.cell(row=row_index, column=9), FILL_CLASSIC_PLUG)
                fill_cell(ws.cell(row=row_index, column=10), FILL_CLASSIC_PLUG)
            ws.row_dimensions[row_index].height = 60

    def _style_liquidcool_alarm_rows(self, ws, start_row: int, end_row: int) -> None:
        threshold_end = start_row + len(LIQUIDCOOL_THRESHOLD_ROWS) - 1
        set_fill_range(ws, start_row, threshold_end, 1, 1, FILL_BLUE)
        set_fill_range(ws, start_row, threshold_end, 2, 9, FILL_ORANGE)
        bit_header_row = threshold_end + 1
        fill_cell(ws.cell(row=bit_header_row, column=11), FILL_THEME3)
        fill_cell(ws.cell(row=bit_header_row + 5, column=12), FILL_YELLOW)
        fill_cell(ws.cell(row=bit_header_row + 6, column=12), FILL_YELLOW)
        fill_cell(ws.cell(row=bit_header_row + 6, column=13), FILL_YELLOW)
        row_index = bit_header_row + 1
        single_rows_remaining = 10
        while row_index <= end_row:
            fill_cell(ws.cell(row=row_index, column=1), FILL_BLUE)
            fill_cell(ws.cell(row=row_index, column=8), FILL_THEME0)
            if single_rows_remaining > 0:
                single_rows_remaining -= 1
                row_index += 1
                continue
            fill_cell(ws.cell(row=row_index + 1, column=8), FILL_THEME0)
            row_index += 2

    def _style_cabinet_rows(self, ws, start_row: int, end_row: int) -> None:
        fill_map = {
            "IA": FILL_LIQUIDCOOL_CABINET_IA,
            "PA": FILL_LIQUIDCOOL_CABINET_PA,
            "EA": FILL_LIQUIDCOOL_CABINET_EA,
            "KA": FILL_YELLOW,
        }
        has_ka_rows = False
        for row_index in range(start_row, end_row + 1):
            metric_code = normalize_metric_code(ws.cell(row=row_index, column=2).value)
            fill = fill_map.get(metric_code, FILL_WHITE)
            set_fill_range(ws, row_index, row_index, 1, 9, fill)
            if self._is_liquidcool_profile():
                set_fill_range(ws, row_index, row_index, 11, 12, FILL_WHITE)
            if metric_code == "KA":
                has_ka_rows = True
        if self._is_liquidcool_profile() and has_ka_rows and end_row >= start_row + 35:
            note_row = start_row + 35
            fill_cell(ws.cell(row=note_row, column=13), FILL_YELLOW)
            merge_with_alignment(ws, f"M{note_row}:N{note_row}")
            ws.cell(row=note_row, column=13, value="机柜状态字说明:0为没电，1为有电，2为缺相")

    def _write_header_row(self, ws, row_no: int, headers: list[str]) -> None:
        for col_no, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_no, column=col_no, value=header)
            cell.font = BOLD_FONT
            cell.fill = NO_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    def _apply_classic_combined_layout(self, ws) -> None:
        self._apply_sheet_layout_spec(ws, self._classic_combined_layout_spec())

    def _apply_classic_aux_layout(self, ws, kind: str) -> None:
        spec = self._classic_aux_layout_spec(kind)
        if spec is not None:
            self._apply_sheet_layout_spec(ws, spec)
        for cell_ref in ("A1", "A2", "A3", "A6", "A8"):
            cell = ws[cell_ref]
            if self._should_left_wrap_aux_cell(cell.value):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    @staticmethod
    def _should_left_wrap_aux_cell(value: Any) -> bool:
        if not isinstance(value, str):
            return False
        text = value.strip()
        if not text:
            return False
        if text == "通道号":
            return False
        return True

    def _classic_combined_layout_spec(self) -> SheetLayoutSpec:
        return CLASSIC_COMBINED_LAYOUT_SPECS.get(self._render_variant_id(), CLASSIC_COMBINED_LAYOUT_SPECS["classic_standard"])

    def _classic_aux_layout_spec(self, kind: str) -> SheetLayoutSpec | None:
        variant_id = self._render_variant_id()
        spec = CLASSIC_AUX_LAYOUT_SPECS.get((kind, variant_id))
        if spec is not None:
            return spec
        return CLASSIC_AUX_LAYOUT_SPECS.get((kind, "classic_standard"))

    def _sheet_style_kind(self, ws) -> str | None:
        if ws.title == self._base_sheet_name():
            return "combined"
        if ws.title in {"中继器", "中继单元", "连接器测温"}:
            return "repeater"
        if ws.title == "报警状态":
            return "alarm"
        if ws.title == "单机柜数据":
            return "cabinet"
        return None

    def _sheet_style_spec(self, ws) -> SheetStyleSpec | None:
        sheet_kind = self._sheet_style_kind(ws)
        if sheet_kind is None:
            return None
        variant_id = self._render_variant_id()
        spec = CLASSIC_SHEET_STYLE_SPECS.get((sheet_kind, variant_id))
        if spec is not None:
            return spec
        return CLASSIC_SHEET_STYLE_SPECS.get((sheet_kind, "classic_standard"))

    def _apply_sheet_layout_spec(self, ws, spec: SheetLayoutSpec) -> None:
        for merge_range in spec.merge_ranges:
            ws.merge_cells(merge_range)
        for row_index, height in spec.row_heights:
            ws.row_dimensions[row_index].height = height
        for cell_ref in spec.left_wrap_cells:
            ws[cell_ref].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for cell_ref, fill in spec.fill_cells:
            fill_cell(ws[cell_ref], fill)

    def _liquidcool_cabinet_label(self, column_id: str, item: dict[str, Any]) -> str:
        metric_labels = {
            "IA": "机柜总电流",
            "PA": "机柜总功率",
            "EA": "机柜总电能",
            "KA": "机柜状态字",
        }
        metric_code = str(item.get("metric_code") or "IA")
        suffix = metric_labels.get(metric_code, item.get("description") or metric_code)
        return f"{column_id}-{int(item['cabinet_index']):02d}{suffix}"

    def _normalize_merged_range_borders(self, ws) -> set[tuple[int, int]]:
        merged_lookup: set[tuple[int, int]] = set()
        handled_ranges: set[str] = set()

        for merged_range in ws.merged_cells.ranges:
            for row_index in range(merged_range.min_row, merged_range.max_row + 1):
                for col_index in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_lookup.add((row_index, col_index))

        for merged_range in ws.merged_cells.ranges:
            range_key = str(merged_range)
            if range_key in handled_ranges:
                continue
            handled_ranges.add(range_key)
            anchor = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            left_side = deepcopy(anchor.border.left) if getattr(anchor.border.left, "style", None) else THIN
            right_side = deepcopy(anchor.border.right) if getattr(anchor.border.right, "style", None) else THIN
            top_side = deepcopy(anchor.border.top) if getattr(anchor.border.top, "style", None) else THIN
            bottom_side = deepcopy(anchor.border.bottom) if getattr(anchor.border.bottom, "style", None) else THIN
            for row_index in range(merged_range.min_row, merged_range.max_row + 1):
                for col_index in range(merged_range.min_col, merged_range.max_col + 1):
                    cell = ws.cell(row=row_index, column=col_index)
                    cell.border = explicit_border(
                        left=left_side if col_index == merged_range.min_col else None,
                        right=right_side if col_index == merged_range.max_col else None,
                        top=top_side if row_index == merged_range.min_row else None,
                        bottom=bottom_side if row_index == merged_range.max_row else None,
                    )

        return merged_lookup

    def _normalize_alarm_row_heights(self, ws) -> None:
        family = self.export_profile.get("family")
        if family == "classic_combined" and not self._is_liquidcool_profile():
            start_row = self._alarm_sheet_spec().data_start_row
        else:
            start_row = 9
        if ws.max_row < start_row:
            return

        if self._is_liquidcool_profile() and ws.title == "报警状态":
            for row_index in range(start_row, ws.max_row + 1):
                if 9 <= row_index <= 25:
                    ws.row_dimensions[row_index].height = 15.0
                elif row_index == 26:
                    ws.row_dimensions[row_index].height = 36.5
                elif 27 <= row_index <= 32:
                    ws.row_dimensions[row_index].height = 15.0
                elif 33 <= row_index <= 36:
                    ws.row_dimensions[row_index].height = 340.0
                elif 37 <= row_index <= 40:
                    ws.row_dimensions[row_index].height = 200.0
                elif 41 <= row_index <= 112:
                    ws.row_dimensions[row_index].height = 210.0
                elif 113 <= row_index <= ws.max_row:
                    ws.row_dimensions[row_index].height = 200.0
            return

        if family == "classic_combined":
            short_rows_end = start_row + 9
            for row_index in range(start_row, ws.max_row + 1):
                has_content = any(ws.cell(row=row_index, column=col_index).value not in (None, "") for col_index in range(1, ws.max_column + 1))
                if not has_content:
                    continue
                ws.row_dimensions[row_index].height = 60.0 if row_index <= short_rows_end else 250.0
            return

        if family == "ab_screen_split":
            for row_index in range(start_row, ws.max_row + 1):
                has_content = any(ws.cell(row=row_index, column=col_index).value not in (None, "") for col_index in range(1, ws.max_column + 1))
                if not has_content:
                    continue
                ws.row_dimensions[row_index].height = 21.0 if row_index <= 18 else 250.05
            return

        if family == "extended_split":
            for row_index in range(start_row, ws.max_row + 1):
                has_content = any(ws.cell(row=row_index, column=col_index).value not in (None, "") for col_index in range(1, ws.max_column + 1))
                if not has_content:
                    continue
                if row_index <= 24:
                    ws.row_dimensions[row_index].height = 21.0
                    continue
                max_lines = 1
                for col_index in range(1, ws.max_column + 1):
                    value = ws.cell(row=row_index, column=col_index).value
                    if value not in (None, ""):
                        max_lines = max(max_lines, text_display_line_count(value))
                ws.row_dimensions[row_index].height = max(21.0, max_lines * 14.0 + 12.0)
            return

        base_height = 21.0
        line_height = 15.0
        padding = 3.0
        for row_index in range(start_row, ws.max_row + 1):
            max_lines = 1
            for col_index in range(1, ws.max_column + 1):
                value = ws.cell(row=row_index, column=col_index).value
                if value not in (None, ""):
                    max_lines = max(max_lines, text_display_line_count(value))
            height = base_height if max_lines <= 1 else max(base_height, max_lines * line_height + padding)
            ws.row_dimensions[row_index].height = height

    def _normalize_alarm_table_borders(self, ws, merged_lookup: set[tuple[int, int]] | None = None) -> None:
        if self.export_profile.get("family") == "classic_combined" and not self._is_liquidcool_profile():
            alarm_spec = self._alarm_sheet_spec()
            table_start_row = alarm_spec.header_row
            data_start_row = alarm_spec.data_start_row
        else:
            table_start_row = 7
            data_start_row = 9
        if ws.max_row < table_start_row or ws.max_column < 1:
            return

        if merged_lookup is None:
            merged_lookup = self._normalize_merged_range_borders(ws)

        for row_index in range(table_start_row, ws.max_row + 1):
            for col_index in range(1, ws.max_column + 1):
                if (row_index, col_index) in merged_lookup:
                    continue
                cell = ws.cell(row=row_index, column=col_index)
                value = cell.value
                if value in (None, "") and not (cell.fill and cell.fill.fill_type == "solid"):
                    continue
                row_height = ws.row_dimensions[row_index].height or 15.0
                if (
                    row_index >= data_start_row
                    and col_index in (8, 10, 11)
                    and value not in (None, "")
                    and row_height > 21.0
                ):
                    cell.border = explicit_border(
                        left=cell.border.left if getattr(cell.border.left, "style", None) else THIN,
                        right=cell.border.right if getattr(cell.border.right, "style", None) else THIN,
                        top=cell.border.top if getattr(cell.border.top, "style", None) else THIN,
                        bottom=None,
                    )

    def _apply_common_style(self, ws) -> None:
        style_spec = self._sheet_style_spec(ws)
        widths = style_spec.column_widths if style_spec is not None else DEFAULT_COLUMN_WIDTHS
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        # 冻结窗格在 workbook 保存前统一清理；同时重置 selection，避免
        # openpyxl 留下 selection.pane 但缺少 <pane> 节点的非法 XML。
        is_repeater_sheet = ws.title in {"中继器", "中继单元"}
        is_unified_main_sheet = self._is_unified_master() and ws.title == self._base_sheet_name()
        body_font_size = 11 if is_repeater_sheet else (10 if self._is_liquidcool_profile() else 11)
        top_line_font_name = "Times New Roman" if (is_repeater_sheet or not self._is_liquidcool_profile()) else "宋体"
        medium_border = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)
        is_alarm_sheet = "报警" in ws.title
        is_extended_start_sheet = self.export_profile.get("family") == "extended_split" and ws.title == "始端箱"
        if self._is_unified_master():
            top_line_cells = {"A1", "A2", "A3", "A4"}
        else:
            top_line_cells = {"A2", "A3"} if (self._is_liquidcool_profile() and is_repeater_sheet) else {"A1", "A3"}
        extended_start_label_rows = []
        if is_extended_start_sheet:
            extended_start_label_rows = [
                row_index
                for row_index in range(12, ws.max_row + 1)
                if ws.cell(row=row_index, column=10).value not in (None, "")
            ]
        extended_first_start_end = (
            extended_start_label_rows[1] - 1
            if len(extended_start_label_rows) > 1
            else (extended_start_label_rows[0] if extended_start_label_rows else None)
        )

        header_rows = {
            row_index
            for row_index in range(1, ws.max_row + 1)
            if str(ws.cell(row=row_index, column=1).value or "") == "通道号"
        }

        for row_index in range(1, ws.max_row + 1):
            for col_index in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_index, column=col_index)
                value = cell.value
                text = str(value or "")
                has_fill = bool(cell.fill and cell.fill.fill_type == "solid")
                has_border = any(
                    getattr(side, "style", None)
                    for side in (cell.border.left, cell.border.right, cell.border.top, cell.border.bottom)
                )
                if is_unified_main_sheet and col_index == 17:
                    # Hidden machine metadata is intentionally unstyled.
                    continue
                if (
                    is_unified_main_sheet
                    and 13 <= col_index <= 16
                    and value in (None, "")
                    and not has_fill
                    and not has_border
                ):
                    # Supplemental note columns are sparse. Avoid constructing
                    # thousands of identical styles for truly empty cells.
                    continue
                needs_border = has_fill or has_border or value not in (None, "")
                is_header_row = cell.row in header_rows

                font_name = "宋体"
                font_size = body_font_size
                bold = False
                italic = False
                color = deepcopy(FONT_COLOR_INDEXED_BLACK) if (value not in (None, "") or has_fill) else None
                horizontal = cell.alignment.horizontal or ("center" if value not in (None, "") else None)
                vertical = cell.alignment.vertical or ("center" if value not in (None, "") else None)
                wrap_text = bool(cell.alignment.wrap_text)
                default_side = THIN if needs_border else Side()
                border = Border(
                    left=cell.border.left if getattr(cell.border.left, "style", None) else default_side,
                    right=cell.border.right if getattr(cell.border.right, "style", None) else default_side,
                    top=cell.border.top if getattr(cell.border.top, "style", None) else default_side,
                    bottom=cell.border.bottom if getattr(cell.border.bottom, "style", None) else default_side,
                )
                var_name = str(ws.cell(row=row_index, column=2).value or "")
                prefix = var_name.rstrip("0123456789")

                if (
                    is_extended_start_sheet
                    and row_index in (9, 10)
                    and cell.column in (10, 11)
                    and value in (None, "")
                ):
                    font_name = "宋体"
                    font_size = 18
                    bold = True
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                    border = Border()
                elif is_extended_start_sheet and row_index == 10 and cell.column == 8 and value in (None, ""):
                    font_name = "宋体"
                    font_size = 11
                    bold = False
                    color = None
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                    border = Border(bottom=THIN)
                elif is_extended_start_sheet and row_index == 11 and cell.column == 8 and value in (None, ""):
                    font_name = "宋体"
                    font_size = 11
                    bold = True
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                elif is_extended_start_sheet and row_index == 11 and cell.column == 10 and value in (None, ""):
                    font_name = "宋体"
                    font_size = 18
                    bold = True
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                elif is_extended_start_sheet and row_index == 11 and cell.column == 11 and value in (None, ""):
                    font_name = "宋体"
                    font_size = 18
                    bold = True
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                    border = Border()
                elif is_extended_start_sheet and row_index == 11 and cell.column in (12, 13) and value in (None, ""):
                    font_name = "宋体"
                    font_size = 11
                    bold = False
                    color = None
                    horizontal = None
                    vertical = None
                    wrap_text = False
                    border = Border()
                elif (
                    is_extended_start_sheet
                    and extended_first_start_end is not None
                    and 16 <= row_index <= extended_first_start_end
                    and cell.column == 11
                    and value in (None, "")
                ):
                    font_name = "宋体"
                    font_size = 18
                    bold = True
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                    border = Border()
                elif (
                    is_extended_start_sheet
                    and row_index >= 12
                    and var_name
                    and cell.column == 8
                ):
                    font_name = "宋体"
                    font_size = 10
                    bold = text == "断路器状态字"
                    color = (
                        None
                        if prefix in {"In", "Pa", "Pb", "Pc", "P", "Qa", "Qb", "Qc", "Q", "Sa", "Sb", "Sc", "S_", "THDUa"}
                        else deepcopy(FONT_COLOR_INDEXED_BLACK)
                    )
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                    border = Border(
                        left=THIN,
                        right=THIN,
                        top=cell.border.top if getattr(cell.border.top, "style", None) else THIN,
                        bottom=cell.border.bottom if getattr(cell.border.bottom, "style", None) else THIN,
                    )
                elif (
                    is_extended_start_sheet
                    and row_index >= 12
                    and var_name
                    and cell.column == 9
                ):
                    font_name = "宋体"
                    font_size = 11 if prefix == "StateS" else 10
                    bold = False
                    if prefix == "StateS":
                        color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    elif prefix in {"LoadS", "UBS", "In", "Pa", "Pb", "Pc", "P", "Qa", "Qb", "Qc", "Q", "Sa", "Sb", "Sc", "S_", "Ta", "Tb", "Tc", "Tn", "THDUa"} and value not in (None, ""):
                        color = None
                    else:
                        color = deepcopy(FONT_COLOR_INDEXED_BLACK) if value not in (None, "") else None
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                    border = Border(
                        left=THIN,
                        right=THIN,
                        top=cell.border.top if getattr(cell.border.top, "style", None) else THIN,
                        bottom=cell.border.bottom if getattr(cell.border.bottom, "style", None) else THIN,
                    )
                elif cell.coordinate in top_line_cells and value not in (None, ""):
                    font_name = top_line_font_name
                    font_size = 10
                    bold = True
                    color = deepcopy(FONT_COLOR_RGB_BLACK)
                    horizontal = "center" if (self._is_liquidcool_profile() and is_repeater_sheet) else "left"
                    vertical = "center"
                    wrap_text = not (self._is_liquidcool_profile() and is_repeater_sheet)
                    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                elif text.startswith(("该项目", "该表", "该界面")):
                    font_name = "宋体"
                    font_size = body_font_size
                    bold = True
                    color = deepcopy(FONT_COLOR_RGB_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = True
                    border = Border(
                        left=THIN,
                        right=THIN,
                        top=cell.border.top if getattr(cell.border.top, "style", None) else Side(),
                        bottom=THIN,
                    )
                elif text.startswith(("注：", "说明：")):
                    font_name = "宋体"
                    font_size = body_font_size
                    bold = True
                    color = deepcopy(FONT_COLOR_RGB_RED)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = True
                elif text.startswith("下面的数据中各Bit位"):
                    font_name = "宋体"
                    font_size = 14
                    bold = True
                    color = deepcopy(FONT_COLOR_RGB_RED)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                    border = medium_border
                elif text.startswith("监控屏上默认设置"):
                    font_name = "宋体"
                    font_size = body_font_size
                    bold = True
                    color = deepcopy(FONT_COLOR_RGB_RED)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = True
                elif is_header_row:
                    font_name = "宋体"
                    font_size = body_font_size
                    bold = not self._is_liquidcool_profile()
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                elif text == "各Bit位含义":
                    font_name = "宋体"
                    font_size = 18
                    bold = True
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                elif text == "断路器状态字说明" or text == "始端箱断路器状态字说明":
                    font_name = "宋体"
                    font_size = 10
                    bold = True
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                elif is_repeater_sheet and cell.column == 9 and value not in (None, ""):
                    font_name = "宋体"
                    font_size = 10
                    bold = False
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                elif is_repeater_sheet and cell.column == 11 and value not in (None, ""):
                    font_name = "宋体"
                    font_size = 11
                    bold = True
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                elif (
                    is_extended_start_sheet
                    and cell.column in (12, 13)
                    and text in {"A相断路器", "B相断路器", "C相断路器", "1：闭合，0：断开"}
                ):
                    font_name = "宋体"
                    font_size = 10
                    bold = False
                    color = deepcopy(FONT_COLOR_RGB_RED)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                elif text.startswith("BIT") or text in {"A相断路器", "B相断路器", "C相断路器", "1：闭合，0：断开"}:
                    font_name = "宋体"
                    font_size = 10
                    bold = False
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                elif is_alarm_sheet and cell.column == 8 and value not in (None, ""):
                    font_name = "宋体"
                    font_size = body_font_size
                    bold = True
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = "\n" in text
                elif is_alarm_sheet and cell.column == 9 and value in (None, "") and ws.cell(row=row_index, column=2).value not in (None, ""):
                    font_name = "宋体"
                    font_size = 10
                    bold = False
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = False
                    border = Border(
                        left=THIN,
                        right=THIN,
                        top=cell.border.top if getattr(cell.border.top, "style", None) else THIN,
                        bottom=cell.border.bottom if getattr(cell.border.bottom, "style", None) else THIN,
                    )
                elif is_alarm_sheet and text.startswith("Bit"):
                    font_name = "宋体"
                    font_size = body_font_size
                    bold = False
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "left"
                    vertical = "center"
                    wrap_text = True
                    border = Border(
                        left=THIN,
                        right=THIN,
                        top=cell.border.top if getattr(cell.border.top, "style", None) else THIN,
                        bottom=cell.border.bottom if getattr(cell.border.bottom, "style", None) else THIN,
                    )
                elif (
                    is_unified_main_sheet
                    and cell.column in (11, 12)
                    and row_index >= self._combined_sheet_spec().data_start_row
                    and value not in (None, "")
                ):
                    font_name = "宋体"
                    font_size = 11 if cell.column == 11 else 12
                    bold = True
                    color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                    horizontal = "center"
                    vertical = "center"
                    wrap_text = True
                    border = medium_border
                elif (
                    not is_alarm_sheet
                    and
                    cell.column in (10, 11)
                    and (
                        "始端箱" in text
                        or "插接箱" in text
                        or "输出支路" in text
                        or "输出分路" in text
                    )
                    and text not in {"断路器状态字说明", "始端箱断路器状态字说明"}
                ):
                    if self._is_liquidcool_profile() and ws.title == "始端箱和插接箱":
                        font_name = "宋体"
                        font_size = 10
                        bold = False
                        color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                        horizontal = "left"
                        vertical = "center"
                        wrap_text = True
                        border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                    else:
                        font_name = "宋体"
                        font_size = 16
                        bold = True
                        color = deepcopy(FONT_COLOR_INDEXED_BLACK)
                        horizontal = "center"
                        vertical = "center"
                        wrap_text = True
                        if ws.title in {"始端箱和插接箱", "A路屏数据", "B路屏数据"}:
                            border = medium_border
                        else:
                            border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                elif value in (None, "") and not has_fill:
                    color = None

                cell.font = Font(
                    name=font_name,
                    size=font_size,
                    bold=bold,
                    italic=italic,
                    color=color,
                )
                cell.alignment = Alignment(
                    horizontal=horizontal,
                    vertical=vertical,
                    wrap_text=wrap_text,
                )
                cell.border = border

        merged_lookup = self._normalize_merged_range_borders(ws)
        if is_alarm_sheet:
            self._normalize_alarm_row_heights(ws)
            self._normalize_alarm_table_borders(ws, merged_lookup=merged_lookup)
