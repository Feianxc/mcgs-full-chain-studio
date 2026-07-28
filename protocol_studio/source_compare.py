from __future__ import annotations

import argparse
import json
import os
from collections import Counter
from difflib import SequenceMatcher
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from mvp_generator.library import PROTOCOL_RESOURCES_ROOT

PROTOCOL_STUDIO_ROOT = Path(__file__).resolve().parent
WORKSPACE_ROOT = PROTOCOL_STUDIO_ROOT.parent
SOURCE_COMPARE_ROOT = PROTOCOL_RESOURCES_ROOT / "source-compare"
WORKBOOK_METADATA_PATH = Path(
    os.environ.get(
        "PROTOCOL_STUDIO_SOURCE_COMPARE_METADATA",
        str(SOURCE_COMPARE_ROOT / "workbook-metadata.json"),
    )
).expanduser().resolve()
SOURCE_WORKBOOKS_ROOT = Path(
    os.environ.get(
        "PROTOCOL_STUDIO_SOURCE_WORKBOOKS_ROOT",
        str(SOURCE_COMPARE_ROOT / "workbooks"),
    )
).expanduser().resolve()
RUNS_ROOT = Path(
    os.environ.get(
        "PROTOCOL_STUDIO_RUNS_ROOT",
        str(PROTOCOL_STUDIO_ROOT / "runs"),
    )
).expanduser().resolve()

DEFAULT_SOURCE_PRIORITY = {
    "classic_combined": [],
    "classic_combined_two_columns": [],
    "classic_combined_liquidcool": [],
    "extended_split": [],
    "ab_screen_split": [],
}

EXPECTED_CANONICAL_ORDER = {
    "classic_combined": ["combined", "repeater", "alarm"],
    "classic_combined_two_columns": ["combined", "alarm"],
    "classic_combined_liquidcool": ["combined", "repeater", "alarm", "cabinet"],
    "extended_split": ["start", "plug", "cabinet", "repeater", "alarm"],
    "ab_screen_split": ["a_data", "a_alarm", "b_data", "b_alarm"],
}

CANONICAL_SHEET_NAME_MAP = {
    "总表": None,
    "始端箱和插接箱": "combined",
    "始端箱": "start",
    "插接箱": "plug",
    "单机柜数据": "cabinet",
    "中继器": "repeater",
    "中继单元": "repeater",
    "连接器测温": "repeater",
    "报警状态": "alarm",
    "A路屏数据": "a_data",
    "A路屏报警": "a_alarm",
    "B路屏数据": "b_data",
    "B路屏报警": "b_alarm",
}

LIQUIDCOOL_EXPORT_PROFILE_ID = "classic_combined_liquidcool_default"
LIQUIDCOOL_ADDRESS_PROFILE_ID = "classic_liquidcool_main1000_repeater5000_cabinet7000_alarm6000_32bit"
CLASSIC_TWO_COLUMNS_EXPORT_PROFILE_ID = "classic_combined_two_columns_default"

STATUS_RANK = {"diverged": 0, "close": 1, "match": 2}
DEFAULT_COLUMN_WIDTH = 13.0
DEFAULT_ROW_HEIGHT = 15.0
TOP_FILL_LIMIT = 8
STYLE_PROBE_MAX_COLUMNS = 15
STYLE_PROBE_MAX_ROWS = 400
DATA_SIGNATURE_MAX_COLUMNS = 10
DATA_SIGNATURE_MAX_ROWS = 500


def classify_compare_key(
    family: str | None,
    export_profile_id: str | None = None,
    address_profile_id: str | None = None,
) -> str | None:
    if not family:
        return None
    if family == "classic_combined" and export_profile_id == CLASSIC_TWO_COLUMNS_EXPORT_PROFILE_ID:
        return "classic_combined_two_columns"
    if family == "classic_combined" and (
        export_profile_id == LIQUIDCOOL_EXPORT_PROFILE_ID or address_profile_id == LIQUIDCOOL_ADDRESS_PROFILE_ID
    ):
        return "classic_combined_liquidcool"
    return family


def infer_compare_key_from_workbooks(
    family: str | None,
    generated_profile: dict[str, Any] | None = None,
    source_profile: dict[str, Any] | None = None,
    export_profile_id: str | None = None,
    address_profile_id: str | None = None,
) -> str:
    compare_key = classify_compare_key(
        family=family,
        export_profile_id=export_profile_id,
        address_profile_id=address_profile_id,
    )
    if compare_key and compare_key != family:
        return compare_key

    for profile in (generated_profile, source_profile):
        if not profile:
            continue
        canonical_order = profile.get("canonical_sheet_order", [])
        canonical_set = set(canonical_order)
        if {"a_data", "a_alarm", "b_data", "b_alarm"}.issubset(canonical_set):
            return "ab_screen_split"
        if {"combined", "repeater", "alarm", "cabinet"}.issubset(canonical_set):
            if export_profile_id == LIQUIDCOOL_EXPORT_PROFILE_ID or address_profile_id == LIQUIDCOOL_ADDRESS_PROFILE_ID:
                return "classic_combined_liquidcool"
            return family or "classic_combined"
        if {"start", "plug", "cabinet", "repeater", "alarm"}.issubset(canonical_set):
            return "extended_split"
        if "combined" in canonical_set and "alarm" in canonical_set and "repeater" not in canonical_set:
            return "classic_combined_two_columns"
        if {"combined", "repeater", "alarm"}.issubset(canonical_set):
            return "classic_combined"

    if family:
        return family
    raise ValueError("无法从生成工作簿结构推断 compare family，请显式提供 --family 或 profile ids。")


@lru_cache(maxsize=1)
def load_workbook_metadata() -> list[dict[str, Any]]:
    if not WORKBOOK_METADATA_PATH.exists():
        return []
    return json.loads(WORKBOOK_METADATA_PATH.read_text(encoding="utf-8"))


@lru_cache(maxsize=1)
def metadata_by_file_name() -> dict[str, dict[str, Any]]:
    return {item["file_name"]: item for item in load_workbook_metadata()}


def load_json(path: str | Path) -> Any:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def save_json(path: str | Path, payload: Any) -> None:
    Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")



def canonicalize_sheet_name(sheet_name: str) -> str | None:
    normalized = str(sheet_name).strip()
    if normalized in CANONICAL_SHEET_NAME_MAP:
        return CANONICAL_SHEET_NAME_MAP[normalized]
    if normalized.endswith("报警") and normalized.startswith("A路"):
        return "a_alarm"
    if normalized.endswith("报警") and normalized.startswith("B路"):
        return "b_alarm"
    if normalized.endswith("数据") and normalized.startswith("A路"):
        return "a_data"
    if normalized.endswith("数据") and normalized.startswith("B路"):
        return "b_data"
    if "中继" in normalized or "测温" in normalized:
        return "repeater"
    return normalized or None



def normalize_address(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None



def pick_representative_source_workbook(
    family: str,
    preferred_file_names: list[str] | None = None,
) -> Path:
    candidates: list[str] = []
    seen: set[str] = set()
    for name in (preferred_file_names or []) + DEFAULT_SOURCE_PRIORITY.get(family, []):
        if name and name not in seen:
            candidates.append(name)
            seen.add(name)
    for name in candidates:
        if Path(name).name != name:
            continue
        candidate = SOURCE_WORKBOOKS_ROOT / name
        if candidate.exists():
            return candidate
    raise FileNotFoundError(
        f"未配置 {family} 的外部参考协议；统一工作流不依赖历史客户 Excel"
    )



def normalize_text(value: Any) -> str:
    text = str(value or "")
    return " ".join(text.replace("\u3000", " ").replace("\n", " ").split()).strip()


def text_similarity(left: Any, right: Any) -> float:
    a = normalize_text(left)
    b = normalize_text(right)
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    return SequenceMatcher(None, a, b).ratio()


def round_score(value: float) -> float:
    return round(max(0.0, min(100.0, value)), 2)


def verdict_from_score(score: float, match_threshold: float = 92.0, close_threshold: float = 70.0) -> str:
    if score >= match_threshold:
        return "match"
    if score >= close_threshold:
        return "close"
    return "diverged"


def worst_status(*statuses: str) -> str:
    valid = [status for status in statuses if status in STATUS_RANK]
    if not valid:
        return "diverged"
    return min(valid, key=lambda item: STATUS_RANK[item])


def safe_float(value: Any, default: float | None = None) -> float | None:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except Exception:  # noqa: BLE001
        return default


def compare_numeric_maps(
    source_map: dict[str, float],
    generated_map: dict[str, float],
    tolerance: float,
    default_value: float,
) -> tuple[float, list[dict[str, Any]]]:
    if not source_map and not generated_map:
        return 1.0, []
    if not source_map:
        return 0.0, []

    details: list[dict[str, Any]] = []
    scores: list[float] = []
    for key in source_map:
        source_value = safe_float(source_map.get(key), default_value)
        generated_value = safe_float(generated_map.get(key), default_value)
        if source_value is None or generated_value is None:
            similarity = 0.0
        else:
            similarity = max(0.0, 1.0 - abs(source_value - generated_value) / max(tolerance, 1e-6))
        scores.append(similarity)
        details.append(
            {
                "key": key,
                "source": source_value,
                "generated": generated_value,
                "delta": None if source_value is None or generated_value is None else round(generated_value - source_value, 4),
                "similarity": round(similarity, 4),
            }
        )
    return (sum(scores) / len(scores) if scores else 1.0), details


def counter_similarity(source_counter: Counter[str], generated_counter: Counter[str]) -> float:
    if not source_counter and not generated_counter:
        return 1.0
    if not source_counter:
        return 0.0
    source_total = sum(source_counter.values())
    generated_total = sum(generated_counter.values())
    if source_total == 0:
        return 1.0 if generated_total == 0 else 0.0
    intersection = sum(min(source_counter[key], generated_counter.get(key, 0)) for key in source_counter)
    recall = intersection / source_total
    precision = intersection / generated_total if generated_total else 0.0
    if recall == 0.0 and precision == 0.0:
        return 0.0
    return 2 * recall * precision / (recall + precision)


def compare_text_lists(source_lines: list[str], generated_lines: list[str], limit: int = 6) -> float:
    source = [normalize_text(line) for line in source_lines if normalize_text(line)][:limit]
    generated = [normalize_text(line) for line in generated_lines if normalize_text(line)][:limit]
    if not source and not generated:
        return 1.0
    if not source:
        return 0.0
    scores = []
    for index, line in enumerate(source):
        candidate = generated[index] if index < len(generated) else ""
        scores.append(text_similarity(line, candidate))
    return sum(scores) / len(scores)


def compare_anchor_texts(source_anchors: dict[str, str], generated_anchors: dict[str, str]) -> float:
    if not source_anchors and not generated_anchors:
        return 1.0
    if not source_anchors:
        return 0.0
    scores = [text_similarity(text, generated_anchors.get(coord, "")) for coord, text in source_anchors.items()]
    return sum(scores) / len(scores) if scores else 1.0


def compare_exact_token_maps(source_map: dict[str, str], generated_map: dict[str, str]) -> float:
    if not source_map and not generated_map:
        return 1.0
    if not source_map:
        return 0.0
    source_counter = Counter(f"{coord}={value}" for coord, value in source_map.items())
    generated_counter = Counter(f"{coord}={value}" for coord, value in generated_map.items())
    return counter_similarity(source_counter, generated_counter)


def compare_row_signatures(source_rows: list[str], generated_rows: list[str]) -> float:
    source = [row for row in source_rows if row][:DATA_SIGNATURE_MAX_ROWS]
    generated = [row for row in generated_rows if row][:DATA_SIGNATURE_MAX_ROWS]
    if not source and not generated:
        return 1.0
    if not source:
        return 0.0
    left = "\n".join(source)
    right = "\n".join(generated)
    return SequenceMatcher(None, left, right).ratio()


def normalize_color(color: Any) -> str | None:
    if color is None:
        return None
    try:
        color_type = getattr(color, "type", None)
        if color_type == "rgb":
            value = getattr(color, "rgb", None)
        elif color_type == "indexed":
            value = getattr(color, "indexed", None)
        elif color_type == "theme":
            value = f"theme:{getattr(color, 'theme', None)}"
        elif color_type == "auto":
            value = "auto"
        else:
            value = getattr(color, "value", None) or getattr(color, "rgb", None) or getattr(color, "indexed", None)
    except Exception:  # noqa: BLE001
        value = None
    if value in (None, ""):
        return None
    text = str(value).upper()
    if len(text) == 8 and text.startswith("00"):
        text = f"FF{text[2:]}"
    return text


def normalize_font_signature(cell: Any) -> str | None:
    try:
        font = cell.font
    except Exception:  # noqa: BLE001
        return None
    if font is None:
        return None
    color = normalize_color(getattr(font, "color", None))
    return "|".join(
        [
            str(font.name or ""),
            str(font.sz or ""),
            "1" if bool(font.bold) else "0",
            "1" if bool(font.italic) else "0",
            color or "",
        ]
    )


def normalize_alignment_signature(cell: Any) -> str | None:
    try:
        alignment = cell.alignment
    except Exception:  # noqa: BLE001
        return None
    if alignment is None:
        return None
    return "|".join(
        [
            str(alignment.horizontal or ""),
            str(alignment.vertical or ""),
            "1" if bool(alignment.wrap_text) else "0",
            str(alignment.text_rotation or 0),
        ]
    )


def normalize_border_signature(cell: Any) -> str | None:
    try:
        border = cell.border
    except Exception:  # noqa: BLE001
        return None
    if border is None:
        return None
    return "|".join(
        [
            str(getattr(border.left, "style", "") or ""),
            str(getattr(border.right, "style", "") or ""),
            str(getattr(border.top, "style", "") or ""),
            str(getattr(border.bottom, "style", "") or ""),
        ]
    )


def _style_probe_bounds(ws, header_row: int | None, max_columns: int = STYLE_PROBE_MAX_COLUMNS) -> tuple[int, int]:
    upper_row = min(ws.max_row, max(120, (header_row or 0) + 220, 40, STYLE_PROBE_MAX_ROWS))
    upper_col = min(max_columns, max(1, ws.max_column))
    explicit_columns = []
    for key in ws.column_dimensions.keys():
        if not isinstance(key, str) or ":" in key:
            continue
        try:
            explicit_columns.append(column_index_from_string(key))
        except ValueError:
            continue
    if explicit_columns:
        upper_col = min(max_columns, max(upper_col, max(explicit_columns)))
    merged_max_col = max((merge.max_col for merge in ws.merged_cells.ranges), default=upper_col)
    upper_col = min(max_columns, max(upper_col, merged_max_col))
    return upper_row, upper_col


def _find_header_row(ws) -> dict[str, Any]:
    max_probe_columns = min(24, max(8, ws.max_column))
    max_probe_rows = min(40, ws.max_row)
    for row_index in range(1, max_probe_rows + 1):
        values = [ws.cell(row=row_index, column=col).value for col in range(1, max_probe_columns + 1)]
        labels = [str(value).strip() if value is not None else "" for value in values]
        if "变量名" in labels and "寄存器地址" in labels:
            return {
                "row": row_index,
                "labels": labels,
                "column_map": {label: idx + 1 for idx, label in enumerate(labels) if label},
            }
    return {"row": None, "labels": [], "column_map": {}}



def _collect_intro_lines(ws, before_row: int | None, limit: int = 6) -> list[str]:
    if not before_row or before_row <= 1:
        return []
    intro_lines: list[str] = []
    for row_index in range(1, before_row):
        value = ws.cell(row=row_index, column=1).value
        if value in (None, ""):
            continue
        intro_lines.append(normalize_text(value))
        if len(intro_lines) >= limit:
            break
    return intro_lines



def _prefixes(lines: list[str], limit: int = 18) -> list[str]:
    return [line[:limit] for line in lines if line]



def _collect_anchor_texts(ws, header_row: int | None) -> dict[str, str]:
    if header_row is None:
        limit_row = min(ws.max_row, 12)
    else:
        limit_row = min(ws.max_row, max(12, header_row))
    limit_col = min(ws.max_column, 15)
    anchors: dict[str, str] = {}
    for row_index in range(1, limit_row + 1):
        for col_index in range(1, limit_col + 1):
            value = ws.cell(row=row_index, column=col_index).value
            text = normalize_text(value)
            if not text:
                continue
            anchors[f"{get_column_letter(col_index)}{row_index}"] = text[:80]
    return anchors



def _collect_merge_profile(ws, header_row: int | None) -> dict[str, Any]:
    top_limit = min(ws.max_row, max(12, (header_row or 0) + 1))
    top_merges: list[str] = []
    all_merges: list[str] = []
    data_exact: list[str] = []
    data_patterns: Counter[str] = Counter()
    for merge_range in ws.merged_cells.ranges:
        merge_text = str(merge_range)
        all_merges.append(merge_text)
        min_col, min_row, max_col, max_row = merge_range.bounds
        if max_row <= top_limit:
            top_merges.append(merge_text)
        else:
            data_exact.append(merge_text)
            pattern = f"{min_col}:{max_col}:{max_row - min_row + 1}"
            data_patterns[pattern] += 1
    return {
        "count": len(all_merges),
        "all": all_merges,
        "top": sorted(top_merges),
        "data_exact": sorted(data_exact),
        "data_patterns": dict(sorted(data_patterns.items())),
    }
def _effective_max_column(ws, max_columns: int | None = None) -> int:
    candidates = [max(1, int(ws.max_column or 1))]
    for key, dimension in ws.column_dimensions.items():
        if not key or ":" in str(key):
            continue
        try:
            column_index = column_index_from_string(str(key))
        except ValueError:
            continue
        if (
            dimension.width is not None
            or getattr(dimension, "hidden", False)
            or getattr(dimension, "bestFit", False)
            or getattr(dimension, "style", None)
        ):
            candidates.append(column_index)
    for merge_range in ws.merged_cells.ranges:
        _, _, max_col, _ = merge_range.bounds
        candidates.append(max_col)
    upper_bound = max(candidates)
    if max_columns is not None:
        upper_bound = min(upper_bound, max_columns)
    return max(1, upper_bound)


def _effective_max_row(ws) -> int:
    candidates = [max(1, int(ws.max_row or 1))]
    for row_index, dimension in ws.row_dimensions.items():
        if not isinstance(row_index, int):
            continue
        if (
            dimension.height is not None
            or getattr(dimension, "hidden", False)
            or getattr(dimension, "style", None)
        ):
            candidates.append(row_index)
    for merge_range in ws.merged_cells.ranges:
        _, _, _, max_row = merge_range.bounds
        candidates.append(max_row)
    return max(candidates)


def _extract_column_widths(ws, max_columns: int = 15) -> dict[str, float]:
    widths: dict[str, float] = {}
    upper_bound = _effective_max_column(ws, max_columns=max_columns)
    for index in range(1, upper_bound + 1):
        letter = get_column_letter(index)
        width = ws.column_dimensions[letter].width
        if width is None:
            width = DEFAULT_COLUMN_WIDTH
        widths[letter] = round(float(width), 4)
    return widths



def _extract_row_heights(ws, header_row: int | None, extra_rows: int = 3) -> dict[str, float]:
    effective_max_row = _effective_max_row(ws)
    if header_row is None:
        upper_bound = min(effective_max_row, 12)
    else:
        upper_bound = min(effective_max_row, max(12, header_row + extra_rows))
    heights: dict[str, float] = {}
    for row_index in range(1, upper_bound + 1):
        height = ws.row_dimensions[row_index].height
        if height is None:
            height = DEFAULT_ROW_HEIGHT
        heights[str(row_index)] = round(float(height), 4)
    return heights



def _extract_fill_histogram(ws, header_row: int | None, max_columns: int = 15) -> dict[str, int]:
    upper_row = min(_effective_max_row(ws), max(40, (header_row or 0) + 40))
    upper_col = _effective_max_column(ws, max_columns=max_columns)
    counter: Counter[str] = Counter()
    for row_index in range(1, upper_row + 1):
        for col_index in range(1, upper_col + 1):
            cell = ws.cell(row=row_index, column=col_index)
            fill = cell.fill
            if not fill or fill.fill_type != "solid":
                continue
            color = normalize_color(fill.fgColor) or normalize_color(fill.start_color)
            if not color:
                continue
            counter[color] += 1
    return dict(counter.most_common(TOP_FILL_LIMIT))


def _extract_fill_grid(ws, header_row: int | None, max_columns: int = STYLE_PROBE_MAX_COLUMNS) -> dict[str, str]:
    upper_row, upper_col = _style_probe_bounds(ws, header_row, max_columns=max_columns)
    grid: dict[str, str] = {}
    for row_index in range(1, upper_row + 1):
        for col_index in range(1, upper_col + 1):
            cell = ws.cell(row=row_index, column=col_index)
            fill = cell.fill
            if not fill or fill.fill_type != "solid":
                continue
            color = normalize_color(fill.fgColor) or normalize_color(fill.start_color)
            if not color:
                continue
            grid[f"{get_column_letter(col_index)}{row_index}"] = color
    return grid


def _extract_style_histogram(
    ws,
    header_row: int | None,
    extractor: Any,
    max_columns: int = STYLE_PROBE_MAX_COLUMNS,
    top_limit: int = TOP_FILL_LIMIT,
) -> dict[str, int]:
    upper_row, upper_col = _style_probe_bounds(ws, header_row, max_columns=max_columns)
    counter: Counter[str] = Counter()
    for row_index in range(1, upper_row + 1):
        for col_index in range(1, upper_col + 1):
            signature = extractor(ws.cell(row=row_index, column=col_index))
            if not signature:
                continue
            counter[signature] += 1
    return dict(counter.most_common(top_limit))


def _extract_data_row_signatures(
    ws,
    header_row: int | None,
    max_columns: int = DATA_SIGNATURE_MAX_COLUMNS,
    max_rows: int = DATA_SIGNATURE_MAX_ROWS,
) -> list[str]:
    if header_row is None:
        return []
    limit_col = min(max_columns, max(1, ws.max_column))
    limit_row = min(ws.max_row, header_row + max_rows)
    signatures: list[str] = []
    for row_index in range(header_row + 1, limit_row + 1):
        values = [normalize_text(ws.cell(row=row_index, column=col).value) for col in range(1, limit_col + 1)]
        if not any(values):
            continue
        signatures.append(" | ".join(values))
    return signatures



def analyze_sheet(ws) -> dict[str, Any]:
    header = _find_header_row(ws)
    header_row = header["row"]
    intro_lines = _collect_intro_lines(ws, header_row)
    sheet_info = {
        "sheet_name": ws.title,
        "canonical_name": canonicalize_sheet_name(ws.title),
        "header_row": header_row,
        "header_labels": [label for label in header.get("labels", []) if label],
        "intro_lines": intro_lines,
        "intro_prefixes": _prefixes(intro_lines),
        "anchor_texts": _collect_anchor_texts(ws, header_row),
        "merge_profile": _collect_merge_profile(ws, header_row),
        "column_widths": _extract_column_widths(ws),
        "row_heights": _extract_row_heights(ws, header_row),
        "fill_histogram": _extract_fill_histogram(ws, header_row),
        "fill_grid": _extract_fill_grid(ws, header_row),
        "font_histogram": _extract_style_histogram(ws, header_row, normalize_font_signature),
        "alignment_histogram": _extract_style_histogram(ws, header_row, normalize_alignment_signature),
        "border_histogram": _extract_style_histogram(ws, header_row, normalize_border_signature),
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "first_data_row": None,
        "first_var_name": None,
        "first_address": None,
        "data_row_count": 0,
        "data_row_signatures": _extract_data_row_signatures(ws, header_row),
        "max_row": ws.max_row,
        "max_column": ws.max_column,
    }
    if header_row is None:
        return sheet_info

    column_map = header["column_map"]
    variable_column = column_map.get("变量名")
    address_column = column_map.get("寄存器地址")
    if variable_column is None or address_column is None:
        return sheet_info

    for row_index in range(header_row + 1, ws.max_row + 1):
        var_value = ws.cell(row=row_index, column=variable_column).value
        if var_value in (None, ""):
            continue
        var_text = normalize_text(var_value)
        if not var_text:
            continue
        sheet_info["data_row_count"] += 1
        if sheet_info["first_data_row"] is None:
            sheet_info["first_data_row"] = row_index
            sheet_info["first_var_name"] = var_text
            sheet_info["first_address"] = normalize_address(ws.cell(row=row_index, column=address_column).value)
    return sheet_info



def profile_workbook(path: Path, family: str | None = None) -> dict[str, Any]:
    expected_order = EXPECTED_CANONICAL_ORDER.get(family)
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        sheets: list[dict[str, Any]] = []
        for ws in workbook.worksheets:
            canonical_name = canonicalize_sheet_name(ws.title)
            if canonical_name is None:
                continue
            if expected_order and canonical_name not in expected_order:
                continue
            sheet_info = analyze_sheet(ws)
            sheets.append(sheet_info)
        return {
            "path": path.name,
            "file_name": path.name,
            "sheet_names": [sheet["sheet_name"] for sheet in sheets],
            "canonical_sheet_order": [sheet["canonical_name"] for sheet in sheets],
            "sheets": sheets,
        }
    finally:
        workbook.close()


@lru_cache(maxsize=16)
def profile_workbook_cached(path_text: str, family: str | None = None) -> dict[str, Any]:
    return profile_workbook(Path(path_text), family=family)



def _compare_intro(source_sheet: dict[str, Any], generated_sheet: dict[str, Any]) -> tuple[bool, float]:
    score = compare_text_lists(source_sheet.get("intro_lines", []), generated_sheet.get("intro_lines", []), limit=6)
    return score >= 0.86, score



def _compare_header_labels(source_sheet: dict[str, Any], generated_sheet: dict[str, Any]) -> tuple[bool, float]:
    source_labels = source_sheet.get("header_labels", [])
    generated_labels = generated_sheet.get("header_labels", [])
    score = text_similarity(" | ".join(source_labels), " | ".join(generated_labels))
    return score >= 0.96, score



def build_priority_fix(metric_key: str, label: str, similarity: float, weight: float, note: str) -> dict[str, Any]:
    gap = max(0.0, 1.0 - similarity)
    return {
        "metric": metric_key,
        "label": label,
        "similarity": round(similarity, 4),
        "score": round_score(similarity * 100),
        "impact": round(weight * gap, 4),
        "note": note,
    }



def compare_sheet_pair(source_sheet: dict[str, Any], generated_sheet: dict[str, Any]) -> dict[str, Any]:
    notes: list[str] = []
    canonical_match = source_sheet.get("canonical_name") == generated_sheet.get("canonical_name")
    title_exact = source_sheet.get("sheet_name") == generated_sheet.get("sheet_name")
    header_row_delta = None
    if source_sheet.get("header_row") is not None and generated_sheet.get("header_row") is not None:
        header_row_delta = abs(int(source_sheet["header_row"]) - int(generated_sheet["header_row"]))
    first_var_match = source_sheet.get("first_var_name") == generated_sheet.get("first_var_name")
    first_address_match = source_sheet.get("first_address") == generated_sheet.get("first_address")
    intro_match, intro_similarity = _compare_intro(source_sheet, generated_sheet)
    header_labels_match, header_labels_similarity = _compare_header_labels(source_sheet, generated_sheet)

    source_count = int(source_sheet.get("data_row_count") or 0)
    generated_count = int(generated_sheet.get("data_row_count") or 0)
    row_count_ratio = None
    row_ratio_score = 0.0
    if source_count > 0:
        row_count_ratio = round(generated_count / source_count, 4)
        row_ratio_score = max(0.0, 1.0 - abs(1.0 - generated_count / source_count))
    data_row_similarity = compare_row_signatures(
        source_sheet.get("data_row_signatures", []),
        generated_sheet.get("data_row_signatures", []),
    )

    if not canonical_match:
        notes.append("sheet canonical 类别不一致")
    if header_row_delta not in (None, 0):
        notes.append(f"header 行差 {header_row_delta}")
    if not header_labels_match:
        notes.append("表头文本或列位不一致")
    if not first_var_match:
        notes.append("首变量名不一致")
    if not first_address_match:
        notes.append("首地址不一致")
    if row_count_ratio is not None and not 0.9 <= row_count_ratio <= 1.1:
        notes.append(f"数据行规模差异较大（ratio={row_count_ratio}）")
    if not intro_match:
        notes.append("说明区文案前缀存在差异")
    if data_row_similarity < 0.95:
        notes.append("数据区行序列 / 关键文案未贴源")

    structure_components = {
        "canonical": 1.0 if canonical_match else 0.0,
        "title": 1.0 if title_exact else 0.0,
        "header_row": 1.0 if header_row_delta in (None, 0) else max(0.0, 1.0 - (header_row_delta / 3.0)),
        "header_labels": header_labels_similarity,
        "first_var": 1.0 if first_var_match else 0.0,
        "first_address": 1.0 if first_address_match else 0.0,
        "intro": intro_similarity,
        "row_ratio": row_ratio_score,
        "data_rows": data_row_similarity,
    }
    structure_weights = {
        "canonical": 20,
        "title": 5,
        "header_row": 12,
        "header_labels": 12,
        "first_var": 10,
        "first_address": 10,
        "intro": 8,
        "row_ratio": 6,
        "data_rows": 17,
    }
    structure_score = round_score(
        sum(structure_components[key] * structure_weights[key] for key in structure_weights)
        / sum(structure_weights.values())
        * 100
    )

    if (
        canonical_match
        and header_row_delta in (0, 1)
        and first_var_match
        and first_address_match
        and header_labels_similarity >= 0.96
        and data_row_similarity >= 0.95
    ):
        structure_status = (
            "match"
            if (
                title_exact
                and intro_match
                and row_count_ratio is not None
                and 0.9 <= row_count_ratio <= 1.1
                and data_row_similarity >= 0.985
            )
            else "close"
        )
    else:
        structure_status = "diverged"

    source_merges = source_sheet.get("merge_profile", {})
    generated_merges = generated_sheet.get("merge_profile", {})
    top_merge_similarity = counter_similarity(Counter(source_merges.get("top", [])), Counter(generated_merges.get("top", [])))
    data_merge_similarity = counter_similarity(
        Counter(source_merges.get("data_patterns", {})),
        Counter(generated_merges.get("data_patterns", {})),
    )
    exact_data_merge_similarity = counter_similarity(
        Counter(source_merges.get("data_exact", [])),
        Counter(generated_merges.get("data_exact", [])),
    )
    width_similarity, width_details = compare_numeric_maps(
        source_sheet.get("column_widths", {}),
        generated_sheet.get("column_widths", {}),
        tolerance=3.0,
        default_value=DEFAULT_COLUMN_WIDTH,
    )
    row_height_similarity, row_height_details = compare_numeric_maps(
        source_sheet.get("row_heights", {}),
        generated_sheet.get("row_heights", {}),
        tolerance=10.0,
        default_value=DEFAULT_ROW_HEIGHT,
    )
    fill_similarity = counter_similarity(
        Counter(source_sheet.get("fill_histogram", {})),
        Counter(generated_sheet.get("fill_histogram", {})),
    )
    fill_position_similarity = compare_exact_token_maps(
        source_sheet.get("fill_grid", {}),
        generated_sheet.get("fill_grid", {}),
    )
    font_similarity = counter_similarity(
        Counter(source_sheet.get("font_histogram", {})),
        Counter(generated_sheet.get("font_histogram", {})),
    )
    alignment_similarity = counter_similarity(
        Counter(source_sheet.get("alignment_histogram", {})),
        Counter(generated_sheet.get("alignment_histogram", {})),
    )
    border_similarity = counter_similarity(
        Counter(source_sheet.get("border_histogram", {})),
        Counter(generated_sheet.get("border_histogram", {})),
    )
    anchor_similarity = compare_anchor_texts(
        source_sheet.get("anchor_texts", {}),
        generated_sheet.get("anchor_texts", {}),
    )
    freeze_panes_similarity = (
        1.0 if source_sheet.get("freeze_panes") == generated_sheet.get("freeze_panes") else 0.0
    )

    format_components = {
        "header_text": header_labels_similarity,
        "intro_copy": intro_similarity,
        "anchor_texts": anchor_similarity,
        "top_merges": top_merge_similarity,
        "data_merges": data_merge_similarity,
        "exact_data_merges": exact_data_merge_similarity,
        "column_widths": width_similarity,
        "row_heights": row_height_similarity,
        "fill_positions": fill_position_similarity,
        "fills": fill_similarity,
        "fonts": font_similarity,
        "alignments": alignment_similarity,
        "borders": border_similarity,
        "freeze_panes": freeze_panes_similarity,
    }
    format_weights = {
        "header_text": 7,
        "intro_copy": 7,
        "anchor_texts": 9,
        "top_merges": 10,
        "data_merges": 6,
        "exact_data_merges": 10,
        "column_widths": 9,
        "row_heights": 7,
        "fill_positions": 12,
        "fills": 5,
        "fonts": 6,
        "alignments": 6,
        "borders": 6,
        "freeze_panes": 4,
    }
    format_score = round_score(
        sum(format_components[key] * format_weights[key] for key in format_weights)
        / sum(format_weights.values())
        * 100
    )
    format_status = verdict_from_score(format_score, match_threshold=93.0, close_threshold=72.0)

    overall_score = round_score((structure_score * 0.35) + (format_score * 0.65))
    combined_status = verdict_from_score(overall_score, match_threshold=92.0, close_threshold=70.0)

    priority_fixes = [
        build_priority_fix("top_merges", "顶部合并区块", top_merge_similarity, format_weights["top_merges"], "先贴齐 intro / 表头 / 说明区的固定合并区。"),
        build_priority_fix("data_merges", "数据区合并模式", data_merge_similarity, format_weights["data_merges"], "重点看状态字说明列、组标签列、J01/J02 / A/B 路合并模式。"),
        build_priority_fix("exact_data_merges", "精确 merge ranges", exact_data_merge_similarity, format_weights["exact_data_merges"], "同一种 merge pattern 还不够，需要把起止行列也贴齐。"),
        build_priority_fix("data_rows", "数据区行序列/文案", data_row_similarity, structure_weights["data_rows"], "逐行贴齐变量名、描述、单位、序列与中继/插接箱命名。"),
        build_priority_fix("column_widths", "列宽", width_similarity, format_weights["column_widths"], "优先同步模板列宽，不要只靠默认宽度。"),
        build_priority_fix("row_heights", "行高", row_height_similarity, format_weights["row_heights"], "优先同步顶区、表头、说明区行高。"),
        build_priority_fix("fill_positions", "精确填充位置", fill_position_similarity, format_weights["fill_positions"], "不仅要颜色数量接近，还要把具体坐标位置贴齐。"),
        build_priority_fix("fills", "颜色/填充", fill_similarity, format_weights["fills"], "同步表头色、状态区色、阈值区块色与交替底色。"),
        build_priority_fix("fonts", "字体", font_similarity, format_weights["fonts"], "同步宋体/粗体/字号与重点文案字体。"),
        build_priority_fix("alignments", "对齐", alignment_similarity, format_weights["alignments"], "同步左/中对齐、换行与旋转设置。"),
        build_priority_fix("borders", "边框", border_similarity, format_weights["borders"], "同步表头、数据区和说明区的边框样式。"),
        build_priority_fix("freeze_panes", "冻结窗格", freeze_panes_similarity, format_weights["freeze_panes"], "如果目标是模板一模一样，freeze panes 也需要贴齐。"),
        build_priority_fix("anchor_texts", "固定文案锚点", anchor_similarity, format_weights["anchor_texts"], "同步 A1/A3/A5/A8/A9/J11 等关键位置文案与列标题。"),
        build_priority_fix("intro_copy", "intro 文案", intro_similarity, format_weights["intro_copy"], "说明区应直接贴源模板，不要泛化改写。"),
        build_priority_fix("header_text", "表头文本", header_labels_similarity, format_weights["header_text"], "列名、列位、辅助列标题必须贴源。"),
    ]
    priority_fixes = sorted(priority_fixes, key=lambda item: item["impact"], reverse=True)

    major_differences: list[str] = []
    if top_merge_similarity < 0.9:
        major_differences.append("顶部合并区块与源模板不一致")
    if data_merge_similarity < 0.85:
        major_differences.append("数据区的组标签 / 状态字合并模式不一致")
    if exact_data_merge_similarity < 0.9:
        major_differences.append("精确 merge ranges 仍与源模板不一致")
    if width_similarity < 0.9:
        major_differences.append("列宽贴源度不足")
    if row_height_similarity < 0.9:
        major_differences.append("行高贴源度不足")
    if fill_position_similarity < 0.9:
        major_differences.append("填充坐标位置与源模板不一致")
    if fill_similarity < 0.8:
        major_differences.append("颜色与填充策略差异较大")
    if font_similarity < 0.9:
        major_differences.append("字体样式分布与源模板不一致")
    if alignment_similarity < 0.9:
        major_differences.append("对齐 / 换行设置与源模板不一致")
    if border_similarity < 0.9:
        major_differences.append("边框样式与源模板不一致")
    if freeze_panes_similarity < 1.0:
        major_differences.append("freeze panes 与源模板不一致")
    if anchor_similarity < 0.85:
        major_differences.append("固定文案锚点与关键标签位置存在偏差")
    if intro_similarity < 0.9:
        major_differences.append("intro / 说明区文案没有完全贴源")
    if header_labels_similarity < 0.95:
        major_differences.append("表头文本或辅助列标题存在偏差")

    return {
        "canonical_name": generated_sheet.get("canonical_name") or source_sheet.get("canonical_name"),
        "sheet_name": generated_sheet.get("sheet_name") or source_sheet.get("sheet_name"),
        "source_sheet": source_sheet.get("sheet_name"),
        "generated_sheet": generated_sheet.get("sheet_name"),
        "status": combined_status,
        "verdict": combined_status,
        "structure_status": structure_status,
        "format_status": format_status,
        "combined_status": combined_status,
        "overall_score": overall_score,
        "structure_score": structure_score,
        "format_score": format_score,
        "title_exact": title_exact,
        "canonical_match": canonical_match,
        "header_row_delta": header_row_delta,
        "header_labels_match": header_labels_match,
        "header_labels_similarity": round(header_labels_similarity, 4),
        "source_header_row": source_sheet.get("header_row"),
        "generated_header_row": generated_sheet.get("header_row"),
        "first_var_match": first_var_match,
        "source_first_var_name": source_sheet.get("first_var_name"),
        "generated_first_var_name": generated_sheet.get("first_var_name"),
        "first_address_match": first_address_match,
        "source_first_address": source_sheet.get("first_address"),
        "generated_first_address": generated_sheet.get("first_address"),
        "source_data_row_count": source_count,
        "generated_data_row_count": generated_count,
        "row_count_ratio": row_count_ratio,
        "data_row_similarity": round(data_row_similarity, 4),
        "intro_match": intro_match,
        "intro_similarity": round(intro_similarity, 4),
        "source_intro_prefixes": source_sheet.get("intro_prefixes", []),
        "generated_intro_prefixes": generated_sheet.get("intro_prefixes", []),
        "source_intro_lines": source_sheet.get("intro_lines", []),
        "generated_intro_lines": generated_sheet.get("intro_lines", []),
        "major_differences": major_differences,
        "priority_fixes": priority_fixes[:5],
        "format_breakdown": {
            "header_text": round_score(header_labels_similarity * 100),
            "intro_copy": round_score(intro_similarity * 100),
            "anchor_texts": round_score(anchor_similarity * 100),
            "top_merges": round_score(top_merge_similarity * 100),
            "data_merges": round_score(data_merge_similarity * 100),
            "exact_data_merges": round_score(exact_data_merge_similarity * 100),
            "column_widths": round_score(width_similarity * 100),
            "row_heights": round_score(row_height_similarity * 100),
            "fill_positions": round_score(fill_position_similarity * 100),
            "fills": round_score(fill_similarity * 100),
            "fonts": round_score(font_similarity * 100),
            "alignments": round_score(alignment_similarity * 100),
            "borders": round_score(border_similarity * 100),
            "freeze_panes": round_score(freeze_panes_similarity * 100),
        },
        "format_details": {
            "merge_similarity": {
                "top": round(top_merge_similarity, 4),
                "data": round(data_merge_similarity, 4),
                "data_exact": round(exact_data_merge_similarity, 4),
                "source_top_merges": source_merges.get("top", []),
                "generated_top_merges": generated_merges.get("top", []),
                "source_data_exact": source_merges.get("data_exact", []),
                "generated_data_exact": generated_merges.get("data_exact", []),
                "source_data_patterns": source_merges.get("data_patterns", {}),
                "generated_data_patterns": generated_merges.get("data_patterns", {}),
            },
            "column_widths": width_details,
            "row_heights": row_height_details,
            "fills": {
                "source": source_sheet.get("fill_histogram", {}),
                "generated": generated_sheet.get("fill_histogram", {}),
                "similarity": round(fill_similarity, 4),
            },
            "fill_positions": {
                "source": source_sheet.get("fill_grid", {}),
                "generated": generated_sheet.get("fill_grid", {}),
                "similarity": round(fill_position_similarity, 4),
            },
            "fonts": {
                "source": source_sheet.get("font_histogram", {}),
                "generated": generated_sheet.get("font_histogram", {}),
                "similarity": round(font_similarity, 4),
            },
            "alignments": {
                "source": source_sheet.get("alignment_histogram", {}),
                "generated": generated_sheet.get("alignment_histogram", {}),
                "similarity": round(alignment_similarity, 4),
            },
            "borders": {
                "source": source_sheet.get("border_histogram", {}),
                "generated": generated_sheet.get("border_histogram", {}),
                "similarity": round(border_similarity, 4),
            },
            "freeze_panes": {
                "source": source_sheet.get("freeze_panes"),
                "generated": generated_sheet.get("freeze_panes"),
                "similarity": round(freeze_panes_similarity, 4),
            },
            "anchor_texts": {
                "source": source_sheet.get("anchor_texts", {}),
                "generated": generated_sheet.get("anchor_texts", {}),
                "similarity": round(anchor_similarity, 4),
            },
            "data_rows": {
                "source_preview": source_sheet.get("data_row_signatures", [])[:12],
                "generated_preview": generated_sheet.get("data_row_signatures", [])[:12],
                "similarity": round(data_row_similarity, 4),
            },
        },
        "source": {
            "header_row": source_sheet.get("header_row"),
            "data_start_row": source_sheet.get("first_data_row"),
            "first_variable": source_sheet.get("first_var_name"),
            "first_address": source_sheet.get("first_address"),
            "row_count": source_count,
            "intro_lines": source_sheet.get("intro_lines", []),
            "column_widths": source_sheet.get("column_widths", {}),
            "row_heights": source_sheet.get("row_heights", {}),
            "merge_profile": source_sheet.get("merge_profile", {}),
            "fill_histogram": source_sheet.get("fill_histogram", {}),
            "fill_grid": source_sheet.get("fill_grid", {}),
            "font_histogram": source_sheet.get("font_histogram", {}),
            "alignment_histogram": source_sheet.get("alignment_histogram", {}),
            "border_histogram": source_sheet.get("border_histogram", {}),
            "freeze_panes": source_sheet.get("freeze_panes"),
        },
        "generated": {
            "header_row": generated_sheet.get("header_row"),
            "data_start_row": generated_sheet.get("first_data_row"),
            "first_variable": generated_sheet.get("first_var_name"),
            "first_address": generated_sheet.get("first_address"),
            "row_count": generated_count,
            "intro_lines": generated_sheet.get("intro_lines", []),
            "column_widths": generated_sheet.get("column_widths", {}),
            "row_heights": generated_sheet.get("row_heights", {}),
            "merge_profile": generated_sheet.get("merge_profile", {}),
            "fill_histogram": generated_sheet.get("fill_histogram", {}),
            "fill_grid": generated_sheet.get("fill_grid", {}),
            "font_histogram": generated_sheet.get("font_histogram", {}),
            "alignment_histogram": generated_sheet.get("alignment_histogram", {}),
            "border_histogram": generated_sheet.get("border_histogram", {}),
            "freeze_panes": generated_sheet.get("freeze_panes"),
        },
        "notes": notes + major_differences,
    }



def compare_workbooks(
    generated_excel_path: Path,
    family: str | None,
    source_workbook_path: Path | None = None,
    preferred_file_names: list[str] | None = None,
    export_profile_id: str | None = None,
    address_profile_id: str | None = None,
) -> dict[str, Any]:
    raw_generated_profile = profile_workbook(generated_excel_path, family=None)
    compare_key = infer_compare_key_from_workbooks(
        family=family,
        generated_profile=raw_generated_profile,
        export_profile_id=export_profile_id,
        address_profile_id=address_profile_id,
    )
    source_workbook_path = source_workbook_path or pick_representative_source_workbook(
        family=compare_key,
        preferred_file_names=preferred_file_names,
    )
    raw_source_profile = profile_workbook_cached(str(source_workbook_path.resolve()), family=None)
    compare_key = infer_compare_key_from_workbooks(
        family=family,
        generated_profile=raw_generated_profile,
        source_profile=raw_source_profile,
        export_profile_id=export_profile_id,
        address_profile_id=address_profile_id,
    )
    source_profile = profile_workbook_cached(str(source_workbook_path.resolve()), family=compare_key)
    generated_profile = profile_workbook(generated_excel_path, family=compare_key)
    expected_order = EXPECTED_CANONICAL_ORDER.get(compare_key, [])

    source_map = {sheet["canonical_name"]: sheet for sheet in source_profile["sheets"]}
    generated_map = {sheet["canonical_name"]: sheet for sheet in generated_profile["sheets"]}

    sheet_results: list[dict[str, Any]] = []
    missing_categories: list[str] = []
    for canonical_name in expected_order:
        source_sheet = source_map.get(canonical_name)
        generated_sheet = generated_map.get(canonical_name)
        if source_sheet is None or generated_sheet is None:
            missing_categories.append(canonical_name)
            sheet_results.append(
                {
                    "canonical_name": canonical_name,
                    "source_sheet": source_sheet.get("sheet_name") if source_sheet else None,
                    "generated_sheet": generated_sheet.get("sheet_name") if generated_sheet else None,
                    "status": "diverged",
                    "verdict": "diverged",
                    "structure_status": "diverged",
                    "format_status": "diverged",
                    "combined_status": "diverged",
                    "overall_score": 0.0,
                    "structure_score": 0.0,
                    "format_score": 0.0,
                    "major_differences": ["source 或 generated 缺少该类工作表"],
                    "priority_fixes": [
                        {
                            "metric": "missing_sheet",
                            "label": "缺少工作表",
                            "similarity": 0.0,
                            "score": 0.0,
                            "impact": 1.0,
                            "note": "先补齐该模板族要求的工作表，再谈版式贴源。",
                        }
                    ],
                    "notes": ["source 或 generated 缺少该类工作表"],
                    "format_breakdown": {},
                }
            )
            continue
        sheet_results.append(compare_sheet_pair(source_sheet, generated_sheet))

    sheet_order_match = source_profile["canonical_sheet_order"] == generated_profile["canonical_sheet_order"]
    structure_statuses = [item.get("structure_status", item.get("status", "diverged")) for item in sheet_results]
    combined_statuses = [item.get("combined_status", item.get("status", "diverged")) for item in sheet_results]

    format_score = round_score(sum(item.get("format_score", 0.0) for item in sheet_results) / max(len(sheet_results), 1))
    structure_score = round_score(sum(item.get("structure_score", 0.0) for item in sheet_results) / max(len(sheet_results), 1))
    overall_score = round_score((structure_score * 0.35) + (format_score * 0.65))

    if missing_categories or not sheet_order_match:
        structure_status = "diverged"
        overall_status = "diverged"
    else:
        structure_status = verdict_from_score(structure_score, match_threshold=92.0, close_threshold=72.0)
        overall_status = verdict_from_score(overall_score, match_threshold=92.0, close_threshold=70.0)

    format_status = verdict_from_score(format_score, match_threshold=93.0, close_threshold=72.0)

    worst_sheet = min(sheet_results, key=lambda item: item.get("overall_score", 0.0), default=None)
    all_priority_fixes: list[dict[str, Any]] = []
    for sheet in sheet_results:
        for fix in sheet.get("priority_fixes", []):
            all_priority_fixes.append(
                {
                    **fix,
                    "sheet": sheet.get("generated_sheet") or sheet.get("source_sheet") or sheet.get("canonical_name"),
                }
            )
    all_priority_fixes = sorted(all_priority_fixes, key=lambda item: item.get("impact", 0.0), reverse=True)

    return {
        "family": family or compare_key,
        "compare_key": compare_key,
        "overall_status": overall_status,
        "verdict": overall_status,
        "structure_status": structure_status,
        "format_status": format_status,
        "overall_score": overall_score,
        "structure_score": structure_score,
        "format_score": format_score,
        "representative_source": {
            "path": source_profile["path"],
            "file_name": source_profile["file_name"],
        },
        "source_workbook": {
            "path": source_profile["path"],
            "file_name": source_profile["file_name"],
            "sheet_order": source_profile["sheet_names"],
            "canonical_sheet_order": source_profile["canonical_sheet_order"],
        },
        "generated_workbook": {
            "path": generated_excel_path.name,
            "file_name": generated_excel_path.name,
            "sheet_order": generated_profile["sheet_names"],
            "canonical_sheet_order": generated_profile["canonical_sheet_order"],
        },
        "sheet_order": {
            "matches": sheet_order_match,
            "source": source_profile["sheet_names"],
            "generated": generated_profile["sheet_names"],
        },
        "sheet_order_match": sheet_order_match,
        "missing_categories": missing_categories,
        "metrics": {
            "sheet_count": len(sheet_results),
            "match_count": sum(1 for item in sheet_results if item["status"] == "match"),
            "close_count": sum(1 for item in sheet_results if item["status"] == "close"),
            "diverged_count": sum(1 for item in sheet_results if item["status"] == "diverged"),
            "structure_score": structure_score,
            "format_score": format_score,
            "overall_score": overall_score,
        },
        "score_summary": {
            "structure": {
                "score": structure_score,
                "status": structure_status,
            },
            "format": {
                "score": format_score,
                "status": format_status,
            },
            "overall": {
                "score": overall_score,
                "status": overall_status,
            },
            "worst_sheet": {
                "canonical_name": worst_sheet.get("canonical_name"),
                "sheet_name": worst_sheet.get("generated_sheet"),
                "overall_score": worst_sheet.get("overall_score"),
                "format_score": worst_sheet.get("format_score"),
            }
            if worst_sheet
            else None,
        },
        "top_priority_fixes": all_priority_fixes[:8],
        "sheets": sheet_results,
        "sheet_results": sheet_results,
        "notes": [
            "已升级为版式级对比：同时考虑表头/intro 文案、合并单元格、列宽、行高、颜色填充与固定锚点文案。",
            "overall_status 现在体现结构 + 版式综合结论；structure_status 与 format_status 可分别查看。",
            "数据行数量仍保留为规模差异指标，不单独代表版式完全失真。",
        ],
    }


def compare_generated_excel_to_source(
    generated_excel_path: Path,
    family: str | None,
    source_workbook_path: Path | None = None,
    preferred_file_names: list[str] | None = None,
    export_profile_id: str | None = None,
    address_profile_id: str | None = None,
) -> dict[str, Any]:
    return compare_workbooks(
        generated_excel_path=Path(generated_excel_path),
        family=family,
        source_workbook_path=source_workbook_path,
        preferred_file_names=preferred_file_names,
        export_profile_id=export_profile_id,
        address_profile_id=address_profile_id,
    )



def compare_run_artifacts(run_dir: Path) -> dict[str, Any]:
    run_dir = Path(run_dir)
    manifest_path = run_dir / "manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"未找到 manifest: {manifest_path}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    family = manifest["family"]
    excel_path = Path(manifest["artifacts"]["excel_path"])
    if not excel_path.is_absolute():
        excel_path = run_dir / excel_path.name
    profile_selection = manifest.get("summary", {}).get("profile_selection", {})
    return compare_workbooks(
        generated_excel_path=excel_path,
        family=family,
        export_profile_id=profile_selection.get("export_profile_id"),
        address_profile_id=profile_selection.get("address_profile_id"),
    )


def compare_run_directory(run_dir: Path) -> dict[str, Any]:
    return compare_run_artifacts(run_dir)



def _latest_run_dirs_by_family() -> dict[str, Path]:
    latest: dict[str, tuple[float, Path]] = {}
    for manifest_path in RUNS_ROOT.glob("*/manifest.json"):
        payload = json.loads(manifest_path.read_text(encoding="utf-8"))
        family = payload.get("family")
        if not family:
            continue
        profile_selection = payload.get("summary", {}).get("profile_selection", {})
        compare_key = classify_compare_key(
            family=family,
            export_profile_id=profile_selection.get("export_profile_id"),
            address_profile_id=profile_selection.get("address_profile_id"),
        )
        mtime = manifest_path.stat().st_mtime
        current = latest.get(compare_key)
        if current is None or mtime > current[0]:
            latest[compare_key] = (mtime, manifest_path.parent)
    return {family: item[1] for family, item in latest.items()}



def render_report_markdown(report: dict[str, Any]) -> str:
    reports = report["reports"] if "reports" in report else [report]
    lines = ["# 源文件版式对比报告", ""]
    for item in reports:
        lines.extend(
            [
                f"## {item['family']}",
                "",
                f"- 综合结论：`{item['overall_status']}` / {item.get('overall_score', '-')}",
                f"- 结构评分：`{item.get('structure_status')}` / {item.get('structure_score', '-')}",
                f"- 版式评分：`{item.get('format_status')}` / {item.get('format_score', '-')}",
                f"- 源文件：`{item['source_workbook']['path']}`",
                f"- 生成文件：`{item['generated_workbook']['path']}`",
                f"- Sheet 顺序一致：`{item['sheet_order_match']}`",
                "",
                "| 类别 | 综合 | 结构分 | 版式分 | 源 Sheet | 生成 Sheet | 主要差异 |",
                "|---|---|---:|---:|---|---|---|",
            ]
        )
        for sheet in item["sheet_results"]:
            lines.append(
                "| {canonical} | {status} | {structure_score} | {format_score} | {source_sheet} | {generated_sheet} | {diff} |".format(
                    canonical=sheet.get("canonical_name", "-"),
                    status=sheet.get("status", "-"),
                    structure_score=sheet.get("structure_score", "-"),
                    format_score=sheet.get("format_score", "-"),
                    source_sheet=sheet.get("source_sheet") or "-",
                    generated_sheet=sheet.get("generated_sheet") or "-",
                    diff="；".join(sheet.get("major_differences", [])[:2]) or "-",
                )
            )
        if item.get("top_priority_fixes"):
            lines.append("")
            lines.append("### Top Priority Fixes")
            for fix in item["top_priority_fixes"][:5]:
                lines.append(
                    f"- `{fix.get('sheet')}` · {fix.get('label')} · score={fix.get('score')} · {fix.get('note')}"
                )
        lines.append("")
    return "\n".join(lines).strip() + "\n"


def render_compare_markdown(report: dict[str, Any]) -> str:
    return render_report_markdown(report if "reports" in report else {"reports": [report]})



def save_report(report: dict[str, Any], json_path: Path, markdown_path: Path) -> None:
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    markdown_path.write_text(render_report_markdown(report), encoding="utf-8")



def main() -> int:
    parser = argparse.ArgumentParser(description="对比生成 Excel 与代表源协议 Excel 的结构/版式差异")
    parser.add_argument("--run-dir", help="指定 run 目录")
    parser.add_argument("--run-id", help="指定 run ID（位于 protocol_studio/runs 下）")
    parser.add_argument("--generated-excel", help="指定生成 Excel 路径")
    parser.add_argument("--family", choices=sorted(EXPECTED_CANONICAL_ORDER.keys()))
    parser.add_argument("--export-profile-id", help="当使用 --generated-excel 时显式指定 export_profile_id")
    parser.add_argument("--address-profile-id", help="当使用 --generated-excel 时显式指定 address_profile_id")
    parser.add_argument("--source", help="指定源 Excel 路径")
    parser.add_argument("--report-json", help="JSON 报告输出路径")
    parser.add_argument("--report-md", help="Markdown 报告输出路径")
    args = parser.parse_args()

    reports: list[dict[str, Any]] = []
    if args.run_dir or args.run_id:
        run_dir = Path(args.run_dir) if args.run_dir else RUNS_ROOT / str(args.run_id)
        reports.append(compare_run_artifacts(run_dir))
    elif args.generated_excel:
        reports.append(
            compare_workbooks(
                generated_excel_path=Path(args.generated_excel),
                family=args.family,
                source_workbook_path=Path(args.source) if args.source else None,
                export_profile_id=args.export_profile_id,
                address_profile_id=args.address_profile_id,
            )
        )
    else:
        for family, run_dir in _latest_run_dirs_by_family().items():
            reports.append(compare_run_artifacts(run_dir))

    bundle = {"reports": reports}
    report_json = Path(args.report_json) if args.report_json else Path(__file__).resolve().parent / "source_compare_report.json"
    report_md = Path(args.report_md) if args.report_md else Path(__file__).resolve().parent / "source_compare_report.md"
    save_report(bundle, report_json, report_md)
    print(json.dumps({"report_json": str(report_json.resolve()), "report_md": str(report_md.resolve()), "count": len(reports)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
