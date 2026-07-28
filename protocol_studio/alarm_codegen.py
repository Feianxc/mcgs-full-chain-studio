from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


STATE_NAME_RE = re.compile(r"^(State_[A-Za-z]+)(\d+)$")
BIT_LINE_RE = re.compile(r"^\s*Bit(?P<bit>\d+)\s*[：:]\s*(?P<detail>.+?)\s*$")
PHASE_RE = re.compile(r"([ABCN])相")
START_BOX_RE = re.compile(r"(?P<route>[AB])路始端箱S(?P<num>\d+)")
PLUG_BOX_RE = re.compile(r"(?:(?P<route>[AB])路)?插接箱(?P<box>C?\d+)(?:-?分路(?P<branch>\d+))?")
OUTPUT_BRANCH_RE = re.compile(r"(?P<route>[AB])路输出分路(?P<branch>\d+)")
EXTENDED_MODULE_RE = re.compile(r"(?P<route>[AB])路(?P<module>\d+)#监控模块(?:分路(?P<branch>\d+))?")
MODULE_LOCAL_BRANCH_RE = re.compile(r"模块内分路(?P<branch>\d+)")
SCREEN_COLUMN_RE = re.compile(r"第(?P<column>一|二|三|四|\d+)列")
REPEATER_RE = re.compile(r"(?P<route>[AB])路(?:(?:Z)?(?P<num>\d+)中继(?:器|单元)?|中继(?P<num_alt>\d+))")
POWER_RE = re.compile(r"(?P<route>[AB])路电源模块异常")
STATE_C_RE = re.compile(r"^StateC(?P<id>\d+(?:_\d+)?)$")
VAR_NAME_RE = re.compile(r"^(?P<prefix>[A-Za-z_]+?)(?P<id>\d+(?:_\d+)?)$")
CANONICAL_TOPOLOGY_METADATA_PREFIX = "__MCGS_CANONICAL_TOPOLOGY__"

ROUTE_INDEX = {"A": 1, "B": 2}
TEMP_PREFIX_ORDER = [
    "Ta",
    "Tb",
    "Tc",
    "Tn",
    "TaO",
    "TbO",
    "TcO",
    "TnO",
    "TaD",
    "TbD",
    "TcD",
    "TnD",
    "Th",
]
INCOMING_TEMP_PREFIX_ORDER = ["Ta", "Tb", "Tc", "Tn"]
PLUG_BOX_TEMPERATURE_TARGET_KINDS = {"plug_box", "output_branch", "extended_module"}


class AlarmCodegenUnsupportedError(ValueError):
    """Raised when the workbook family is intentionally unsupported."""


@dataclass
class AlarmEntry:
    state_var: str
    state_prefix: str
    state_index: int
    title: str
    bit_text: str


@dataclass
class AlarmTarget:
    kind: str
    route: str
    logical_suffix: str | None = None
    physical_box_id: str | None = None
    branch_index: int | None = None
    module_index: int | None = None
    screen_column: int | None = None
    phase: str | None = None


@dataclass
class WorkbookContext:
    repeater_comm_prefix: str
    device_vars: dict[str, dict[str, dict[str, list[str]]]] = field(
        default_factory=lambda: {
            "start_box": defaultdict(lambda: defaultdict(list)),
            "plug_box": defaultdict(lambda: defaultdict(list)),
            "repeater": defaultdict(lambda: defaultdict(list)),
        }
    )
    combined_plug_map: dict[tuple[str, str, int | None], str] = field(default_factory=dict)
    combined_plug_physical_map: dict[tuple[str, str], list[str]] = field(default_factory=lambda: defaultdict(list))
    output_branch_map: dict[tuple[str, int], str] = field(default_factory=dict)
    extended_module_branch_map: dict[tuple[str, int, int], str] = field(default_factory=dict)
    extended_module_comm_map: dict[tuple[str, int], str] = field(default_factory=dict)
    display_module_branch_map: dict[tuple[str, int, int, int], str] = field(default_factory=dict)
    display_module_comm_map: dict[tuple[str, int, int], str] = field(default_factory=dict)


STATE_RULES: dict[str, dict[str, Any]] = {
    "State_SPD": {"kind": "simple", "base": "ALARM_SPD", "alarm_level": 0, "external": True},
    "State_THD": {"kind": "simple", "base": "THD", "alarm_level": 0},
    "State_In": {"kind": "simple", "base": "In", "alarm_level": 0},
    "State_FH": {"kind": "simple", "base": "F", "alarm_level": 0},
    "State_FL": {"kind": "simple", "base": "F", "alarm_level": 1},
    "State_LoadH": {"kind": "simple", "base": "LoadS", "alarm_level": 0},
    "State_UnbH": {"kind": "simple", "base": "UBS", "alarm_level": 0},
    "State_Power": {"kind": "power", "alarm_level": 0},
    "State_Com": {"kind": "comm", "alarm_level": 0},
    "State_VH": {"kind": "phase", "base_family": "U", "alarm_level": 0},
    "State_VL": {"kind": "phase", "base_family": "U", "alarm_level": 1},
    "State_VLL": {"kind": "phase", "base_family": "U", "alarm_level": 2},
    "State_IH": {"kind": "phase", "base_family": "I", "alarm_level": 0},
    "State_PFL": {"kind": "phase", "base_family": "PF", "alarm_level": 0},
    "State_PH": {"kind": "phase", "base_family": "P", "alarm_level": 0},
    "State_LoadHC": {"kind": "phase", "base_family": "Load", "alarm_level": 0},
    "State_TH": {"kind": "temp", "alarm_level": 0},
}

LEGACY_SLIDE_RAIL_ALARM_LEVELS: dict[str, int] = {
    "State_FH": 1,
    "State_FL": 0,
    "State_VH": 2,
    "State_VL": 1,
    "State_VLL": 0,
}


def alarm_level_for_rule(
    state_prefix: str,
    default_level: int,
    *,
    legacy_slide_rail_order: bool = False,
) -> int:
    if not legacy_slide_rail_order:
        return default_level
    return LEGACY_SLIDE_RAIL_ALARM_LEVELS.get(state_prefix, default_level)


def normalize_title(title: str | None) -> str:
    if not title:
        return ""
    return " ".join(str(title).split())


def extract_bit_lines(text: str | None) -> list[str]:
    if not isinstance(text, str):
        return []
    matched: list[str] = []
    for raw_line in text.splitlines():
        normalized = raw_line.strip()
        if not normalized:
            continue
        if BIT_LINE_RE.match(normalized):
            matched.append(normalized)
    return matched


def normalize_box_id(raw_box_id: str) -> str:
    digits = re.sub(r"\D+", "", str(raw_box_id))
    return digits.lstrip("0") or digits


def normalize_prefix(prefix: str) -> str:
    if prefix.endswith("Z") and prefix.startswith("T"):
        return prefix[:-1]
    return prefix


def parse_canonical_topology_metadata(row_values: list[Any]) -> dict[str, str]:
    """Read renderer-owned hidden metadata without guessing topology from output order."""

    raw_metadata = next(
        (
            value
            for value in row_values
            if isinstance(value, str)
            and value.startswith(f"{CANONICAL_TOPOLOGY_METADATA_PREFIX}|")
        ),
        None,
    )
    if raw_metadata is None:
        return {}

    metadata: dict[str, str] = {}
    for token in raw_metadata.split("|")[1:]:
        key, separator, value = token.partition("=")
        key = key.strip()
        value = value.strip()
        if separator and key and value:
            metadata[key] = value
    return metadata


def canonical_metadata_int(metadata: dict[str, str], key: str) -> int | None:
    value = metadata.get(key)
    if value is None:
        return None
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid canonical topology metadata: {key}={value!r}") from exc
    if parsed < 1:
        raise ValueError(f"Invalid canonical topology metadata: {key} must be >= 1")
    return parsed


def canonical_communication_device_code(metadata: dict[str, str]) -> str | None:
    value = metadata.get("communication_variable_device_code")
    if value is None:
        return None
    normalized = str(value).strip()
    if normalized.startswith("Comm_EC"):
        normalized = normalized[len("Comm_EC") :]
    if not STATE_C_RE.fullmatch(f"StateC{normalized}"):
        raise ValueError(
            "Invalid canonical topology metadata: "
            f"communication_variable_device_code={value!r}"
        )
    return normalized


def judge_name(index: int) -> str:
    return "separate_judge" if index == 0 else f"separate_judge_{index + 1}"


def phase_prefix(base_family: str, phase: str) -> str:
    phase = phase.upper()
    if base_family == "U":
        return {"A": "Ua", "B": "Ub", "C": "Uc"}[phase]
    if base_family == "I":
        return {"A": "Ia", "B": "Ib", "C": "Ic"}[phase]
    if base_family == "P":
        return {"A": "Pa", "B": "Pb", "C": "Pc"}[phase]
    if base_family == "PF":
        return {"A": "PFa", "B": "PFb", "C": "PFc"}[phase]
    if base_family == "Load":
        return {"A": "Loada", "B": "Loadb", "C": "Loadc"}[phase]
    raise ValueError(f"Unsupported phase base family: {base_family}")


def generate_simple_code(var_name: str, alarm_level: int, state_var: str, bit: int) -> str:
    return (
        f"!GetAlmValue({var_name}, {alarm_level}, separate_judge, 16)\n"
        f"IF separate_judge<>1 THEN\n"
        f"    {state_var} = !BitSet({state_var}, {bit})\n"
        f"ELSE\n"
        f"    {state_var} = !BitClear({state_var}, {bit})\n"
        f"ENDIF\n"
    )


def generate_temp_code(temp_vars: list[str], state_var: str, bit: int) -> str:
    judges = [judge_name(index) for index in range(len(temp_vars))]
    lines = [f"!GetAlmValue({var_name}, 0, {judge}, 16)" for var_name, judge in zip(temp_vars, judges, strict=True)]
    condition = " OR ".join(f"{judge}<>1" for judge in judges)
    lines.append(f"IF ({condition}) THEN")
    lines.append(f"    {state_var} = !BitSet({state_var}, {bit})")
    lines.append("ELSE")
    lines.append(f"    {state_var} = !BitClear({state_var}, {bit})")
    lines.append("ENDIF")
    return "\n".join(lines) + "\n"


def is_alarm_sheet(sheet_name: str) -> bool:
    return sheet_name == "报警状态" or sheet_name.endswith("报警")


def is_metadata_text(text: str) -> bool:
    normalized = normalize_title(text)
    if not normalized:
        return True
    return bool(
        normalized.startswith("State_")
        or normalized in {"SINGLE", "DOUBLE", "只读", "只读4DF", "只读4DUB", "只读4WUB", "只读4DWB"}
        or normalized.startswith("[4区]")
        or "位 " in normalized
        or normalized.startswith("Bit")
        or normalized in {"A", "V", "Hz", "KW", "%", "℃", "KWH"}
    )


def iter_alarm_entries(context: WorkbookContext, workbook_path: Path) -> list[AlarmEntry]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if {"A路屏报警", "B路屏报警"}.issubset(set(workbook.sheetnames)):
        raise AlarmCodegenUnsupportedError("当前版本暂不支持 A/B 分屏报警页的报警代码自动生成。")
    if "报警状态" not in workbook.sheetnames:
        raise AlarmCodegenUnsupportedError("当前工作簿未包含统一的“报警状态”页，无法自动生成报警代码。")

    sheet = workbook["报警状态"]
    entries: list[AlarmEntry] = []
    last_title = ""
    for row in sheet.iter_rows(values_only=True):
        row_values = list(row)
        state_var = next(
            (value for value in row_values if isinstance(value, str) and value.startswith("State_")),
            None,
        )
        if not state_var:
            continue

        match = STATE_NAME_RE.match(state_var)
        if not match:
            raise ValueError(f"Unsupported state variable name: {state_var}")

        bit_candidates: list[tuple[int, int, str]] = []
        for value in row_values:
            if not isinstance(value, str):
                continue
            matched_lines = extract_bit_lines(value)
            if not matched_lines:
                continue
            bit_candidates.append((len(matched_lines), len(value), "\n".join(matched_lines)))

        bit_text = max(bit_candidates, key=lambda item: (item[0], item[1]))[2] if bit_candidates else ""
        if not bit_text:
            raise ValueError(f"Missing bit description for {state_var}")

        title = next(
            (
                str(value)
                for value in row_values
                if isinstance(value, str)
                and value != state_var
                and not extract_bit_lines(value)
                and not is_metadata_text(value)
            ),
            "",
        )
        if title:
            last_title = title
        else:
            title = last_title

        entries.append(
            AlarmEntry(
                state_var=state_var,
                state_prefix=match.group(1),
                state_index=int(match.group(2)),
                title=normalize_title(title),
                bit_text=bit_text,
            )
        )
    return entries


def collect_device_vars(context: WorkbookContext, workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if {"A路屏报警", "B路屏报警"}.issubset(set(workbook.sheetnames)):
        raise AlarmCodegenUnsupportedError("当前版本暂不支持 A/B 分屏报警页的报警代码自动生成。")

    for sheet_name in workbook.sheetnames:
        if is_alarm_sheet(sheet_name):
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_col=2, max_col=2, values_only=True):
            var_name = row[0]
            if not isinstance(var_name, str):
                continue
            match = VAR_NAME_RE.match(var_name)
            if not match:
                continue
            raw_prefix = match.group("prefix")
            logical_id = match.group("id")
            prefix = normalize_prefix(raw_prefix)

            if raw_prefix.endswith("Z") and raw_prefix.startswith("T"):
                device_kind = "repeater"
            elif "_" in logical_id or int(logical_id.split("_", 1)[0]) >= 100:
                device_kind = "plug_box"
            else:
                device_kind = "start_box"

            bucket = context.device_vars[device_kind][logical_id]
            if var_name not in bucket[prefix]:
                bucket[prefix].append(var_name)

    for sheet_name in workbook.sheetnames:
        if is_alarm_sheet(sheet_name) or sheet_name in {"中继器", "中继单元", "连接器测温"}:
            continue
        collect_combined_plug_map(context, workbook[sheet_name])
    if "插接箱" in workbook.sheetnames:
        collect_extended_module_map(context, workbook["插接箱"])


def collect_combined_plug_map(context: WorkbookContext, sheet) -> None:
    current_label = ""
    current_metadata: dict[str, str] = {}
    branch_counters: dict[tuple[str, str], int] = defaultdict(int)
    for row in sheet.iter_rows(values_only=True):
        row_values = list(row)
        row_label = next(
            (
                str(value)
                for value in row_values
                if isinstance(value, str) and ("插接箱" in value or "输出分路" in value)
            ),
            "",
        )
        row_metadata = parse_canonical_topology_metadata(row_values)
        if row_label:
            current_label = row_label
            current_metadata = row_metadata
        elif row_metadata:
            # Canonical unified rows may deliberately keep the user-facing K
            # column free of internal global output numbers.  Do not let a
            # previous block label leak into a metadata-led row.
            current_label = ""
            current_metadata = row_metadata

        state_var = next(
            (value for value in row_values if isinstance(value, str) and STATE_C_RE.match(value)),
            None,
        )
        if not state_var:
            continue

        label = row_label or current_label
        metadata_output_no = canonical_metadata_int(current_metadata, "output_no")
        if not label and metadata_output_no is None:
            continue

        match = STATE_C_RE.match(state_var)
        if not match:
            continue
        logical_suffix = match.group("id")
        output_match = OUTPUT_BRANCH_RE.search(label) if label else None
        metadata_route = current_metadata.get("route")
        if metadata_route is not None and metadata_route not in ROUTE_INDEX:
            raise ValueError(f"Invalid canonical topology metadata: route={metadata_route!r}")
        metadata_screen_column = canonical_metadata_int(
            current_metadata,
            "screen_column",
        )
        if metadata_screen_column is not None and metadata_screen_column < 1:
            raise ValueError(
                "Invalid canonical topology metadata: "
                f"screen_column={metadata_screen_column!r}"
            )

        route: str | None = None
        route_branch_no: int | None = None
        if metadata_output_no is not None:
            if output_match and metadata_output_no != int(output_match.group("branch")):
                raise ValueError(
                    "Canonical topology output mismatch: "
                    f"label={output_match.group('branch')}, metadata={metadata_output_no}"
                )
            if output_match and metadata_route is not None and metadata_route != output_match.group("route"):
                raise ValueError(
                    "Canonical topology route mismatch: "
                    f"label={output_match.group('route')!r}, metadata={metadata_route!r}"
                )
            route = metadata_route or (output_match.group("route") if output_match else None)
            if route is None:
                raise ValueError(
                    "Canonical topology metadata for an output branch is missing route."
                )
            route_branch_no = metadata_output_no
        elif output_match:
            route = output_match.group("route")
            route_branch_no = int(output_match.group("branch"))

        if route is not None and route_branch_no is not None:
            context.output_branch_map[(route, route_branch_no)] = logical_suffix

            module_match = EXTENDED_MODULE_RE.search(label) if label else None
            metadata_module_no = canonical_metadata_int(current_metadata, "module_no")
            metadata_display_module_no = (
                canonical_metadata_int(current_metadata, "display_module_no")
                or metadata_module_no
            )
            label_module_no = int(module_match.group("module")) if module_match else None
            if (
                metadata_module_no is not None
                and label_module_no is not None
                and metadata_module_no != label_module_no
            ):
                raise ValueError(
                    "Canonical topology module mismatch: "
                    f"label={label_module_no}, metadata={metadata_module_no}"
                )
            module_no = metadata_module_no or label_module_no

            local_match = MODULE_LOCAL_BRANCH_RE.search(label)
            metadata_local_branch_no = canonical_metadata_int(
                current_metadata,
                "module_local_branch_no",
            )
            label_local_branch_no = None
            if module_match and module_match.group("branch"):
                label_local_branch_no = int(module_match.group("branch"))
            elif local_match:
                label_local_branch_no = int(local_match.group("branch"))
            if (
                metadata_local_branch_no is not None
                and label_local_branch_no is not None
                and metadata_local_branch_no != label_local_branch_no
            ):
                raise ValueError(
                    "Canonical topology local branch mismatch: "
                    f"label={label_local_branch_no}, metadata={metadata_local_branch_no}"
                )
            local_branch_no = metadata_local_branch_no or label_local_branch_no

            if module_no is not None:
                if local_branch_no is not None:
                    context.extended_module_branch_map[
                        (route, module_no, local_branch_no)
                    ] = logical_suffix
                    if (
                        metadata_screen_column is not None
                        and metadata_display_module_no is not None
                    ):
                        context.display_module_branch_map[
                            (
                                route,
                                metadata_screen_column,
                                metadata_display_module_no,
                                local_branch_no,
                            )
                        ] = logical_suffix
                alarm_slot = (
                    canonical_metadata_int(current_metadata, "communication_alarm_slot")
                    or module_no
                )
                communication_device_code = (
                    canonical_communication_device_code(current_metadata)
                    or logical_suffix
                )
                if "communication_variable_device_code" in current_metadata:
                    context.extended_module_comm_map[(route, alarm_slot)] = communication_device_code
                else:
                    context.extended_module_comm_map.setdefault(
                        (route, alarm_slot),
                        communication_device_code,
                    )
                if (
                    metadata_screen_column is not None
                    and metadata_display_module_no is not None
                ):
                    context.display_module_comm_map[
                        (
                            route,
                            metadata_screen_column,
                            metadata_display_module_no,
                        )
                    ] = communication_device_code
            else:
                if current_metadata:
                    raise ValueError(
                        "Canonical topology metadata for an output branch is missing module_no."
                    )
                # Legacy workbook compatibility only: historical Malaysia sheets
                # encoded exactly two outputs per monitor module and had no parent
                # module metadata.  New canonical workbooks must never enter here.
                legacy_module_no = (route_branch_no - 1) // 2 + 1
                legacy_local_branch_no = (route_branch_no - 1) % 2 + 1
                context.extended_module_branch_map[
                    (route, legacy_module_no, legacy_local_branch_no)
                ] = logical_suffix
                context.extended_module_comm_map.setdefault(
                    (route, legacy_module_no),
                    logical_suffix,
                )
            continue
        plug_match = PLUG_BOX_RE.search(label)
        if not plug_match:
            continue

        route = plug_match.group("route") or infer_route_from_logical_suffix(logical_suffix)
        if not route:
            continue

        physical_box_id = normalize_box_id(plug_match.group("box"))
        branch_key = (route, physical_box_id)
        explicit_branch_index = int(plug_match.group("branch")) if plug_match.group("branch") else None
        if explicit_branch_index is None:
            branch_counters[branch_key] += 1
            branch_index = branch_counters[branch_key]
        else:
            branch_index = explicit_branch_index
            branch_counters[branch_key] = max(branch_counters[branch_key], branch_index)
        context.combined_plug_map[(route, physical_box_id, branch_index)] = logical_suffix
        context.combined_plug_map.setdefault((route, physical_box_id, None), logical_suffix)
        if logical_suffix not in context.combined_plug_physical_map[(route, physical_box_id)]:
            context.combined_plug_physical_map[(route, physical_box_id)].append(logical_suffix)


def collect_extended_module_map(context: WorkbookContext, sheet) -> None:
    current_module: tuple[str, int] | None = None
    current_module_is_explicit = False
    current_branch_count = 0
    current_metadata: dict[str, str] = {}

    for row in sheet.iter_rows(values_only=True):
        row_values = list(row)
        module_label = next(
            (str(value) for value in row_values if isinstance(value, str) and "#监控模块" in value),
            "",
        )
        output_label = next(
            (str(value) for value in row_values if isinstance(value, str) and "#输出支路" in value),
            "",
        )
        row_metadata = parse_canonical_topology_metadata(row_values)
        if module_label or output_label:
            current_metadata = row_metadata
        elif row_metadata:
            current_metadata = row_metadata

        state_var = next(
            (value for value in row_values if isinstance(value, str) and STATE_C_RE.match(value)),
            None,
        )
        if not state_var:
            continue

        state_match = STATE_C_RE.match(state_var)
        if state_match is None:
            continue
        logical_suffix = state_match.group("id")

        module_match = None
        if module_label:
            module_match = EXTENDED_MODULE_RE.search(module_label)
            if module_match:
                current_module = (module_match.group("route"), int(module_match.group("module")))
                current_module_is_explicit = True
                current_branch_count = 0

        metadata_module_no = canonical_metadata_int(current_metadata, "module_no")
        metadata_route = current_metadata.get("route")
        if metadata_route is not None and metadata_route not in ROUTE_INDEX:
            raise ValueError(f"Invalid canonical topology metadata: route={metadata_route!r}")
        if metadata_module_no is not None:
            route = metadata_route or infer_route_from_text(module_label or output_label)
            if route is None:
                route = infer_route_from_logical_suffix(logical_suffix)
            if route is None:
                raise ValueError("Canonical topology metadata is missing route.")
            metadata_module = (route, metadata_module_no)
            if current_module != metadata_module:
                current_branch_count = 0
            current_module = metadata_module
            current_module_is_explicit = True

        output_number = parse_output_index(output_label)
        used_legacy_fallback = False
        if not current_module_is_explicit:
            route = infer_route_from_text(output_label) or infer_route_from_logical_suffix(logical_suffix)
            output_number = parse_output_index(output_label)
            if route is None or output_number is None:
                continue
            # Legacy workbook compatibility only: old split workbooks omitted
            # parent-module metadata and encoded two output rows per module.
            inferred_module_index = (output_number + 1) // 2
            current_module = (route, inferred_module_index)
            current_branch_count = 0
            used_legacy_fallback = True

        metadata_local_branch_no = canonical_metadata_int(
            current_metadata,
            "module_local_branch_no",
        )
        local_match = MODULE_LOCAL_BRANCH_RE.search(module_label or output_label)
        label_local_branch_no = None
        if module_match and module_match.group("branch"):
            label_local_branch_no = int(module_match.group("branch"))
        elif local_match:
            label_local_branch_no = int(local_match.group("branch"))

        if metadata_local_branch_no is not None:
            branch_index = metadata_local_branch_no
        elif label_local_branch_no is not None:
            branch_index = label_local_branch_no
        elif used_legacy_fallback and output_number is not None:
            branch_index = (output_number - 1) % 2 + 1
        else:
            branch_index = current_branch_count + 1
        current_branch_count = max(current_branch_count, branch_index)

        if current_module is None:
            continue
        route, module_index = current_module
        context.extended_module_branch_map[(route, module_index, branch_index)] = logical_suffix
        alarm_slot = (
            canonical_metadata_int(current_metadata, "communication_alarm_slot")
            or module_index
        )
        communication_device_code = (
            canonical_communication_device_code(current_metadata)
            or logical_suffix.split("_", 1)[0]
        )
        if "communication_variable_device_code" in current_metadata:
            context.extended_module_comm_map[(route, alarm_slot)] = communication_device_code
        else:
            context.extended_module_comm_map.setdefault(
                (route, alarm_slot),
                communication_device_code,
            )


def parse_output_index(output_label: str) -> int | None:
    match = re.search(r"(\d+)#输出支路", output_label or "")
    return int(match.group(1)) if match else None


def infer_route_from_text(text: str) -> str | None:
    match = re.search(r"([AB])路", text or "")
    return match.group(1) if match else None


def infer_route_from_logical_suffix(logical_suffix: str) -> str | None:
    try:
        numeric_value = int(str(logical_suffix).split("_", 1)[0])
    except ValueError:
        return None
    if numeric_value >= 200:
        return "B"
    if numeric_value >= 100:
        return "A"
    return None


def parse_screen_column(text: str) -> int | None:
    match = SCREEN_COLUMN_RE.search(text or "")
    if not match:
        return None
    raw_column = match.group("column")
    chinese_columns = {"一": 1, "二": 2, "三": 3, "四": 4}
    return chinese_columns.get(raw_column, int(raw_column) if raw_column.isdigit() else None)


def parse_alarm_target(detail: str) -> AlarmTarget:
    phase_match = PHASE_RE.search(detail)
    phase = phase_match.group(1) if phase_match else None

    power_match = POWER_RE.search(detail)
    if power_match:
        return AlarmTarget(kind="power", route=power_match.group("route"), phase=phase)

    repeater_match = REPEATER_RE.search(detail)
    if repeater_match:
        logical_suffix = repeater_match.group("num") or repeater_match.group("num_alt")
        return AlarmTarget(kind="repeater", route=repeater_match.group("route"), logical_suffix=logical_suffix, phase=phase)

    start_box_match = START_BOX_RE.search(detail)
    if start_box_match:
        return AlarmTarget(
            kind="start_box",
            route=start_box_match.group("route"),
            logical_suffix=start_box_match.group("num"),
            phase=phase,
        )

    output_branch_match = OUTPUT_BRANCH_RE.search(detail)
    if output_branch_match:
        return AlarmTarget(
            kind="output_branch",
            route=output_branch_match.group("route"),
            branch_index=int(output_branch_match.group("branch")),
            phase=phase,
        )

    # New canonical labels intentionally include both the stable output prefix
    # and human-readable parent-module metadata.  Output alarms must resolve by
    # their explicit output number; module-only parsing is reserved for module
    # communication alarms and legacy extended sheets.
    module_match = EXTENDED_MODULE_RE.search(detail)
    if module_match:
        branch_match = MODULE_LOCAL_BRANCH_RE.search(detail)
        raw_branch_index = module_match.group("branch")
        if raw_branch_index is None and branch_match:
            raw_branch_index = branch_match.group("branch")
        branch_index = int(raw_branch_index) if raw_branch_index else None
        return AlarmTarget(
            kind="extended_module" if branch_index else "extended_module_comm",
            route=module_match.group("route"),
            module_index=int(module_match.group("module")),
            screen_column=parse_screen_column(detail),
            branch_index=branch_index,
            phase=phase,
        )

    plug_match = PLUG_BOX_RE.search(detail)
    if plug_match:
        branch_index = int(plug_match.group("branch")) if plug_match.group("branch") else None
        route = plug_match.group("route")
        if route is None:
            route = infer_route_from_logical_suffix(normalize_box_id(plug_match.group("box")))
            if route is None:
                route = "A"
        return AlarmTarget(
            kind="plug_box",
            route=route,
            physical_box_id=normalize_box_id(plug_match.group("box")),
            branch_index=branch_index,
            phase=phase,
        )

    raise ValueError(f"Unsupported bit description: {detail}")


def resolve_plug_logical_suffix(context: WorkbookContext, target: AlarmTarget) -> str:
    if not target.physical_box_id:
        raise ValueError("Missing physical plug-box identifier.")

    exact_key = (target.route, target.physical_box_id, target.branch_index)
    if exact_key in context.combined_plug_map:
        return context.combined_plug_map[exact_key]

    generic_key = (target.route, target.physical_box_id, None)
    if generic_key in context.combined_plug_map:
        return context.combined_plug_map[generic_key]

    suffixes = context.combined_plug_physical_map.get((target.route, target.physical_box_id), [])
    if len(suffixes) == 1:
        return suffixes[0]

    if target.branch_index is None and target.physical_box_id in context.device_vars["plug_box"]:
        return target.physical_box_id
    if target.branch_index is not None and target.physical_box_id in context.device_vars["plug_box"]:
        return target.physical_box_id

    raise ValueError(
        f"Cannot resolve logical plug-box suffix for {target.route} route physical box {target.physical_box_id}"
        f"{'' if target.branch_index is None else f' branch {target.branch_index}'}."
    )


def resolve_target_suffix(context: WorkbookContext, target: AlarmTarget, *, for_comm: bool = False) -> str:
    if target.kind == "start_box":
        if not target.logical_suffix:
            raise ValueError("Missing start-box logical suffix.")
        return target.logical_suffix
    if target.kind == "repeater":
        if not target.logical_suffix:
            raise ValueError("Missing repeater logical suffix.")
        return target.logical_suffix
    if target.kind == "plug_box":
        return resolve_plug_logical_suffix(context, target)
    if target.kind == "output_branch":
        if target.branch_index is None:
            raise ValueError("Missing output branch index.")
        key = (target.route, target.branch_index)
        if key not in context.output_branch_map:
            raise ValueError(
                f"Cannot resolve output branch mapping for {target.route} route branch {target.branch_index}."
            )
        return context.output_branch_map[key]
    if target.kind == "extended_module":
        if target.module_index is None or target.branch_index is None:
            raise ValueError("Missing extended module routing information.")
        if target.screen_column is not None:
            display_key = (
                target.route,
                target.screen_column,
                target.module_index,
                target.branch_index,
            )
            if display_key in context.display_module_branch_map:
                return context.display_module_branch_map[display_key]
        key = (target.route, target.module_index, target.branch_index)
        if key not in context.extended_module_branch_map:
            raise ValueError(f"Cannot resolve branch mapping for {target.route} route module {target.module_index} branch {target.branch_index}.")
        return context.extended_module_branch_map[key]
    if target.kind == "extended_module_comm":
        if not for_comm:
            raise ValueError("Extended module communication target only supports communication mappings.")
        if target.module_index is None:
            raise ValueError("Missing extended module index.")
        if target.screen_column is not None:
            display_key = (
                target.route,
                target.screen_column,
                target.module_index,
            )
            if display_key in context.display_module_comm_map:
                return context.display_module_comm_map[display_key]
        key = (target.route, target.module_index)
        if key not in context.extended_module_comm_map:
            raise ValueError(f"Cannot resolve communication mapping for {target.route} route module {target.module_index}.")
        return context.extended_module_comm_map[key]
    raise ValueError(f"Unsupported target kind: {target.kind}")


def get_bucket(context: WorkbookContext, target: AlarmTarget) -> dict[str, list[str]]:
    logical_suffix = resolve_target_suffix(context, target)
    if target.kind == "start_box":
        return context.device_vars["start_box"].get(logical_suffix, {})
    if target.kind in {"plug_box", "output_branch", "extended_module"}:
        return context.device_vars["plug_box"].get(logical_suffix, {})
    if target.kind == "repeater":
        return context.device_vars["repeater"].get(logical_suffix, {})
    raise ValueError(f"Unsupported bucket target kind: {target.kind}")


def resolve_prefixed_var(context: WorkbookContext, target: AlarmTarget, prefix: str) -> str:
    bucket = get_bucket(context, target)
    candidates = bucket.get(prefix, [])
    if candidates:
        return candidates[0]

    if target.kind == "start_box" and prefix in {"Loada", "Loadb", "Loadc", "Load"}:
        load_candidates = bucket.get("LoadS", [])
        if load_candidates:
            return load_candidates[0]

    logical_suffix = resolve_target_suffix(context, target)
    fallback = f"{prefix}{logical_suffix}"
    if target.kind == "repeater":
        repeater_bucket = context.device_vars["repeater"].get(logical_suffix, {})
        if prefix in repeater_bucket:
            return repeater_bucket[prefix][0]
    if target.kind == "start_box" and prefix == "LoadS":
        return fallback
    raise ValueError(f"Missing variable {fallback} for {target.kind}.")


def resolve_temp_vars(
    context: WorkbookContext,
    target: AlarmTarget,
    *,
    incoming_only: bool = False,
) -> list[str]:
    bucket = get_bucket(context, target)
    if target.phase:
        prefix = {"A": "Ta", "B": "Tb", "C": "Tc", "N": "Tn"}[target.phase]
        return [resolve_prefixed_var(context, target, prefix)]

    ordered: list[str] = []
    seen: set[str] = set()
    prefix_order = INCOMING_TEMP_PREFIX_ORDER if incoming_only else TEMP_PREFIX_ORDER
    for prefix in prefix_order:
        for var_name in bucket.get(prefix, []):
            if var_name not in seen:
                seen.add(var_name)
                ordered.append(var_name)
    if ordered:
        return ordered
    raise ValueError(f"No temperature variables found for {target.kind} target.")


def resolve_simple_var(context: WorkbookContext, target: AlarmTarget, rule: dict[str, Any], state_index: int) -> str:
    base = rule["base"]
    if rule.get("external"):
        logical_suffix = resolve_target_suffix(context, target)
        return f"{base}{logical_suffix}"
    if rule["base"] in {"THD", "In", "F", "LoadS", "UBS"}:
        return resolve_prefixed_var(context, target, base)
    if rule["kind"] == "power":
        return f"Power{state_index}"
    return f"{base}{resolve_target_suffix(context, target)}"


def resolve_phase_var(context: WorkbookContext, target: AlarmTarget, rule: dict[str, Any]) -> str:
    if not target.phase:
        raise ValueError(f"Phase missing for {target.kind}.")
    prefix = phase_prefix(rule["base_family"], target.phase)
    return resolve_prefixed_var(context, target, prefix)


def resolve_comm_var(context: WorkbookContext, target: AlarmTarget) -> str:
    if target.kind == "start_box":
        return f"Comm_ES{resolve_target_suffix(context, target)}"
    if target.kind in {"plug_box", "output_branch"}:
        return f"Comm_EC{resolve_target_suffix(context, target)}"
    if target.kind == "extended_module_comm":
        return f"Comm_EC{resolve_target_suffix(context, target, for_comm=True)}"
    if target.kind == "repeater":
        return f"{context.repeater_comm_prefix}{resolve_target_suffix(context, target)}"
    raise ValueError(f"Unsupported communication target kind: {target.kind}")


def build_code_for_bit(
    context: WorkbookContext,
    entry: AlarmEntry,
    rule: dict[str, Any],
    bit: int,
    detail: str,
    *,
    legacy_slide_rail_order: bool = False,
) -> str:
    target = parse_alarm_target(detail)
    alarm_level = alarm_level_for_rule(
        entry.state_prefix,
        rule["alarm_level"],
        legacy_slide_rail_order=legacy_slide_rail_order,
    )
    if rule["kind"] == "simple":
        var_name = resolve_simple_var(context, target, rule, entry.state_index)
        return generate_simple_code(var_name, alarm_level, entry.state_var, bit)
    if rule["kind"] == "power":
        var_name = f"Power{entry.state_index}"
        return generate_simple_code(var_name, alarm_level, entry.state_var, bit)
    if rule["kind"] == "phase":
        var_name = resolve_phase_var(context, target, rule)
        return generate_simple_code(var_name, alarm_level, entry.state_var, bit)
    if rule["kind"] == "comm":
        var_name = resolve_comm_var(context, target)
        return generate_simple_code(var_name, alarm_level, entry.state_var, bit)
    if rule["kind"] == "temp":
        temp_vars = resolve_temp_vars(
            context,
            target,
            incoming_only=(
                legacy_slide_rail_order
                and target.kind in PLUG_BOX_TEMPERATURE_TARGET_KINDS
            ),
        )
        return generate_temp_code(temp_vars, entry.state_var, bit)
    raise ValueError(f"Unsupported rule kind: {rule['kind']}")


def generate_alarm_code_from_workbook(
    workbook_path: Path,
    repeater_comm_prefix: str = "Comm_EZ",
    *,
    legacy_slide_rail_order: bool = False,
) -> str:
    workbook_path = Path(workbook_path).resolve()
    context = WorkbookContext(repeater_comm_prefix=repeater_comm_prefix)
    collect_device_vars(context, workbook_path)
    entries = iter_alarm_entries(context, workbook_path)

    output_lines = [
        f"' Generated from Excel workbook: {workbook_path.name}",
        f"' Repeater comm prefix default: {repeater_comm_prefix}",
        "' Workbook-first mode: state rows and bit descriptions come directly from the generated workbook.",
        "",
    ]

    for entry in entries:
        rule = STATE_RULES.get(entry.state_prefix)
        if rule is None:
            raise ValueError(f"No generation rule defined for {entry.state_prefix}")

        title = entry.title or entry.state_var
        output_lines.append(f"' === {entry.state_var} | {title} ===")
        parsed_bits = [BIT_LINE_RE.match(raw_line) for raw_line in entry.bit_text.splitlines()]
        matched_any = False
        for matched in parsed_bits:
            if not matched:
                continue
            matched_any = True
            bit = int(matched.group("bit"))
            detail = matched.group("detail")
            output_lines.append(
                build_code_for_bit(
                    context,
                    entry,
                    rule,
                    bit,
                    detail,
                    legacy_slide_rail_order=legacy_slide_rail_order,
                ).rstrip()
            )
            output_lines.append("")

        if not matched_any:
            raise ValueError(f"No valid bit rows found for {entry.state_var}")

    return "\n".join(output_lines).rstrip() + "\n"
