from __future__ import annotations

import csv
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_DEVICE_NAME = "upload"
DEFAULT_DRIVER_LIBRARY_PATH = (
    r"Program\Drivers\通用设备\modbus\modbuscommslave\modbuscommslave_str.ui"
)
DRIVER_LIBRARY_PATH_ENV = "PROTOCOL_STUDIO_MCGS_DRIVER_LIBRARY_PATH"
DEFAULT_DRIVER_COMPONENT_NAME = "ModbusRTU上传"
DEFAULT_DRIVER_COMPONENT_VERSION = "7.105"
DEFAULT_ENCODING = "gb18030"

UPLOAD_HEADERS = [
    "通道号",
    "变量名",
    "变量类型",
    "通道名称",
    "读写类型",
    "寄存器名称",
    "数据类型",
    "寄存器地址",
    "地址偏移",
    "通道采集频次",
    "通道处理",
]


def resolve_driver_library_path(value: str | None = None) -> str:
    """Resolve the MCGS driver reference without embedding a developer path.

    The portable default is relative to the MCGS Pro installation root.  A
    deployment that requires an absolute path can provide it through the
    generated project config or ``PROTOCOL_STUDIO_MCGS_DRIVER_LIBRARY_PATH``.
    """

    configured = str(value or "").strip()
    if configured:
        return configured
    environment_value = os.environ.get(DRIVER_LIBRARY_PATH_ENV, "").strip()
    return environment_value or DEFAULT_DRIVER_LIBRARY_PATH

HEADER_ALIASES = {
    "通道号": "channel_no",
    "变量名": "var_name",
    "变量类型": "var_type",
    "通道名称": "channel_name",
    "读写类型": "access_type",
    "寄存器名称": "register_name",
    "数据类型": "data_type",
    "寄存器地址": "register_address",
}


@dataclass(frozen=True)
class ProtocolPointRow:
    sheet_name: str
    row_number: int
    var_name: str
    var_type: str
    channel_name: str
    access_type: str
    register_name: str
    data_type: str
    register_address: int


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def _parse_register_address(value: Any) -> int | None:
    text = _clean_cell(value)
    if not text:
        return None
    if re.fullmatch(r"[+-]?\d+", text):
        return int(text)
    try:
        number = float(text)
    except ValueError:
        return None
    if number.is_integer():
        return int(number)
    return None


def _find_header_row(sheet) -> tuple[int, dict[str, int]] | None:
    max_scan_row = min(sheet.max_row or 0, 120)
    max_scan_col = min(sheet.max_column or 0, 40)
    rows = sheet.iter_rows(
        min_row=1,
        max_row=max_scan_row,
        min_col=1,
        max_col=max_scan_col,
        values_only=True,
    )
    for row_number, raw_values in enumerate(rows, start=1):
        values = [_clean_cell(value) for value in raw_values]
        if "通道号" not in values or "变量名" not in values or "寄存器地址" not in values:
            continue

        positions: dict[str, int] = {}
        for col_index, header in enumerate(values, start=1):
            normalized = HEADER_ALIASES.get(header)
            if normalized and normalized not in positions:
                positions[normalized] = col_index
        if {"var_name", "data_type", "register_address"}.issubset(positions):
            return row_number, positions
    return None


def _value_by_position(sheet, row_number: int, positions: dict[str, int], key: str) -> str:
    column = positions.get(key)
    if not column:
        return ""
    return _clean_cell(sheet.cell(row=row_number, column=column).value)


def _tuple_value_by_position(values: tuple[Any, ...], positions: dict[str, int], key: str) -> Any:
    column = positions.get(key)
    if not column or column > len(values):
        return None
    return values[column - 1]


def _channel_prefix_for(data_type: str, channel_name: str) -> str:
    channel_name = _clean_cell(channel_name)
    data_type = _clean_cell(data_type)
    if channel_name:
        return re.sub(r"\d+$", "", channel_name)
    if "无符号二进制" in data_type:
        return "只读4DUB" if "32位" in data_type else "只读4WUB"
    if "浮点" in data_type or "32位" in data_type:
        return "只读4DF"
    return "只读4DF"


def extract_protocol_point_rows(workbook_path: Path) -> list[ProtocolPointRow]:
    """Read generated protocol workbook rows that should enter MCGS upload CSV.

    The program-upload table is intentionally workbook-first: it follows the
    same visible worksheet order and data rows that the generated Excel exposes.
    This keeps the CSV aligned with future renderer changes without duplicating
    all renderer-specific row-building logic.
    """

    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        rows: list[ProtocolPointRow] = []
        for sheet in workbook.worksheets:
            header = _find_header_row(sheet)
            if header is None:
                continue
            header_row, positions = header
            data_rows = sheet.iter_rows(
                min_row=header_row + 1,
                max_row=sheet.max_row or header_row,
                values_only=True,
            )
            for row_number, values in enumerate(data_rows, start=header_row + 1):
                var_name = _clean_cell(_tuple_value_by_position(values, positions, "var_name"))
                register_address = _parse_register_address(
                    _tuple_value_by_position(values, positions, "register_address")
                )
                if not var_name or register_address is None:
                    continue
                data_type = _clean_cell(_tuple_value_by_position(values, positions, "data_type"))
                rows.append(
                    ProtocolPointRow(
                        sheet_name=sheet.title,
                        row_number=row_number,
                        var_name=var_name,
                        var_type=_clean_cell(_tuple_value_by_position(values, positions, "var_type")) or "SINGLE",
                        channel_name=_clean_cell(_tuple_value_by_position(values, positions, "channel_name")),
                        access_type=_clean_cell(_tuple_value_by_position(values, positions, "access_type")) or "只读",
                        register_name=_clean_cell(_tuple_value_by_position(values, positions, "register_name")) or "[4区]输出寄存器",
                        data_type=data_type,
                        register_address=register_address,
                    )
                )
        return rows
    finally:
        workbook.close()


def upload_row_from_protocol_row(index: int, row: ProtocolPointRow) -> list[str]:
    upload_address = row.register_address + 1
    channel_prefix = _channel_prefix_for(row.data_type, row.channel_name)
    return [
        str(index),
        row.var_name,
        row.var_type,
        f"{channel_prefix}{upload_address}",
        row.access_type,
        row.register_name,
        row.data_type,
        str(upload_address),
        "",
        "1",
        "",
    ]


def build_program_upload_rows(
    workbook_path: Path,
    *,
    device_name: str = DEFAULT_DEVICE_NAME,
    driver_library_path: str | None = None,
    driver_component_name: str = DEFAULT_DRIVER_COMPONENT_NAME,
    driver_component_version: str = DEFAULT_DRIVER_COMPONENT_VERSION,
) -> list[list[str]]:
    protocol_rows = extract_protocol_point_rows(workbook_path)
    if not protocol_rows:
        raise ValueError("未从协议 Excel 中识别到可导出的程序上传点位行")
    return build_program_upload_rows_from_protocol_rows(
        protocol_rows,
        device_name=device_name,
        driver_library_path=driver_library_path,
        driver_component_name=driver_component_name,
        driver_component_version=driver_component_version,
    )


def build_program_upload_rows_from_protocol_rows(
    protocol_rows: list[ProtocolPointRow],
    *,
    device_name: str = DEFAULT_DEVICE_NAME,
    driver_library_path: str | None = None,
    driver_component_name: str = DEFAULT_DRIVER_COMPONENT_NAME,
    driver_component_version: str = DEFAULT_DRIVER_COMPONENT_VERSION,
) -> list[list[str]]:
    if not protocol_rows:
        raise ValueError("未从协议 Excel 中识别到可导出的程序上传点位行")

    effective_driver_library_path = resolve_driver_library_path(driver_library_path)
    rows: list[list[str]] = [
        [f"组态设备名称:{device_name or DEFAULT_DEVICE_NAME}"],
        [f"驱动库文件路径:{effective_driver_library_path}"],
        [f"驱动构件名称:{driver_component_name or DEFAULT_DRIVER_COMPONENT_NAME}"],
        [f"驱动构件版本:{driver_component_version or DEFAULT_DRIVER_COMPONENT_VERSION}"],
        UPLOAD_HEADERS,
    ]
    rows.extend(upload_row_from_protocol_row(index, row) for index, row in enumerate(protocol_rows))
    return rows


def summarize_sheet_boundaries(protocol_rows: list[ProtocolPointRow]) -> list[dict[str, Any]]:
    boundaries: list[dict[str, Any]] = []
    current_sheet: str | None = None
    start_index = 0
    count = 0
    for index, row in enumerate(protocol_rows):
        if current_sheet is None:
            current_sheet = row.sheet_name
            start_index = index
            count = 1
            continue
        if row.sheet_name == current_sheet:
            count += 1
            continue
        boundaries.append(
            {
                "sheet": current_sheet,
                "start_index": start_index,
                "end_index": index - 1,
                "count": count,
            }
        )
        current_sheet = row.sheet_name
        start_index = index
        count = 1
    if current_sheet is not None:
        boundaries.append(
            {
                "sheet": current_sheet,
                "start_index": start_index,
                "end_index": start_index + count - 1,
                "count": count,
            }
        )
    return boundaries


def write_program_upload_csv(
    workbook_path: Path,
    output_path: Path,
    *,
    device_name: str = DEFAULT_DEVICE_NAME,
    driver_library_path: str | None = None,
    driver_component_name: str = DEFAULT_DRIVER_COMPONENT_NAME,
    driver_component_version: str = DEFAULT_DRIVER_COMPONENT_VERSION,
    encoding: str = DEFAULT_ENCODING,
) -> dict[str, Any]:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    protocol_rows = extract_protocol_point_rows(workbook_path)
    if not protocol_rows:
        raise ValueError("未从协议 Excel 中识别到可导出的程序上传点位行")
    rows = build_program_upload_rows_from_protocol_rows(
        protocol_rows,
        device_name=device_name,
        driver_library_path=driver_library_path,
        driver_component_name=driver_component_name,
        driver_component_version=driver_component_version,
    )
    with output_path.open("w", encoding=encoding, newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)

    point_rows = max(0, len(rows) - 5)
    return {
        "status": "generated",
        "message": "程序上传点表已生成",
        "artifact_path": output_path.name,
        "file_name": output_path.name,
        "point_count": point_rows,
        "encoding": encoding,
        "driver_component_name": driver_component_name or DEFAULT_DRIVER_COMPONENT_NAME,
        "address_transform": "excel_register_address_plus_1",
        "sheet_boundaries": summarize_sheet_boundaries(protocol_rows),
    }


def write_program_upload_csv_from_config(workbook_path: Path, output_path: Path, config: dict[str, Any]) -> dict[str, Any]:
    options = config.get("program_upload", {}) if isinstance(config, dict) else {}
    if not isinstance(options, dict):
        options = {}
    return write_program_upload_csv(
        workbook_path,
        output_path,
        device_name=str(options.get("device_name") or DEFAULT_DEVICE_NAME),
        driver_library_path=str(options.get("driver_library_path") or "") or None,
        driver_component_name=str(options.get("driver_component_name") or DEFAULT_DRIVER_COMPONENT_NAME),
        driver_component_version=str(options.get("driver_component_version") or DEFAULT_DRIVER_COMPONENT_VERSION),
        encoding=str(options.get("encoding") or DEFAULT_ENCODING),
    )
